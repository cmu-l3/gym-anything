#!/usr/bin/env python3
"""
Verifier for New Construction Defect Tracker task
"""

import sys
import os
import logging
import tempfile
import re
from datetime import datetime, timedelta

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_defect_tracker(traj, env_info, task_info):
    """
    Verify that the new construction defect tracker spreadsheet was created correctly.

    Checks:
    1. Correct column structure (9 columns with appropriate headers)
    2. All 6 required defect entries present
    3. Formula usage in 'Days Since Reported' column
    4. Header formatting (bold)
    5. Priority color coding (at least some priorities colored)
    6. Summary section with metrics
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/warranty_defects.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_defect_')

    try:
        # Copy and parse the spreadsheet
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"❌ Failed to load spreadsheet: {error}"
            }

        sheet = wb.active
        score = 0
        max_score = 100
        feedback_parts = []

        # Get all data for analysis
        all_data = get_sheet_data(wb, sheet.title, max_rows=50, max_cols=15)
        
        # Convert to lowercase text for searching
        all_text_lower = " ".join([
            str(cell).lower() 
            for row in all_data 
            for cell in row 
            if cell is not None
        ])

        # ============================================================
        # CRITERION 1: Structure Validation (25 points)
        # ============================================================
        structure_score = 0
        
        # Expected headers (flexible matching)
        expected_headers_keywords = [
            ["defect", "id"],
            ["location"],
            ["category", "type"],
            ["description", "issue", "problem"],
            ["noticed", "discovered", "found"],
            ["reported", "submitted"],
            ["response", "status", "builder"],
            ["priority", "severity"],
            ["days", "since", "reported"]
        ]
        
        # Find header row (should be row 1, but check first few rows)
        header_row = None
        header_row_idx = None
        
        for row_idx in range(0, min(5, len(all_data))):
            row_text = " ".join([str(cell).lower() if cell else "" for cell in all_data[row_idx]])
            # Check if this row has header-like content
            if any(keyword in row_text for keywords in expected_headers_keywords for keyword in keywords):
                header_row = all_data[row_idx]
                header_row_idx = row_idx
                break
        
        if header_row is None:
            feedback_parts.append("❌ No header row found")
        else:
            # Check how many expected columns are present
            header_text = " ".join([str(cell).lower() if cell else "" for cell in header_row])
            matches = sum(
                1 for keywords in expected_headers_keywords
                if any(keyword in header_text for keyword in keywords)
            )
            
            if matches >= 8:
                structure_score = 25
                feedback_parts.append(f"✅ Column structure correct ({matches}/9 columns matched)")
            elif matches >= 6:
                structure_score = 18
                feedback_parts.append(f"⚠️ Column structure mostly correct ({matches}/9 columns matched)")
            else:
                structure_score = 10
                feedback_parts.append(f"⚠️ Column structure incomplete ({matches}/9 columns matched)")
        
        score += structure_score

        # ============================================================
        # CRITERION 2: Data Completeness (30 points)
        # ============================================================
        data_score = 0
        
        # Required defect keywords (flexible matching)
        required_defects = [
            {
                "name": "Basement Water Seepage",
                "keywords": ["basement", "water", "seepage", "leak", "moisture"],
                "min_matches": 2
            },
            {
                "name": "HVAC Short-Cycling",
                "keywords": ["hvac", "heat", "cool", "short", "cycle", "cycling", "ac"],
                "min_matches": 2
            },
            {
                "name": "Door Won't Close",
                "keywords": ["door", "close", "shut", "bedroom", "master"],
                "min_matches": 2
            },
            {
                "name": "Faucet Leak",
                "keywords": ["faucet", "leak", "drip", "kitchen", "plumbing"],
                "min_matches": 2
            },
            {
                "name": "Paint Bubbling",
                "keywords": ["paint", "bubble", "peel", "living", "room"],
                "min_matches": 2
            },
            {
                "name": "Light Flickering",
                "keywords": ["light", "flicker", "electric", "fixture", "entry", "front"],
                "min_matches": 2
            }
        ]
        
        defects_found = 0
        missing_defects = []
        
        for defect in required_defects:
            # Count how many keywords are present
            keyword_matches = sum(1 for keyword in defect["keywords"] if keyword in all_text_lower)
            
            if keyword_matches >= defect["min_matches"]:
                defects_found += 1
            else:
                missing_defects.append(defect["name"])
        
        if defects_found >= 6:
            data_score = 30
            feedback_parts.append(f"✅ All 6 required defects found")
        elif defects_found >= 4:
            data_score = 20
            feedback_parts.append(f"⚠️ {defects_found}/6 required defects found (missing: {', '.join(missing_defects[:2])})")
        elif defects_found >= 2:
            data_score = 10
            feedback_parts.append(f"⚠️ Only {defects_found}/6 defects found")
        else:
            feedback_parts.append(f"❌ Only {defects_found}/6 required defects found")
        
        score += data_score

        # ============================================================
        # CRITERION 3: Formula Verification (20 points)
        # ============================================================
        formula_score = 0
        has_formula = False
        
        # Check for formulas in the spreadsheet
        # Need to reload without data_only to see formulas
        try:
            from openpyxl import load_workbook
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            copy_from_env(container_path, temp_file.name)
            
            wb_formulas = load_workbook(temp_file.name, data_only=False)
            sheet_formulas = wb_formulas.active
            
            # Search for TODAY() formulas in likely columns (rightmost columns)
            for row_idx in range(2, min(15, sheet_formulas.max_row + 1)):
                for col_idx in range(6, min(15, sheet_formulas.max_column + 1)):
                    cell = sheet_formulas.cell(row=row_idx, column=col_idx)
                    if cell.data_type == 'f' and cell.value:
                        formula_upper = str(cell.value).upper()
                        if 'TODAY()' in formula_upper or 'TODAY' in formula_upper:
                            has_formula = True
                            break
                if has_formula:
                    break
            
            wb_formulas.close()
            os.unlink(temp_file.name)
            
        except Exception as e:
            logger.warning(f"Could not check formulas: {e}")
        
        if has_formula:
            formula_score = 20
            feedback_parts.append("✅ 'Days Since Reported' uses TODAY() formula")
        else:
            # Check if there are numeric values that could be days since reported
            # This is a fallback if formula detection fails
            has_reasonable_values = False
            for row in all_data[1:10]:  # Skip header, check data rows
                for cell in row:
                    if isinstance(cell, (int, float)) and 0 <= cell <= 365:
                        has_reasonable_values = True
                        break
            
            if has_reasonable_values:
                formula_score = 10
                feedback_parts.append("⚠️ Days values present, but formula not detected (may be hardcoded)")
            else:
                feedback_parts.append("❌ 'Days Since Reported' should use =TODAY()-[Date Reported] formula")
        
        score += formula_score

        # ============================================================
        # CRITERION 4: Header Formatting (10 points)
        # ============================================================
        formatting_score = 0
        
        if header_row_idx is not None:
            try:
                # Check if first cell in header row is bold
                first_header_cell = sheet.cell(row=header_row_idx + 1, column=1)
                if first_header_cell.font and first_header_cell.font.bold:
                    formatting_score += 5
                    feedback_parts.append("✅ Headers are bold")
                else:
                    feedback_parts.append("⚠️ Headers should be bold")
                
                # Check for background fill on header row
                if first_header_cell.fill and first_header_cell.fill.start_color:
                    color_hex = str(first_header_cell.fill.start_color.rgb)
                    # Check if it's grayish (not white FFFFFFFF)
                    if color_hex != 'FFFFFFFF' and color_hex != '00000000':
                        formatting_score += 5
                        feedback_parts.append("✅ Headers have background color")
                    else:
                        feedback_parts.append("⚠️ Headers should have gray background")
                else:
                    feedback_parts.append("⚠️ Headers should have gray background")
                    
            except Exception as e:
                logger.warning(f"Could not check header formatting: {e}")
        
        score += formatting_score

        # ============================================================
        # CRITERION 5: Priority Color Coding (10 points)
        # ============================================================
        priority_color_score = 0
        colored_priorities_found = 0
        
        try:
            # Search for cells containing "High", "Medium", "Low" with color
            for row_idx in range(2, min(15, sheet.max_row + 1)):
                for col_idx in range(1, min(12, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        value_lower = cell.value.lower()
                        if value_lower in ['high', 'medium', 'low']:
                            # Check if cell has font color
                            if cell.font and cell.font.color:
                                color_rgb = str(cell.font.color.rgb) if cell.font.color.rgb else ""
                                # Check if it's not default black
                                if color_rgb not in ['00000000', 'FF000000', '']:
                                    colored_priorities_found += 1
            
            if colored_priorities_found >= 2:
                priority_color_score = 10
                feedback_parts.append(f"✅ Priority values have color formatting ({colored_priorities_found} colored)")
            elif colored_priorities_found >= 1:
                priority_color_score = 5
                feedback_parts.append(f"⚠️ Some priorities colored, but not all ({colored_priorities_found} colored)")
            else:
                feedback_parts.append("⚠️ Priority values should be color-coded (High=red, Medium=orange, Low=green)")
                
        except Exception as e:
            logger.warning(f"Could not check priority colors: {e}")
        
        score += priority_color_score

        # ============================================================
        # CRITERION 6: Summary Section (5 points)
        # ============================================================
        summary_score = 0
        
        # Look for summary keywords and numeric values below the data
        summary_keywords = ["total", "summary", "count", "high priority", "no response"]
        
        # Check rows 9-20 for summary content
        summary_found = False
        for row_idx in range(8, min(20, len(all_data))):
            row_text = " ".join([str(cell).lower() if cell else "" for cell in all_data[row_idx]])
            if any(keyword in row_text for keyword in summary_keywords):
                # Check if there are numeric values in adjacent cells
                if any(isinstance(cell, (int, float)) for cell in all_data[row_idx]):
                    summary_found = True
                    break
        
        if summary_found:
            summary_score = 5
            feedback_parts.append("✅ Summary section present")
        else:
            feedback_parts.append("⚠️ Summary section recommended (total defects, high priority count, no response count)")
        
        score += summary_score

        # ============================================================
        # Final Assessment
        # ============================================================
        passed = score >= 70
        normalized_score = score / 100.0
        
        feedback = " | ".join(feedback_parts)
        
        return {
            "passed": passed,
            "score": normalized_score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"❌ Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)


# Additional helper function for debugging
def print_sheet_info(sheet, num_rows=10):
    """Debug helper to print sheet contents"""
    logger.info(f"Sheet title: {sheet.title}")
    logger.info(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
    
    for row_idx in range(1, min(num_rows + 1, sheet.max_row + 1)):
        row_data = [sheet.cell(row=row_idx, column=col_idx).value 
                   for col_idx in range(1, min(10, sheet.max_column + 1))]
        logger.info(f"Row {row_idx}: {row_data}")