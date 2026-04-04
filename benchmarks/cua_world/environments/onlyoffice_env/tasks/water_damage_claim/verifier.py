#!/usr/bin/env python3
"""
Verifier for Water Damage Insurance Claim task
Validates that a proper insurance claim inventory spreadsheet was created
"""

import sys
import os
import logging
import tempfile
import re
from collections import Counter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def find_header_row(sheet_data, max_search_rows=20):
    """
    Find the row containing column headers.
    Returns (row_index, column_mapping) or (None, None) if not found.
    
    Column mapping is a dict: {'item': col_idx, 'location': col_idx, ...}
    """
    keywords = {
        'item': ['item', 'description', 'product', 'name'],
        'location': ['location', 'room', 'area', 'place'],
        'damage': ['damage', 'condition', 'status', 'level'],
        'date': ['date', 'purchase', 'bought', 'acquired'],
        'value': ['value', 'cost', 'amount', 'price', 'estimate']
    }
    
    for row_idx, row in enumerate(sheet_data[:max_search_rows]):
        if not row or all(cell is None for cell in row):
            continue
        
        # Convert row to lowercase strings
        row_lower = [str(cell).lower() if cell is not None else '' for cell in row]
        
        # Try to match all 5 required columns
        column_mapping = {}
        for col_type, keyword_list in keywords.items():
            for col_idx, cell_value in enumerate(row_lower):
                if any(kw in cell_value for kw in keyword_list):
                    column_mapping[col_type] = col_idx
                    break
        
        # If we found all 5 column types, this is the header row
        if len(column_mapping) == 5:
            return row_idx, column_mapping
    
    return None, None


def is_bold_cell(workbook, sheet_name, row_idx, col_idx):
    """
    Check if a cell is bold formatted.
    Row and col are 0-indexed for array, but we need 1-indexed for openpyxl
    """
    try:
        sheet = workbook[sheet_name]
        cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
        return cell.font and cell.font.bold
    except Exception as e:
        logger.debug(f"Error checking bold formatting: {e}")
        return False


def extract_numeric_value(value):
    """
    Extract numeric value from cell, handling currency symbols and text.
    Returns float or None
    """
    if value is None:
        return None
    
    if isinstance(value, (int, float)):
        return float(value)
    
    # Try to extract number from string (e.g., "$1,200.50" -> 1200.50)
    if isinstance(value, str):
        # Remove currency symbols, commas, spaces
        cleaned = re.sub(r'[$,\s€£¥]', '', value)
        try:
            return float(cleaned)
        except ValueError:
            return None
    
    return None


def find_sum_formula(workbook, sheet_name, value_col_idx, data_row_count, header_row_idx):
    """
    Find a SUM formula that references the value column.
    Returns (found, formula_cell_address, calculated_value, is_bold)
    """
    try:
        sheet = workbook[sheet_name]
        
        # Search for formulas in the area below the data
        search_start_row = header_row_idx + data_row_count + 2  # Start searching after data
        search_end_row = min(search_start_row + 20, sheet.max_row + 1)  # Search next 20 rows
        
        # Also search in the value column and adjacent columns
        search_start_col = max(1, value_col_idx)  # 1-indexed
        search_end_col = min(value_col_idx + 4, sheet.max_column + 1)
        
        for row in range(search_start_row, search_end_row):
            for col in range(search_start_col, search_end_col):
                cell = sheet.cell(row=row, column=col)
                
                # Check if cell has a formula
                if hasattr(cell, 'value') and cell.value is not None:
                    # Try to get formula from data_only=False workbook
                    # Since we loaded with data_only=True, we need to reload
                    # For now, check if the value is a number and reasonable
                    pass
                
                # Check cell formula attribute
                if hasattr(cell, '_value'):
                    formula_str = str(cell._value) if cell._value else ''
                    if 'SUM' in formula_str.upper():
                        is_bold = cell.font and cell.font.bold
                        return True, f"{cell.column_letter}{cell.row}", cell.value, is_bold
        
        # Alternative: Look for numeric values that are likely totals
        # This is a fallback if we can't detect formulas directly
        return False, None, None, False
        
    except Exception as e:
        logger.error(f"Error finding SUM formula: {e}")
        return False, None, None, False


def verify_insurance_claim_spreadsheet(traj, env_info, task_info):
    """
    Verify that the insurance claim spreadsheet was created correctly.
    
    Checks:
    1. File exists and is valid XLSX
    2. Has 5 required column headers (with appropriate keywords)
    3. Headers are bold formatted
    4. At least 6 rows of data items
    5. All 5 columns have values for each item
    6. Estimated Value column contains numeric values
    7. At least 3 distinct locations/rooms mentioned
    8. Mix of damage levels (at least 2 different types)
    9. SUM formula exists and calculates total correctly
    10. Total cell is bold formatted
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/water_damage_claim.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_claim_')

    try:
        # Copy and parse the spreadsheet
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {"passed": False, "score": 0, "feedback": f"Failed to load spreadsheet: {error}"}

        sheet_name = wb.sheetnames[0]
        sheet_data = get_sheet_data(wb, sheet_name, max_rows=100, max_cols=20)

        if not sheet_data or len(sheet_data) < 2:
            return {"passed": False, "score": 0, "feedback": "Spreadsheet is empty or has insufficient data"}

        criteria_passed = 0
        total_criteria = 10
        feedback_parts = []

        # Criterion 1: Find header row with 5 required columns
        header_row_idx, column_mapping = find_header_row(sheet_data)
        
        if header_row_idx is None or len(column_mapping) != 5:
            feedback_parts.append(f"❌ Missing required column headers (found {len(column_mapping) if column_mapping else 0}/5)")
            feedback_parts.append("   Required: Item Description, Location/Room, Damage Level, Purchase Date, Estimated Value")
            return {
                "passed": False,
                "score": 0,
                "feedback": " | ".join(feedback_parts)
            }
        else:
            criteria_passed += 1
            feedback_parts.append(f"✅ Found all 5 required column headers at row {header_row_idx + 1}")

        # Criterion 2: Check if headers are bold
        headers_bold = True
        for col_type, col_idx in column_mapping.items():
            if not is_bold_cell(wb, sheet_name, header_row_idx, col_idx):
                headers_bold = False
                break
        
        if headers_bold:
            criteria_passed += 1
            feedback_parts.append("✅ Header row is bold formatted")
        else:
            feedback_parts.append("❌ Header row is not bold formatted")

        # Extract data rows (everything after header row, excluding empty rows)
        data_rows = []
        for row_idx in range(header_row_idx + 1, len(sheet_data)):
            row = sheet_data[row_idx]
            # Check if row has any non-empty cells in the column positions
            has_data = any(
                row[col_idx] is not None and str(row[col_idx]).strip() != ''
                for col_idx in column_mapping.values()
                if col_idx < len(row)
            )
            if has_data:
                data_rows.append((row_idx, row))

        # Criterion 3: At least 6 data rows
        data_row_count = len(data_rows)
        if data_row_count >= 6:
            criteria_passed += 1
            feedback_parts.append(f"✅ Has {data_row_count} item entries (minimum 6 required)")
        else:
            feedback_parts.append(f"❌ Only {data_row_count} item entries (minimum 6 required)")

        # Criterion 4: All columns filled for each item
        all_columns_filled = True
        for row_idx, row in data_rows:
            for col_type, col_idx in column_mapping.items():
                if col_idx >= len(row) or row[col_idx] is None or str(row[col_idx]).strip() == '':
                    all_columns_filled = False
                    break
            if not all_columns_filled:
                break
        
        if all_columns_filled and data_row_count > 0:
            criteria_passed += 1
            feedback_parts.append("✅ All columns filled for each item")
        else:
            feedback_parts.append("❌ Some items have missing data in required columns")

        # Criterion 5: Estimated Value column contains numeric values
        value_col_idx = column_mapping['value']
        numeric_values = []
        all_numeric = True
        
        for row_idx, row in data_rows:
            if value_col_idx < len(row):
                num_val = extract_numeric_value(row[value_col_idx])
                if num_val is not None:
                    numeric_values.append(num_val)
                else:
                    all_numeric = False
        
        if all_numeric and len(numeric_values) == data_row_count and data_row_count > 0:
            criteria_passed += 1
            feedback_parts.append(f"✅ All Estimated Value entries are numeric (sum: ${sum(numeric_values):.2f})")
        else:
            feedback_parts.append(f"❌ Some Estimated Value entries are not numeric ({len(numeric_values)}/{data_row_count} are valid)")

        # Criterion 6: At least 3 distinct locations
        location_col_idx = column_mapping['location']
        locations = []
        for row_idx, row in data_rows:
            if location_col_idx < len(row) and row[location_col_idx]:
                loc = str(row[location_col_idx]).strip().lower()
                if loc:
                    locations.append(loc)
        
        distinct_locations = len(set(locations))
        if distinct_locations >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ Items from {distinct_locations} distinct locations (minimum 3 required)")
        else:
            feedback_parts.append(f"❌ Items from only {distinct_locations} location(s) (minimum 3 required)")

        # Criterion 7: Mix of damage levels (at least 2 different types)
        damage_col_idx = column_mapping['damage']
        damage_levels = []
        for row_idx, row in data_rows:
            if damage_col_idx < len(row) and row[damage_col_idx]:
                damage = str(row[damage_col_idx]).strip().lower()
                if damage:
                    damage_levels.append(damage)
        
        distinct_damage_types = len(set(damage_levels))
        if distinct_damage_types >= 2:
            criteria_passed += 1
            feedback_parts.append(f"✅ Mix of damage levels ({distinct_damage_types} different types)")
        else:
            feedback_parts.append(f"❌ Only {distinct_damage_types} damage level type (minimum 2 required)")

        # Criterion 8 & 9: Find SUM formula
        # We need to reload workbook without data_only to see formulas
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=temp_dir)
        try:
            copy_from_env(container_path, temp_file.name)
            
            # Load with data_only=False to see formulas
            from openpyxl import load_workbook
            wb_formulas = load_workbook(temp_file.name, data_only=False)
            sheet_formulas = wb_formulas[sheet_name]
            
            # Search for SUM formula
            found_formula = False
            formula_correct = False
            formula_bold = False
            expected_sum = sum(numeric_values) if numeric_values else 0
            
            # Search in rows after the data
            search_start_row = header_row_idx + 1 + data_row_count + 1
            search_end_row = min(search_start_row + 20, sheet_formulas.max_row)
            
            for row in range(search_start_row, search_end_row + 1):
                for col in range(1, sheet_formulas.max_column + 1):
                    cell = sheet_formulas.cell(row=row, column=col)
                    
                    # Check if cell contains a formula
                    cell_value = cell.value
                    if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
                        formula_upper = cell_value.upper()
                        if 'SUM' in formula_upper:
                            found_formula = True
                            
                            # Check if it's bold
                            if cell.font and cell.font.bold:
                                formula_bold = True
                            
                            # Get calculated value from data_only workbook
                            calc_value = wb[sheet_name].cell(row=row, column=col).value
                            if calc_value is not None and isinstance(calc_value, (int, float)):
                                if abs(calc_value - expected_sum) <= 1:
                                    formula_correct = True
                            
                            break
                
                if found_formula:
                    break
            
            if found_formula and formula_correct:
                criteria_passed += 1
                feedback_parts.append(f"✅ SUM formula present with correct total (${expected_sum:.2f})")
            elif found_formula:
                feedback_parts.append(f"❌ SUM formula found but result is incorrect (expected ${expected_sum:.2f})")
            else:
                feedback_parts.append("❌ No SUM formula found for total claim amount")
            
            if formula_bold:
                criteria_passed += 1
                feedback_parts.append("✅ Total claim amount cell is bold formatted")
            else:
                if found_formula:
                    feedback_parts.append("❌ Total claim amount cell is not bold formatted")
                else:
                    feedback_parts.append("❌ Total claim amount cell not found or not bold")
            
        except Exception as e:
            logger.error(f"Error checking formulas: {e}")
            feedback_parts.append(f"⚠️ Could not verify SUM formula: {str(e)}")
        finally:
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)

        # Calculate final score
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 75

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)