#!/usr/bin/env python3
"""
Verifier for Book Club Voting Sheet task
"""

import sys
import os
import logging
import tempfile
import re
from difflib import SequenceMatcher

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    parse_xlsx_file,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def fuzzy_match(str1, str2, threshold=0.75):
    """
    Fuzzy string matching to handle minor variations
    
    Args:
        str1: First string
        str2: Second string
        threshold: Similarity threshold (0-1)
    
    Returns:
        True if strings are similar enough
    """
    if str1 is None or str2 is None:
        return False
    
    # Normalize strings
    s1 = str(str1).lower().strip()
    s2 = str(str2).lower().strip()
    
    # Exact match
    if s1 == s2:
        return True
    
    # Fuzzy match using SequenceMatcher
    ratio = SequenceMatcher(None, s1, s2).ratio()
    return ratio >= threshold


def extract_formula(workbook, sheet_name, cell_ref):
    """
    Extract formula from a cell (need to reload workbook without data_only)
    
    Args:
        workbook: Workbook object
        sheet_name: Sheet name
        cell_ref: Cell reference
    
    Returns:
        Formula string or None
    """
    try:
        # Get the sheet
        if isinstance(sheet_name, int):
            sheet = workbook.worksheets[sheet_name]
        else:
            sheet = workbook[sheet_name]
        
        # Get the cell
        cell = sheet[cell_ref]
        
        # Check if cell has a formula
        if cell.data_type == 'f':  # Formula
            return cell.value
        elif hasattr(cell, 'formula') and cell.formula:
            return cell.formula
        else:
            # Try to access the value and see if it looks like a formula
            val = cell.value
            if val and isinstance(val, str) and val.startswith('='):
                return val
        
        return None
    except Exception as e:
        logger.debug(f"Could not extract formula from {cell_ref}: {e}")
        return None


def is_bold(workbook, sheet_name, cell_ref):
    """
    Check if a cell's text is bold
    
    Args:
        workbook: Workbook object
        sheet_name: Sheet name
        cell_ref: Cell reference
    
    Returns:
        True if cell text is bold
    """
    try:
        if isinstance(sheet_name, int):
            sheet = workbook.worksheets[sheet_name]
        else:
            sheet = workbook[sheet_name]
        
        cell = sheet[cell_ref]
        if cell.font and cell.font.bold:
            return True
        return False
    except Exception as e:
        logger.debug(f"Could not check bold for {cell_ref}: {e}")
        return False


def copy_file_from_env(container_path, copy_from_env_fn):
    """
    Copy file from container environment to temporary location
    
    Args:
        container_path: Path in container
        copy_from_env_fn: Function to copy file
    
    Returns:
        Tuple of (temp_file_path, success)
    """
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_path = temp_file.name
    temp_file.close()
    
    try:
        copy_from_env_fn(container_path, temp_path)
        
        if not os.path.exists(temp_path) or os.path.getsize(temp_path) == 0:
            return None, False
        
        return temp_path, True
    except Exception as e:
        logger.error(f"Error copying file: {e}")
        if os.path.exists(temp_path):
            os.unlink(temp_path)
        return None, False


def verify_book_club_voting_sheet(traj, env_info, task_info):
    """
    Verify that book club voting spreadsheet was created correctly.

    Checks:
    1. File exists and is valid XLSX
    2. Has correct 5 book titles in column A (fuzzy match)
    3. Has correct 5 author names in column B (fuzzy match)
    4. Has 8 member rating columns (C-J or similar)
    5. Has average rating column with AVERAGE formulas
    6. Header row is bold
    7. Structure is usable (proper layout)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/BookClubVote.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_bookclub_')
    temp_file_path = None

    try:
        # Copy file from container
        temp_file_path, success = copy_file_from_env(container_path, copy_from_env)
        
        if not success:
            return {
                "passed": False, 
                "score": 0, 
                "feedback": f"Failed to load spreadsheet: File not found or empty at {container_path}"
            }

        # Parse the spreadsheet (data_only=True to get calculated values)
        wb_data = parse_xlsx_file(temp_file_path)
        if wb_data is None:
            return {"passed": False, "score": 0, "feedback": "Failed to parse XLSX file"}

        # Also load without data_only to check formulas
        from openpyxl import load_workbook
        try:
            wb_formula = load_workbook(temp_file_path, data_only=False)
        except:
            wb_formula = wb_data  # Fallback

        # Get the first sheet
        sheet_name = wb_data.sheetnames[0]
        
        criteria_passed = 0
        max_criteria = 7
        feedback_parts = []

        # Expected book data
        expected_books = [
            ("The Midnight Library", "Matt Haig"),
            ("Lessons in Chemistry", "Bonnie Garmus"),
            ("Tomorrow, and Tomorrow, and Tomorrow", "Gabrielle Zevin"),
            ("The Lincoln Highway", "Amor Towles"),
            ("Demon Copperhead", "Barbara Kingsolver")
        ]

        # Criterion 1: Check if 5 book titles are present in column A (rows 2-6)
        books_found = 0
        book_rows = []
        
        for i in range(2, 7):  # Rows 2-6
            cell_val = get_cell_value(wb_data, sheet_name, f'A{i}')
            if cell_val:
                book_rows.append(i)
                # Check if this matches any expected book title
                for title, author in expected_books:
                    if fuzzy_match(cell_val, title, threshold=0.7):
                        books_found += 1
                        break
        
        if books_found >= 4:  # Allow 4/5 to be lenient
            criteria_passed += 1
            feedback_parts.append(f"✅ Book titles present: {books_found}/5 found")
        else:
            feedback_parts.append(f"❌ Book titles missing: only {books_found}/5 found")

        # Criterion 2: Check if 5 author names are present in column B
        authors_found = 0
        
        for i in range(2, 7):  # Rows 2-6
            cell_val = get_cell_value(wb_data, sheet_name, f'B{i}')
            if cell_val:
                # Check if this matches any expected author
                for title, author in expected_books:
                    if fuzzy_match(cell_val, author, threshold=0.7):
                        authors_found += 1
                        break
        
        if authors_found >= 4:  # Allow 4/5 to be lenient
            criteria_passed += 1
            feedback_parts.append(f"✅ Author names present: {authors_found}/5 found")
        else:
            feedback_parts.append(f"❌ Author names missing: only {authors_found}/5 found")

        # Criterion 3: Check for 8 member rating columns (C-J or similar)
        # Count non-empty headers in row 1, columns C onwards
        member_columns = []
        for col_idx in range(3, 15):  # Check columns C through N (generous range)
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            header_val = get_cell_value(wb_data, sheet_name, f'{col_letter}1')
            if header_val and str(header_val).strip():
                # Could be "Member 1", "Member_1", "Sarah", etc.
                member_columns.append(col_letter)
        
        if len(member_columns) >= 8:
            criteria_passed += 1
            feedback_parts.append(f"✅ Member rating columns present: {len(member_columns)} columns")
        else:
            feedback_parts.append(f"❌ Member rating columns insufficient: {len(member_columns)}/8 found")

        # Criterion 4: Check for Average Rating column
        # Look for a column header containing "average" or "avg"
        avg_column = None
        for col_idx in range(11, 20):  # Check columns K onwards
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx)
            header_val = get_cell_value(wb_data, sheet_name, f'{col_letter}1')
            if header_val:
                header_str = str(header_val).lower()
                if 'average' in header_str or 'avg' in header_str:
                    avg_column = col_letter
                    break
        
        if avg_column:
            criteria_passed += 1
            feedback_parts.append(f"✅ Average Rating column found: Column {avg_column}")
        else:
            feedback_parts.append("❌ Average Rating column not found")

        # Criterion 5: Check for AVERAGE formulas in the average column
        formulas_found = 0
        if avg_column:
            for i in range(2, 7):  # Rows 2-6
                cell_ref = f'{avg_column}{i}'
                
                # Try to extract formula
                formula = extract_formula(wb_formula, sheet_name, cell_ref)
                
                if formula:
                    formula_upper = formula.upper()
                    if 'AVERAGE' in formula_upper:
                        formulas_found += 1
                        logger.info(f"Found AVERAGE formula in {cell_ref}: {formula}")
        
        if formulas_found >= 4:  # Allow 4/5 to be lenient
            criteria_passed += 1
            feedback_parts.append(f"✅ AVERAGE formulas present: {formulas_found}/5 found")
        else:
            feedback_parts.append(f"❌ AVERAGE formulas missing: only {formulas_found}/5 found")

        # Criterion 6: Check if header row is bold
        # Check a few header cells for bold formatting
        bold_count = 0
        headers_to_check = ['A1', 'B1', 'C1', 'D1']
        for cell_ref in headers_to_check:
            if is_bold(wb_data, sheet_name, cell_ref):
                bold_count += 1
        
        if bold_count >= 2:  # At least half are bold
            criteria_passed += 1
            feedback_parts.append(f"✅ Header row formatting: {bold_count}/{len(headers_to_check)} headers are bold")
        else:
            feedback_parts.append(f"❌ Header row not bold: only {bold_count}/{len(headers_to_check)} headers are bold")

        # Criterion 7: Overall structure check
        # Verify that data exists in the expected layout
        structure_ok = (
            books_found >= 3 and 
            authors_found >= 3 and 
            len(member_columns) >= 6
        )
        
        if structure_ok:
            criteria_passed += 1
            feedback_parts.append("✅ Overall structure is correct and usable")
        else:
            feedback_parts.append("❌ Overall structure is incomplete or incorrect")

        # Calculate score
        score = int((criteria_passed / max_criteria) * 100)
        passed = score >= 70  # Need 70% to pass (5/7 criteria)

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        # Cleanup
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except:
                pass
        cleanup_temp_dir(temp_dir)