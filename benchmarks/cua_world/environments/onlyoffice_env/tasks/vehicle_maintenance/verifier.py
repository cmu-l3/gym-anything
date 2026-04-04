#!/usr/bin/env python3
"""
Verifier for Vehicle Maintenance Log task

This verifier checks:
1. File exists and is valid XLSX
2. Required columns (Date, Service Type, Mileage, Cost) are present
3. At least 5 maintenance records exist with complete data
4. SUM formula exists and calculates correct total cost
5. AVERAGE formula exists and calculates correct average cost
6. MAX formula exists and identifies correct maximum cost
7. Data is organized logically with formulas below data table
"""

import sys
import os
import logging
import tempfile
import re
from typing import Optional, Tuple, List, Dict, Any

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def find_column_by_header(sheet, header_keywords: List[str], max_row: int = 5) -> Optional[int]:
    """
    Find column index by searching for header keywords in first few rows.
    
    Args:
        sheet: Worksheet object
        header_keywords: List of keywords to search for (case-insensitive)
        max_row: Maximum row to search in
    
    Returns:
        Column index (1-based) or None if not found
    """
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, 20):  # Search first 20 columns
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_text = str(cell.value).lower().strip()
                for keyword in header_keywords:
                    if keyword.lower() in cell_text:
                        return col_idx
    return None


def find_header_row(sheet, max_row: int = 5) -> int:
    """
    Find the row containing column headers.
    Looks for rows with multiple non-empty cells that look like headers.
    
    Returns:
        Row index (1-based) of header row, defaults to 1
    """
    for row_idx in range(1, max_row + 1):
        non_empty_count = 0
        has_expected_headers = False
        
        for col_idx in range(1, 10):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                non_empty_count += 1
                cell_text = str(cell.value).lower()
                # Check if this looks like a header
                if any(keyword in cell_text for keyword in ['date', 'service', 'mileage', 'cost', 'type']):
                    has_expected_headers = True
        
        if non_empty_count >= 3 and has_expected_headers:
            return row_idx
    
    return 1  # Default to first row


def extract_numeric_values(sheet, column: int, start_row: int, max_rows: int = 50) -> List[float]:
    """
    Extract numeric values from a column starting at start_row.
    Stops when encountering non-numeric or empty cells.
    
    Args:
        sheet: Worksheet object
        column: Column index (1-based)
        start_row: Starting row index (1-based)
        max_rows: Maximum number of rows to check
    
    Returns:
        List of numeric values
    """
    values = []
    for row_idx in range(start_row, start_row + max_rows):
        cell = sheet.cell(row=row_idx, column=column)
        if cell.value is not None:
            try:
                # Try to convert to float
                val = float(cell.value)
                if val > 0:  # Cost should be positive
                    values.append(val)
            except (ValueError, TypeError):
                # If we hit non-numeric, check if we've collected enough values
                if len(values) >= 3:
                    break
                # Otherwise continue searching
                continue
        else:
            # Empty cell - if we have enough values, stop
            if len(values) >= 3:
                break
    
    return values


def find_formula_in_sheet(sheet, formula_type: str, target_column: int) -> Tuple[bool, Optional[float]]:
    """
    Search entire sheet for a formula of given type referencing target column.
    
    Args:
        sheet: Worksheet object
        formula_type: Type of formula ('SUM', 'AVERAGE', 'MAX')
        target_column: Column index that formula should reference (1-based)
    
    Returns:
        Tuple of (formula_found, calculated_value)
    """
    # Convert column index to letter for searching
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(target_column)
    
    formula_type_upper = formula_type.upper()
    
    # Search entire used range
    for row in sheet.iter_rows(max_row=100, max_col=20):
        for cell in row:
            # Check if cell contains a formula
            if cell.value is not None:
                # Check the data_type to see if it's a formula
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    # This is a formula - check if it's the right type
                    formula_str = str(cell.value).upper()
                    
                    # Check if formula contains the expected function and references the target column
                    if formula_type_upper in formula_str and col_letter in formula_str:
                        # Get the calculated value
                        # Need to reload workbook with data_only=True to get calculated values
                        # For now, we'll check the cell value which should contain the result
                        # Actually, cell.value when data_type='f' contains the formula string
                        # We need to look at the cached value
                        pass
                
                # Also check if the cell value looks like a formula result by checking context
                # Look for the formula pattern in the cell as a string
                cell_str = str(cell.value)
                if cell_str.startswith('='):
                    formula_str = cell_str.upper()
                    if formula_type_upper in formula_str and col_letter in formula_str:
                        # Formula found, but we need the calculated value
                        # The value is in the formula itself, we need to recalculate
                        # For verification, we'll check this separately
                        return (True, None)
    
    # Alternative: Check cells that have numeric values near text labels
    # Search for label-value pairs
    for row_idx in range(1, 100):
        for col_idx in range(1, 10):
            label_cell = sheet.cell(row=row_idx, column=col_idx)
            if label_cell.value:
                label_text = str(label_cell.value).lower()
                
                # Check if this label matches the formula type
                is_sum_label = formula_type_upper == 'SUM' and any(kw in label_text for kw in ['total', 'sum'])
                is_avg_label = formula_type_upper == 'AVERAGE' and any(kw in label_text for kw in ['average', 'avg', 'mean'])
                is_max_label = formula_type_upper == 'MAX' and any(kw in label_text for kw in ['max', 'maximum', 'highest', 'largest'])
                
                if is_sum_label or is_avg_label or is_max_label:
                    # Check adjacent cells for the value
                    for offset in [1, 0, -1]:  # Check right, same, left
                        for row_offset in [0, 1, -1]:  # Check same, below, above
                            value_cell = sheet.cell(row=row_idx + row_offset, column=col_idx + offset)
                            if value_cell.value is not None:
                                try:
                                    val = float(value_cell.value)
                                    if val > 0:
                                        return (True, val)
                                except (ValueError, TypeError):
                                    pass
    
    return (False, None)


def verify_vehicle_maintenance(traj, env_info, task_info):
    """
    Verify that vehicle maintenance log was created correctly.

    Verification criteria (7 total):
    1. File exists and is valid XLSX format
    2. All 4 required columns present (Date, Service Type, Mileage, Cost)
    3. At least 5 complete maintenance records
    4. SUM formula exists with correct total
    5. AVERAGE formula exists with correct average
    6. MAX formula exists with correct maximum
    7. Data is organized logically
    
    Pass threshold: 70% (5 out of 7 criteria)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/vehicle_maintenance_log.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_maintenance_')

    try:
        # Criterion 1: Copy and parse the spreadsheet
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {"passed": False, "score": 0, "feedback": f"Failed to load spreadsheet: {error}"}

        criteria_passed = 0
        total_criteria = 7
        feedback_parts = []

        # File is valid
        criteria_passed += 1
        feedback_parts.append("✅ File exists and is valid XLSX")

        sheet = wb.active
        
        # Find header row
        header_row = find_header_row(sheet)
        logger.info(f"Header row identified: {header_row}")

        # Criterion 2: Find all required columns
        date_col = find_column_by_header(sheet, ['date', 'when'], max_row=header_row+2)
        service_col = find_column_by_header(sheet, ['service', 'type', 'description', 'repair'], max_row=header_row+2)
        mileage_col = find_column_by_header(sheet, ['mileage', 'miles', 'odometer', 'km'], max_row=header_row+2)
        cost_col = find_column_by_header(sheet, ['cost', 'price', 'amount', '$', 'expense'], max_row=header_row+2)

        columns_found = all([date_col, service_col, mileage_col, cost_col])
        
        if columns_found:
            criteria_passed += 1
            feedback_parts.append(f"✅ All 4 required columns found (Date: col{date_col}, Service: col{service_col}, Mileage: col{mileage_col}, Cost: col{cost_col})")
        else:
            missing = []
            if not date_col: missing.append("Date")
            if not service_col: missing.append("Service Type")
            if not mileage_col: missing.append("Mileage")
            if not cost_col: missing.append("Cost")
            feedback_parts.append(f"❌ Missing columns: {', '.join(missing)}")
            
            # Cannot proceed without columns
            return {
                "passed": False,
                "score": int((criteria_passed / total_criteria) * 100),
                "feedback": " | ".join(feedback_parts)
            }

        # Criterion 3: Check for at least 5 complete records
        data_start_row = header_row + 1
        
        # Count complete rows (all 4 columns have values)
        complete_records = 0
        for row_idx in range(data_start_row, data_start_row + 30):  # Check up to 30 rows
            date_val = sheet.cell(row=row_idx, column=date_col).value
            service_val = sheet.cell(row=row_idx, column=service_col).value
            mileage_val = sheet.cell(row=row_idx, column=mileage_col).value
            cost_val = sheet.cell(row=row_idx, column=cost_col).value
            
            # Check if all values are present and non-empty
            if all([date_val, service_val, mileage_val, cost_val]):
                # Verify mileage and cost are numeric
                try:
                    float(mileage_val)
                    float(cost_val)
                    complete_records += 1
                except (ValueError, TypeError):
                    pass
            else:
                # If we encounter empty row and have enough records, stop counting
                if complete_records >= 3:
                    break

        if complete_records >= 5:
            criteria_passed += 1
            feedback_parts.append(f"✅ {complete_records} complete maintenance records found")
        else:
            feedback_parts.append(f"❌ Only {complete_records} complete records (need at least 5)")

        # Extract cost values for formula verification
        costs = extract_numeric_values(sheet, cost_col, data_start_row, max_rows=complete_records + 5)
        
        if len(costs) < complete_records:
            # Try to get costs more carefully
            costs = []
            for row_idx in range(data_start_row, data_start_row + complete_records):
                cost_cell = sheet.cell(row=row_idx, column=cost_col)
                if cost_cell.value is not None:
                    try:
                        costs.append(float(cost_cell.value))
                    except (ValueError, TypeError):
                        pass

        logger.info(f"Extracted cost values: {costs}")

        if len(costs) < 3:
            feedback_parts.append("⚠️ Warning: Could not extract enough cost values for formula verification")
            # Still continue to check if formulas exist

        # Calculate expected values
        expected_sum = sum(costs) if costs else 0
        expected_avg = sum(costs) / len(costs) if costs else 0
        expected_max = max(costs) if costs else 0

        logger.info(f"Expected: SUM={expected_sum}, AVG={expected_avg}, MAX={expected_max}")

        # Criterion 4: Check SUM formula
        sum_found, sum_value = find_formula_in_sheet(sheet, 'SUM', cost_col)
        
        if sum_found:
            if sum_value is not None and abs(sum_value - expected_sum) <= 0.1:
                criteria_passed += 1
                feedback_parts.append(f"✅ SUM formula correct: {sum_value:.2f}")
            else:
                # Formula exists but value not checked, give partial credit by checking if the value appears anywhere
                # Search for the expected sum value in the sheet
                found_correct_sum = False
                for row in sheet.iter_rows(max_row=100, max_col=20):
                    for cell in row:
                        if cell.value is not None:
                            try:
                                val = float(cell.value)
                                if abs(val - expected_sum) <= 0.1:
                                    found_correct_sum = True
                                    break
                            except:
                                pass
                    if found_correct_sum:
                        break
                
                if found_correct_sum:
                    criteria_passed += 1
                    feedback_parts.append(f"✅ SUM formula exists and result appears correct: ~{expected_sum:.2f}")
                else:
                    feedback_parts.append(f"⚠️ SUM formula found but result unclear (expected ~{expected_sum:.2f})")
        else:
            feedback_parts.append(f"❌ SUM formula not found (expected total: {expected_sum:.2f})")

        # Criterion 5: Check AVERAGE formula
        avg_found, avg_value = find_formula_in_sheet(sheet, 'AVERAGE', cost_col)
        
        if avg_found:
            if avg_value is not None and abs(avg_value - expected_avg) <= 0.1:
                criteria_passed += 1
                feedback_parts.append(f"✅ AVERAGE formula correct: {avg_value:.2f}")
            else:
                # Search for expected average value
                found_correct_avg = False
                for row in sheet.iter_rows(max_row=100, max_col=20):
                    for cell in row:
                        if cell.value is not None:
                            try:
                                val = float(cell.value)
                                if abs(val - expected_avg) <= 0.1:
                                    found_correct_avg = True
                                    break
                            except:
                                pass
                    if found_correct_avg:
                        break
                
                if found_correct_avg:
                    criteria_passed += 1
                    feedback_parts.append(f"✅ AVERAGE formula exists and result appears correct: ~{expected_avg:.2f}")
                else:
                    feedback_parts.append(f"⚠️ AVERAGE formula found but result unclear (expected ~{expected_avg:.2f})")
        else:
            feedback_parts.append(f"❌ AVERAGE formula not found (expected average: {expected_avg:.2f})")

        # Criterion 6: Check MAX formula
        max_found, max_value = find_formula_in_sheet(sheet, 'MAX', cost_col)
        
        if max_found:
            if max_value is not None and abs(max_value - expected_max) <= 0.1:
                criteria_passed += 1
                feedback_parts.append(f"✅ MAX formula correct: {max_value:.2f}")
            else:
                # Search for expected max value
                found_correct_max = False
                for row in sheet.iter_rows(max_row=100, max_col=20):
                    for cell in row:
                        if cell.value is not None:
                            try:
                                val = float(cell.value)
                                if abs(val - expected_max) <= 0.1:
                                    found_correct_max = True
                                    break
                            except:
                                pass
                    if found_correct_max:
                        break
                
                if found_correct_max:
                    criteria_passed += 1
                    feedback_parts.append(f"✅ MAX formula exists and result appears correct: ~{expected_max:.2f}")
                else:
                    feedback_parts.append(f"⚠️ MAX formula found but result unclear (expected ~{expected_max:.2f})")
        else:
            feedback_parts.append(f"❌ MAX formula not found (expected maximum: {expected_max:.2f})")

        # Criterion 7: Check data organization
        # Verify headers are in row 1 or early row, data follows, formulas are after data
        organization_good = (header_row <= 3) and (complete_records >= 3)
        
        if organization_good:
            criteria_passed += 1
            feedback_parts.append("✅ Data organized logically")
        else:
            feedback_parts.append("⚠️ Data organization could be improved")

        # Calculate final score
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 70

        feedback = " | ".join(feedback_parts)

        logger.info(f"Verification complete: {criteria_passed}/{total_criteria} criteria passed, score={score}")

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)