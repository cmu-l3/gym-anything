#!/usr/bin/env python3
"""
Verifier for Elderly Medication Reconciliation task

This verifier checks that a caregiver has properly organized their elderly 
parent's medication information for a medical appointment, including:
- Data completeness
- Professional formatting
- Safety highlighting
- Summary statistics
"""

import sys
import os
import logging
import tempfile
from datetime import datetime, timedelta

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    parse_xlsx_file,
    get_sheet_data,
    get_cell_value,
    count_filled_cells,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_medication_reconciliation(traj, env_info, task_info):
    """
    Verify medication reconciliation spreadsheet for elderly parent.
    
    Scoring breakdown (100 points):
    - Data Completeness: 30 points
    - Professional Formatting: 20 points
    - Safety Features: 25 points
    - Summary Statistics: 25 points
    
    Pass threshold: 60 points
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Try primary expected path, then fallback to messy filename
    primary_path = "/home/ga/Documents/Spreadsheets/dad_medications_reconciled.xlsx"
    fallback_path = "/home/ga/Documents/Spreadsheets/dad_medications_messy.xlsx"
    
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_medrecon_')
    
    try:
        # Try to copy the file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=temp_dir)
        
        container_path = primary_path
        try:
            copy_from_env(primary_path, temp_file.name)
        except:
            logger.info(f"Primary path not found, trying fallback: {fallback_path}")
            container_path = fallback_path
            try:
                copy_from_env(fallback_path, temp_file.name)
            except Exception as e:
                return {
                    "passed": False, 
                    "score": 0, 
                    "feedback": f"❌ File not found at {primary_path} or {fallback_path}: {str(e)}"
                }
        
        if not os.path.exists(temp_file.name) or os.path.getsize(temp_file.name) == 0:
            return {
                "passed": False,
                "score": 0,
                "feedback": f"❌ File not found or empty: {container_path}"
            }
        
        # Parse the spreadsheet
        wb = parse_xlsx_file(temp_file.name)
        if wb is None:
            return {
                "passed": False,
                "score": 0,
                "feedback": "❌ Could not parse XLSX file"
            }
        
        score = 0
        feedback_parts = []
        
        # Get primary sheet (could be renamed or be in multiple sheets)
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        # ====================================================================
        # CRITERION 1: Data Completeness (30 points)
        # ====================================================================
        
        # Check 1a: Required columns present
        required_keywords = ['med', 'dos', 'freq', 'doctor', 'pharm', 'status']
        header_row = []
        
        for cell in sheet[1]:
            if cell.value:
                header_row.append(str(cell.value).lower())
            else:
                header_row.append("")
        
        cols_found = sum(1 for keyword in required_keywords 
                        if any(keyword in header for header in header_row))
        
        if cols_found < 4:
            feedback_parts.append(f"❌ Missing required columns (found {cols_found}/6 keywords)")
            return {
                "passed": False,
                "score": 0,
                "feedback": " | ".join(feedback_parts)
            }
        else:
            score += 5
            feedback_parts.append(f"✅ Required columns present ({cols_found}/6)")
        
        # Check 1b: Extract all medication data
        data = get_sheet_data(wb, sheet_name, max_rows=50, max_cols=15)
        
        # Count non-empty rows (medications)
        med_rows = []
        for row_idx, row in enumerate(data[1:], start=2):  # Skip header
            if row and any(cell for cell in row if cell):
                # Check if this looks like a medication row (has med name)
                if row[0] and str(row[0]).strip():
                    med_rows.append((row_idx, row))
        
        num_medications = len(med_rows)
        
        if num_medications < 10:
            feedback_parts.append(f"⚠️ Only {num_medications} medications found (expected 10-12)")
            score += 5
        else:
            feedback_parts.append(f"✅ {num_medications} medications documented")
            score += 10
        
        # Check 1c: Data completeness ratio
        # Count how many cells are filled vs total expected
        if med_rows:
            total_cells = len(med_rows) * len(header_row)
            filled_cells = sum(1 for _, row in med_rows for cell in row if cell and str(cell).strip())
            completeness_ratio = filled_cells / total_cells if total_cells > 0 else 0
            
            if completeness_ratio > 0.75:
                score += 15
                feedback_parts.append(f"✅ High data completeness ({int(completeness_ratio*100)}%)")
            elif completeness_ratio > 0.55:
                score += 10
                feedback_parts.append(f"⚠️ Moderate data completeness ({int(completeness_ratio*100)}%)")
            else:
                score += 5
                feedback_parts.append(f"❌ Low data completeness ({int(completeness_ratio*100)}%)")
        
        # ====================================================================
        # CRITERION 2: Professional Formatting (20 points)
        # ====================================================================
        
        # Check 2a: Header row formatting (bold)
        header_cell = sheet['A1']
        is_header_bold = False
        if header_cell.font and header_cell.font.bold:
            is_header_bold = True
            score += 5
            feedback_parts.append("✅ Header row is bold")
        else:
            feedback_parts.append("⚠️ Header row not bold")
        
        # Check 2b: Column widths adjusted (not default 8.43)
        from openpyxl.utils import get_column_letter
        custom_width_count = 0
        for col_idx in range(1, min(12, len(header_row) + 1)):
            col_letter = get_column_letter(col_idx)
            if col_letter in sheet.column_dimensions:
                width = sheet.column_dimensions[col_letter].width
                if width != 8.43 and width > 10:  # Default is 8.43
                    custom_width_count += 1
        
        if custom_width_count >= 3:
            score += 5
            feedback_parts.append(f"✅ Column widths adjusted ({custom_width_count} columns)")
        else:
            feedback_parts.append(f"⚠️ Few column widths adjusted ({custom_width_count} columns)")
        
        # Check 2c: Borders or gridlines
        has_borders = False
        for row in sheet.iter_rows(min_row=1, max_row=min(15, len(med_rows) + 2), max_col=5):
            for cell in row:
                if cell.border and (cell.border.left.style or cell.border.top.style):
                    has_borders = True
                    break
            if has_borders:
                break
        
        if has_borders:
            score += 5
            feedback_parts.append("✅ Professional borders applied")
        else:
            score += 2
            feedback_parts.append("⚠️ No visible borders")
        
        # Check 2d: Organization (active vs discontinued separated)
        # Look for organization indicators
        has_organization = False
        
        # Check if multiple sheets (active/discontinued separation)
        if len(wb.sheetnames) > 1:
            has_organization = True
            score += 5
            feedback_parts.append("✅ Multi-sheet organization (Active/Discontinued)")
        else:
            # Check for section headers or sorting
            status_col_idx = None
            for idx, header in enumerate(header_row):
                if 'status' in header:
                    status_col_idx = idx
                    break
            
            if status_col_idx is not None:
                # Check if medications are grouped by status
                statuses = [row[status_col_idx] for _, row in med_rows 
                          if len(row) > status_col_idx and row[status_col_idx]]
                
                # Simple check: if Active medications come before Discontinued
                active_indices = [i for i, s in enumerate(statuses) 
                                if s and 'active' in str(s).lower()]
                discontinued_indices = [i for i, s in enumerate(statuses) 
                                      if s and ('discontin' in str(s).lower() or 'stop' in str(s).lower())]
                
                if active_indices and discontinued_indices:
                    if max(active_indices) < min(discontinued_indices):
                        has_organization = True
                        score += 5
                        feedback_parts.append("✅ Medications organized by status")
                    else:
                        score += 2
                        feedback_parts.append("⚠️ Medications present but not well organized")
        
        # ====================================================================
        # CRITERION 3: Safety Features (25 points)
        # ====================================================================
        
        # Check 3a: Conditional formatting or highlighting
        # Look for cells with fill colors (highlighting)
        highlighted_cells = 0
        for row in sheet.iter_rows(min_row=2, max_row=min(20, len(med_rows) + 5), max_col=12):
            for cell in row:
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    rgb = cell.fill.start_color.rgb
                    # Check if not white/default (FFFFFFFF or 00000000)
                    if rgb not in ['FFFFFFFF', '00000000', 'FF000000']:
                        highlighted_cells += 1
        
        if highlighted_cells >= 5:
            score += 10
            feedback_parts.append(f"✅ Safety highlighting applied ({highlighted_cells} cells)")
        elif highlighted_cells >= 2:
            score += 5
            feedback_parts.append(f"⚠️ Some highlighting present ({highlighted_cells} cells)")
        else:
            feedback_parts.append("❌ No safety highlighting detected")
        
        # Check 3b: Status column with validation or consistent values
        if status_col_idx is not None:
            unique_statuses = set()
            for _, row in med_rows:
                if len(row) > status_col_idx and row[status_col_idx]:
                    unique_statuses.add(str(row[status_col_idx]).lower().strip())
            
            # Good if statuses are consistent (Active, Discontinued, etc.)
            if len(unique_statuses) <= 4:  # Not too many variations
                score += 5
                feedback_parts.append(f"✅ Consistent status values ({len(unique_statuses)} types)")
            else:
                feedback_parts.append(f"⚠️ Inconsistent status values ({len(unique_statuses)} types)")
        
        # Check 3c: Date tracking for refills
        date_cols = [i for i, h in enumerate(header_row) if 'date' in h or 'refill' in h]
        if date_cols:
            score += 10
            feedback_parts.append("✅ Date tracking for refills implemented")
        else:
            score += 5
            feedback_parts.append("⚠️ Limited date tracking")
        
        # ====================================================================
        # CRITERION 4: Summary Statistics with Formulas (25 points)
        # ====================================================================
        
        # Look for summary section (usually at top or in separate area)
        # Check for keywords and formulas
        has_summary = False
        formula_count = 0
        summary_keywords = ['total', 'count', 'summary', 'active', 'average', 'refill']
        
        # Check first 15 rows and columns A-E for summary section
        for row_idx in range(1, min(25, len(med_rows) + 10)):
            for col_idx in range(1, 6):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Check for summary keywords
                if cell.value and isinstance(cell.value, str):
                    cell_lower = cell.value.lower()
                    if any(keyword in cell_lower for keyword in summary_keywords):
                        has_summary = True
                
                # Check for formulas
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                elif hasattr(cell, 'data_type') and cell.data_type == 'f':
                    formula_count += 1
        
        if has_summary and formula_count >= 3:
            score += 25
            feedback_parts.append(f"✅ Comprehensive summary with {formula_count} formulas")
        elif has_summary and formula_count >= 1:
            score += 15
            feedback_parts.append(f"✅ Summary section with {formula_count} formula(s)")
        elif formula_count >= 1:
            score += 10
            feedback_parts.append(f"⚠️ Limited summary ({formula_count} formula found)")
        elif has_summary:
            score += 5
            feedback_parts.append("⚠️ Summary section without formulas")
        else:
            feedback_parts.append("❌ No summary statistics found")
        
        # ====================================================================
        # FINAL ASSESSMENT
        # ====================================================================
        
        passed = score >= 60
        
        # Add appropriate emoji and context to feedback
        if passed:
            if score >= 85:
                feedback_header = "🌟 EXCELLENT - Professional medical documentation"
            elif score >= 75:
                feedback_header = "✅ GOOD - Ready for medical appointment"
            else:
                feedback_header = "✅ ACCEPTABLE - Meets basic requirements"
        else:
            feedback_header = "❌ NEEDS IMPROVEMENT - Not ready for medical use"
        
        feedback = f"{feedback_header} | Score: {score}/100 | " + " | ".join(feedback_parts)
        
        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }
        
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0,
            "feedback": f"❌ Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)
