#!/usr/bin/env python3
"""
Verifier for Grade Dispute Verification task

This verifier checks that the student created a spreadsheet with:
1. Grade data entered (labs, homework, exams)
2. Formulas used for calculations
3. Weighted average calculated correctly (~84%)
4. Discrepancy with posted grade identified
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    parse_xlsx_file,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_grade_verification(traj, env_info, task_info):
    """
    Verify the grade verification spreadsheet task
    
    Checks:
    1. File exists and is valid XLSX
    2. Grade data is entered (labs, homework, exams)
    3. Formulas are used (not manual calculation)
    4. Weighted average is calculated correctly
    5. Discrepancy with posted grade is identified
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/grade_verification.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_grade_')
    
    try:
        # Copy file from container to temp directory
        temp_file = os.path.join(temp_dir, "grade_verification.xlsx")
        
        try:
            copy_from_env(container_path, temp_file)
        except Exception as e:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"❌ Could not copy file from container: {str(e)}"
            }
        
        if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"❌ File not found or empty: {container_path}"
            }
        
        # Parse the spreadsheet
        wb = parse_xlsx_file(temp_file)
        if wb is None:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": "❌ Could not parse grade_verification.xlsx - file may be corrupted"
            }
        
        score = 0
        max_score = 100
        feedback_parts = []
        
        score += 15  # File exists and is valid
        feedback_parts.append("✅ File created successfully")
        
        # Get first sheet
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        # Extract all numeric and string values
        all_values = []
        all_values_numeric = []
        formulas_found = []
        
        for row in sheet.iter_rows(max_row=100, max_col=20):
            for cell in row:
                if cell.value is not None:
                    all_values.append(cell.value)
                    if isinstance(cell.value, (int, float)):
                        all_values_numeric.append(cell.value)
                # Check for formulas by looking at cell data type
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    formulas_found.append(True)
        
        # Also try to reload without data_only to detect formulas
        try:
            from openpyxl import load_workbook
            wb_formula = load_workbook(temp_file, data_only=False)
            sheet_formula = wb_formula[wb_formula.sheetnames[0]]
            
            for row in sheet_formula.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas_found.append(cell.value)
        except:
            pass  # If this fails, we'll just use the results-based checking
        
        logger.info(f"Found {len(all_values_numeric)} numeric values")
        logger.info(f"Found {len(formulas_found)} formulas")
        
        # CRITERION 1: Data Completeness (20 points)
        data_score = 0
        
        # Count numeric values in reasonable grade range (0-100)
        grade_range_values = [v for v in all_values_numeric if 0 <= v <= 100]
        
        # Should have ~10 labs + 12 homework + 2 exams = 24 grade entries
        # Plus calculated values (averages, weighted grade, discrepancy)
        if len(grade_range_values) >= 20:
            data_score += 15
            feedback_parts.append(f"✅ Sufficient grade data entered ({len(grade_range_values)} values in 0-100 range)")
        elif len(grade_range_values) >= 15:
            data_score += 10
            feedback_parts.append(f"⚠️ Some grade data present ({len(grade_range_values)} values)")
        elif len(grade_range_values) >= 10:
            data_score += 5
            feedback_parts.append(f"⚠️ Minimal grade data ({len(grade_range_values)} values, expected ~24+)")
        else:
            feedback_parts.append(f"❌ Insufficient grade data ({len(grade_range_values)} values)")
        
        # Check for posted grade (78)
        if 78 in all_values_numeric or 78.0 in all_values_numeric:
            data_score += 5
            feedback_parts.append("✅ Posted grade (78%) found")
        elif any(77 <= v <= 79 for v in all_values_numeric):
            data_score += 3
            feedback_parts.append("⚠️ Value close to posted grade found")
        else:
            feedback_parts.append("❌ Posted grade (78%) not found")
        
        score += data_score
        
        # CRITERION 2: Formula Usage (25 points)
        formula_score = 0
        
        # Check if formulas were detected
        if len(formulas_found) >= 3:
            formula_score += 10
            feedback_parts.append(f"✅ Multiple formulas detected ({len(formulas_found)} formulas)")
        elif len(formulas_found) >= 1:
            formula_score += 5
            feedback_parts.append(f"⚠️ Some formulas detected ({len(formulas_found)} formula(s))")
        else:
            feedback_parts.append("⚠️ No formulas explicitly detected (may still be correct if results are accurate)")
        
        # Expected intermediate results
        expected_lab_avg = 88.11  # (88+92+78+90+85+91+89+87+93)/9
        expected_hw_avg = 90.58   # Sum of all HW / 12
        expected_midterm = 82
        expected_final = 79
        expected_weighted = 84.22 # Weighted average
        expected_discrepancy = 6.22
        
        # Check for lab average (~88%)
        lab_avg_found = any(abs(v - expected_lab_avg) < 3 for v in all_values_numeric)
        if lab_avg_found:
            formula_score += 5
            feedback_parts.append("✅ Lab average calculated correctly (~88%)")
        else:
            # Check if they might have included the 0 (wrong but common mistake)
            wrong_lab_avg = 79.3  # If all 10 labs included
            if any(abs(v - wrong_lab_avg) < 3 for v in all_values_numeric):
                feedback_parts.append("❌ Lab average appears to include the dropped grade (should exclude Lab 9)")
            else:
                feedback_parts.append("❌ Lab average not found or incorrect (expected ~88%)")
        
        # Check for homework average (~91%)
        hw_avg_found = any(abs(v - expected_hw_avg) < 3 for v in all_values_numeric)
        if hw_avg_found:
            formula_score += 5
            feedback_parts.append("✅ Homework average calculated correctly (~91%)")
        else:
            # Check if close
            if any(87 <= v <= 94 for v in all_values_numeric):
                formula_score += 2
                feedback_parts.append("⚠️ Homework average may be present but not exact")
            else:
                feedback_parts.append("❌ Homework average not found (expected ~91%)")
        
        # Check for exam scores present
        has_82 = any(abs(v - 82) < 1 for v in all_values_numeric)
        has_79 = any(abs(v - 79) < 1 for v in all_values_numeric)
        if has_82 and has_79:
            formula_score += 5
            feedback_parts.append("✅ Exam scores present (82%, 79%)")
        else:
            feedback_parts.append("❌ Exam scores missing or incorrect")
        
        score += formula_score
        
        # CRITERION 3: Weighted Average Calculation (30 points)
        weighted_score = 0
        
        # Check for final calculated grade around 84%
        final_grade_found = any(abs(v - expected_weighted) < 3 for v in all_values_numeric)
        
        if final_grade_found:
            weighted_score += 30
            feedback_parts.append("✅ CORRECT: Final weighted grade calculated (~84%)")
        else:
            # Check if it's somewhat close (within 5%)
            close_values = [v for v in all_values_numeric if 79 < v < 89 and v != 82 and abs(v - 78) > 2]
            if close_values:
                weighted_score += 15
                closest = min(close_values, key=lambda x: abs(x - expected_weighted))
                feedback_parts.append(f"⚠️ Final grade present ({closest:.2f}%) but not exactly correct (expected 84.22%)")
            else:
                feedback_parts.append("❌ Final weighted grade not calculated or far from expected value (84%)")
        
        score += weighted_score
        
        # CRITERION 4: Discrepancy Identification (10 points)
        discrepancy_score = 0
        
        # Look for the discrepancy value (around 6 percentage points)
        discrepancy_values = [v for v in all_values_numeric if 5 < v < 8]
        
        if discrepancy_values:
            discrepancy_score += 10
            feedback_parts.append(f"✅ Discrepancy identified (~{discrepancy_values[0]:.2f} percentage points)")
        else:
            # They might have calculated it but in a way we didn't detect
            # Give partial credit if final grade is correct
            if final_grade_found:
                discrepancy_score += 5
                feedback_parts.append("⚠️ Final grade correct but discrepancy not explicitly shown")
            else:
                feedback_parts.append("❌ Discrepancy with posted grade not calculated")
        
        score += discrepancy_score
        
        # Final assessment
        passed = score >= 70
        final_score = score / max_score
        
        feedback = " | ".join(feedback_parts)
        
        if passed:
            if score >= 90:
                feedback = f"🎯 EXCELLENT: Grade verification spreadsheet perfectly created! (Score: {score}/{max_score}) | " + feedback
            else:
                feedback = f"✅ PASSED: Grade verification successful! (Score: {score}/{max_score}) | " + feedback
        else:
            feedback = f"❌ INCOMPLETE: Grade verification needs improvement (Score: {score}/{max_score}) | " + feedback
        
        logger.info(f"Final score: {score}/{max_score}, Passed: {passed}")
        
        return {
            "passed": passed,
            "score": final_score,
            "feedback": feedback
        }
    
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"❌ Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)
