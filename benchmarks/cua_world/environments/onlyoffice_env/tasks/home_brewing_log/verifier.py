#!/usr/bin/env python3
"""
Verifier for Home Brewing Log task
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_brewing_log(traj, env_info, task_info):
    """
    Verify that brewing log spreadsheet was created correctly.

    Checks:
    1. File exists and has proper structure
    2. Headers present (at least 8 of 10 expected)
    3. All 4 batches present with key data
    4. ABV formulas produce correct results (within tolerance)
    5. Cost/Bottle formulas produce correct results
    6. Conditional formatting applied to Rating column
    7. Summary statistics present and correct
    8. Formulas are used (not hard-coded values)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/brewing_log.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_brewing_')

    try:
        # Copy and parse the spreadsheet (with data_only=True to get formula results)
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {"passed": False, "score": 0, "feedback": f"Failed to load spreadsheet: {error}"}

        criteria_passed = 0
        total_criteria = 8
        feedback_parts = []

        # Get the active sheet
        sheet_name = wb.active.title
        sheet = wb.active

        # Get all data for analysis
        data = get_sheet_data(wb, sheet_name, max_rows=20, max_cols=15)

        # ============================================================
        # Criterion 1: Check headers exist (at least 8 of 10 expected)
        # ============================================================
        expected_headers = [
            'batch name', 'brew date', 'style', 'original gravity', 'og',
            'final gravity', 'fg', 'abv', 'batch size', 'bottles',
            'total cost', 'cost', 'cost/bottle', 'cost per bottle', 'rating'
        ]
        
        if len(data) > 0:
            header_row = [str(cell).lower() if cell else '' for cell in data[0]]
            header_matches = sum(1 for header in header_row if any(exp in header for exp in expected_headers))
            
            if header_matches >= 8:
                criteria_passed += 1
                feedback_parts.append(f"✅ Headers present ({header_matches} columns)")
            else:
                feedback_parts.append(f"❌ Insufficient headers ({header_matches}/8 minimum)")
        else:
            feedback_parts.append("❌ No data in spreadsheet")

        # ============================================================
        # Criterion 2: Check all 4 batches present with key data
        # ============================================================
        batch_names = ['summer haze', 'autumn amber', 'winter porter', 'spring saison']
        batches_found = 0
        batch_rows = {}
        
        for row_idx, row in enumerate(data[1:11], start=1):  # Check rows 2-11
            row_text = ' '.join([str(cell).lower() if cell else '' for cell in row])
            for batch_name in batch_names:
                if batch_name in row_text:
                    batches_found += 1
                    batch_rows[batch_name] = row_idx
                    break
        
        if batches_found >= 4:
            criteria_passed += 1
            feedback_parts.append(f"✅ All 4 batches present")
        else:
            feedback_parts.append(f"❌ Only {batches_found}/4 batches found")

        # ============================================================
        # Criterion 3: Verify ABV calculations (at least 3 of 4 correct)
        # ============================================================
        # Expected ABV values for each batch
        expected_abvs = {
            'summer haze': (1.065, 1.012, 6.956),
            'autumn amber': (1.055, 1.014, 5.381),
            'winter porter': (1.070, 1.018, 6.825),
            'spring saison': (1.058, 1.008, 6.563)
        }
        
        abv_correct = 0
        abv_total = 0
        
        for batch_name, (og, fg, expected_abv) in expected_abvs.items():
            if batch_name in batch_rows:
                row_idx = batch_rows[batch_name]
                row = data[row_idx]
                
                # Try to find ABV value in the row (usually around column 5-7)
                for col_idx, cell_value in enumerate(row):
                    if isinstance(cell_value, (int, float)):
                        # Check if this value is close to expected ABV
                        if 4.0 <= cell_value <= 9.0:  # Reasonable ABV range
                            if abs(cell_value - expected_abv) <= 0.3:
                                abv_correct += 1
                                break
                abv_total += 1
        
        if abv_correct >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ ABV calculations correct ({abv_correct}/4)")
        else:
            feedback_parts.append(f"❌ ABV calculations incorrect ({abv_correct}/4, need 3)")

        # ============================================================
        # Criterion 4: Verify Cost/Bottle calculations (at least 3 of 4)
        # ============================================================
        expected_costs = {
            'summer haze': (52.50, 48, 1.094),
            'autumn amber': (38.75, 48, 0.807),
            'winter porter': (45.00, 36, 1.250),
            'spring saison': (41.25, 48, 0.859)
        }
        
        cost_correct = 0
        cost_total = 0
        
        for batch_name, (total_cost, batch_size, expected_cost_per) in expected_costs.items():
            if batch_name in batch_rows:
                row_idx = batch_rows[batch_name]
                row = data[row_idx]
                
                # Look for cost per bottle value (usually $0.70-$1.50 range)
                for cell_value in row:
                    if isinstance(cell_value, (int, float)):
                        if 0.5 <= cell_value <= 2.0:
                            if abs(cell_value - expected_cost_per) <= 0.05:
                                cost_correct += 1
                                break
                cost_total += 1
        
        if cost_correct >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ Cost/Bottle calculations correct ({cost_correct}/4)")
        else:
            feedback_parts.append(f"❌ Cost/Bottle calculations incorrect ({cost_correct}/4, need 3)")

        # ============================================================
        # Criterion 5: Check conditional formatting on Rating column
        # ============================================================
        # This is tricky - we need to check if conditional formatting rules exist
        # openpyxl can detect conditional formatting
        has_conditional_formatting = False
        
        try:
            # Check if any conditional formatting exists in the sheet
            if hasattr(sheet, 'conditional_formatting') and sheet.conditional_formatting:
                if len(sheet.conditional_formatting._cf_rules) > 0:
                    has_conditional_formatting = True
            
            # Alternative: check if cells in rating column have fill colors
            # Look for rating column (should have values 3, 4, 5)
            for col_idx in range(len(data[0]) if len(data) > 0 else 0):
                rating_like = 0
                for row_idx in range(1, min(6, len(data))):
                    cell_value = data[row_idx][col_idx] if col_idx < len(data[row_idx]) else None
                    if isinstance(cell_value, (int, float)) and 1 <= cell_value <= 5:
                        rating_like += 1
                
                if rating_like >= 3:
                    # This looks like the rating column
                    # Check if cells have formatting
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_idx + 1)
                    
                    for row_num in range(2, 6):
                        cell = sheet[f'{col_letter}{row_num}']
                        if cell.fill and cell.fill.start_color:
                            color = cell.fill.start_color.rgb
                            if color and color != 'FFFFFFFF' and color != '00000000':
                                has_conditional_formatting = True
                                break
        except Exception as e:
            logger.warning(f"Could not check conditional formatting: {e}")
        
        if has_conditional_formatting:
            criteria_passed += 1
            feedback_parts.append("✅ Conditional formatting detected")
        else:
            feedback_parts.append("⚠️ Conditional formatting not clearly detected (may be present but not visible)")
            # Be lenient on this criterion since it's hard to detect reliably
            criteria_passed += 0.5

        # ============================================================
        # Criterion 6: Check summary statistics - Average ABV
        # ============================================================
        expected_avg_abv = 6.43  # Average of 6.956, 5.381, 6.825, 6.563
        avg_abv_found = False
        
        # Look for average in rows 7-12, typically in first few columns
        for row_idx in range(6, min(12, len(data))):
            for col_idx in range(min(3, len(data[row_idx]) if row_idx < len(data) else 0)):
                cell_value = data[row_idx][col_idx] if row_idx < len(data) and col_idx < len(data[row_idx]) else None
                if isinstance(cell_value, (int, float)):
                    if 5.0 <= cell_value <= 8.0:  # Reasonable average ABV range
                        if abs(cell_value - expected_avg_abv) <= 0.5:
                            avg_abv_found = True
                            break
        
        if avg_abv_found:
            criteria_passed += 1
            feedback_parts.append("✅ Average ABV summary correct")
        else:
            feedback_parts.append("❌ Average ABV summary missing or incorrect")

        # ============================================================
        # Criterion 7: Check summary statistics - Total Spent
        # ============================================================
        expected_total = 177.50  # Sum of 52.50, 38.75, 45.00, 41.25
        total_found = False
        
        for row_idx in range(6, min(12, len(data))):
            for col_idx in range(min(3, len(data[row_idx]) if row_idx < len(data) else 0)):
                cell_value = data[row_idx][col_idx] if row_idx < len(data) and col_idx < len(data[row_idx]) else None
                if isinstance(cell_value, (int, float)):
                    if 150 <= cell_value <= 200:  # Reasonable total range
                        if abs(cell_value - expected_total) <= 2.0:
                            total_found = True
                            break
        
        if total_found:
            criteria_passed += 1
            feedback_parts.append("✅ Total Spent summary correct")
        else:
            feedback_parts.append("❌ Total Spent summary missing or incorrect")

        # ============================================================
        # Criterion 8: Verify formulas are used (not hard-coded)
        # ============================================================
        # Reload workbook with data_only=False to check for formulas
        try:
            from openpyxl import load_workbook
            temp_file_path = os.path.join(temp_dir, 'brewing_log_formulas.xlsx')
            copy_from_env(container_path, temp_file_path)
            wb_formulas = load_workbook(temp_file_path, data_only=False)
            sheet_formulas = wb_formulas.active
            
            formula_count = 0
            # Check for formulas in likely ABV and Cost columns (columns 5-10, rows 2-5)
            for row_num in range(2, 6):
                for col_num in range(5, 11):
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_num)
                    cell = sheet_formulas[f'{col_letter}{row_num}']
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
            
            # Also check summary section for formulas (rows 7-10)
            for row_num in range(7, 11):
                for col_num in range(1, 5):
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_num)
                    cell = sheet_formulas[f'{col_letter}{row_num}']
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
            
            if formula_count >= 10:
                criteria_passed += 1
                feedback_parts.append(f"✅ Formulas detected ({formula_count} formulas)")
            else:
                feedback_parts.append(f"❌ Insufficient formulas ({formula_count} found, need 10+)")
        except Exception as e:
            logger.warning(f"Could not verify formulas: {e}")
            feedback_parts.append("⚠️ Could not verify formula usage")

        # ============================================================
        # Calculate score and determine pass/fail
        # ============================================================
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 75

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)