#!/usr/bin/env python3
"""
Verifier for Debt Avalanche Planner task

Verifies:
1. All 5 debts present with correct data
2. Priority ranking correct (highest interest = rank 1)
3. Total debt formula present and correct
4. Total minimum payment formula present and correct
5. Visual highlighting of highest-priority debt (Credit Card A)
"""

import sys
import os
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import openpyxl for advanced verification
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def has_background_color(cell):
    """Check if a cell has background color fill"""
    try:
        if cell.fill and cell.fill.fgColor:
            # Check if it has a color (not default/white)
            rgb = cell.fill.fgColor.rgb
            if rgb and rgb not in ['00000000', 'FFFFFFFF', '00FFFFFF', 'FFFFFF']:
                return True
        return False
    except:
        return False


def is_bold(cell):
    """Check if a cell has bold formatting"""
    try:
        if cell.font and cell.font.bold:
            return True
        return False
    except:
        return False


def verify_debt_avalanche_planner(traj, env_info, task_info):
    """
    Verify that debt avalanche planner spreadsheet was created correctly.

    Checks:
    1. File exists and is valid
    2. All 5 debts present with correct balances
    3. Priority ranking correct (Credit Card A = 1)
    4. Total debt formula present and correct (~$25,800)
    5. Total minimum payment formula present and correct (~$700)
    6. Visual highlighting on Credit Card A row
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/debt_payoff_plan.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_debt_')
    temp_file = os.path.join(temp_dir, 'debt_payoff_plan.xlsx')

    try:
        # Copy file from container
        copy_from_env(container_path, temp_file)

        if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
            return {"passed": False, "score": 0, "feedback": f"File not found or empty: {container_path}"}

        # Parse with data_only=True to get calculated values
        wb_values = load_workbook(temp_file, data_only=True)
        ws_values = wb_values.active

        # Parse with data_only=False to check formulas
        wb_formulas = load_workbook(temp_file, data_only=False)
        ws_formulas = wb_formulas.active

        criteria_passed = 0
        total_criteria = 6
        feedback_parts = []

        # Expected debt balances (allowing some tolerance for formatting)
        expected_balances = {
            'credit card a': 4200,
            'credit card b': 2800,
            'personal loan': 6500,
            'car loan': 9200,
            'student loan': 3100
        }

        # Expected interest rates
        expected_rates = {
            'credit card a': 23.99,
            'credit card b': 18.5,
            'personal loan': 12.0,
            'car loan': 6.5,
            'student loan': 4.25
        }

        # Criterion 1: Check if all 5 debts are present with correct balances (35%)
        debts_found = {}
        for row in range(1, 20):  # Search first 20 rows
            debt_name_cell = ws_values.cell(row=row, column=1).value
            balance_cell = ws_values.cell(row=row, column=2).value
            
            if debt_name_cell and isinstance(debt_name_cell, str):
                debt_name_lower = debt_name_cell.lower().strip()
                if debt_name_lower in expected_balances and isinstance(balance_cell, (int, float)):
                    if abs(balance_cell - expected_balances[debt_name_lower]) <= 100:
                        debts_found[debt_name_lower] = {
                            'row': row,
                            'balance': balance_cell
                        }

        if len(debts_found) >= 5:
            criteria_passed += 1
            feedback_parts.append(f"✅ All 5 debts present with correct balances")
        else:
            feedback_parts.append(f"❌ Only {len(debts_found)}/5 debts found with correct data")

        # Criterion 2: Check priority ranking (15%)
        # Credit Card A (highest interest at 23.99%) should be rank 1
        priority_correct = False
        if 'credit card a' in debts_found:
            cc_a_row = debts_found['credit card a']['row']
            # Check columns 5-7 for priority rank
            for col in range(5, 8):
                priority_val = ws_values.cell(row=cc_a_row, column=col).value
                if priority_val == 1 or priority_val == '1':
                    priority_correct = True
                    criteria_passed += 1
                    feedback_parts.append("✅ Priority ranking correct (Credit Card A = rank 1)")
                    break
        
        if not priority_correct:
            feedback_parts.append("❌ Priority ranking incorrect or missing")

        # Criterion 3: Check Total Debt formula (15%)
        total_debt_found = False
        total_debt_correct = False
        for row in range(8, 15):  # Search summary area
            cell_value = ws_formulas.cell(row=row, column=2).value
            if cell_value and isinstance(cell_value, str) and '=SUM' in cell_value.upper():
                # Found a SUM formula, check if result is correct
                calc_value = ws_values.cell(row=row, column=2).value
                if calc_value and isinstance(calc_value, (int, float)):
                    if abs(calc_value - 25800) <= 200:  # Total debt should be ~$25,800
                        criteria_passed += 1
                        total_debt_correct = True
                        feedback_parts.append(f"✅ Total debt formula correct: ${calc_value:,.0f}")
                        break
                    else:
                        total_debt_found = True

        if not total_debt_correct:
            if total_debt_found:
                feedback_parts.append(f"❌ Total debt formula found but result incorrect")
            else:
                feedback_parts.append("❌ Total debt formula missing (should be =SUM of balances)")

        # Criterion 4: Check Total Minimum Payments formula (10%)
        total_min_found = False
        total_min_correct = False
        for row in range(8, 15):
            cell_value = ws_formulas.cell(row=row, column=2).value
            if cell_value and isinstance(cell_value, str) and '=SUM' in cell_value.upper():
                # Check if this is a different SUM than the total debt
                calc_value = ws_values.cell(row=row, column=2).value
                if calc_value and isinstance(calc_value, (int, float)):
                    if abs(calc_value - 700) <= 50:  # Total minimums should be ~$700
                        criteria_passed += 1
                        total_min_correct = True
                        feedback_parts.append(f"✅ Total minimum payments formula correct: ${calc_value:,.0f}")
                        break

        if not total_min_correct:
            feedback_parts.append("❌ Total minimum payments formula missing or incorrect")

        # Criterion 5: Check visual highlighting on Credit Card A row (10%)
        highlighting_found = False
        if 'credit card a' in debts_found:
            cc_a_row = debts_found['credit card a']['row']
            # Check if any cell in the row has background color or bold
            has_color = False
            has_bold = False
            
            for col in range(1, 7):
                cell = ws_formulas.cell(row=cc_a_row, column=col)
                if has_background_color(cell):
                    has_color = True
                if is_bold(cell):
                    has_bold = True
            
            if has_color or has_bold:
                criteria_passed += 1
                highlighting_found = True
                style_type = "background color" if has_color else "bold text"
                feedback_parts.append(f"✅ Credit Card A row highlighted with {style_type}")

        if not highlighting_found:
            feedback_parts.append("❌ Credit Card A row not visually highlighted")

        # Criterion 6: Check extra payment calculation (15%)
        # Should be around $100 (800 - 700)
        extra_payment_correct = False
        for row in range(8, 15):
            label = ws_values.cell(row=row, column=1).value
            if label and isinstance(label, str) and 'extra' in label.lower():
                calc_value = ws_values.cell(row=row, column=2).value
                if calc_value and isinstance(calc_value, (int, float)):
                    if abs(calc_value - 100) <= 20:
                        criteria_passed += 1
                        extra_payment_correct = True
                        feedback_parts.append(f"✅ Extra payment calculation correct: ${calc_value:,.0f}")
                        break

        if not extra_payment_correct:
            feedback_parts.append("❌ Extra payment calculation missing or incorrect")

        # Calculate score with weighted criteria
        weights = [0.35, 0.15, 0.15, 0.10, 0.10, 0.15]  # Total = 1.0
        score = int(sum([weights[i] * 100 if i < criteria_passed else 0 for i in range(total_criteria)]))
        
        # Alternative simpler scoring
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 75

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)
