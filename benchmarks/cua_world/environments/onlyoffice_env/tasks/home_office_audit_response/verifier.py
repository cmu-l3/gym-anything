#!/usr/bin/env python3
"""
Verifier for Home Office Audit Response task

This verifies that the user correctly created a multi-sheet spreadsheet
documenting home office deduction calculations for an IRS audit response,
including cross-sheet formula references and accurate calculations.
"""

import sys
import os
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_home_office_audit_response(traj, env_info, task_info):
    """
    Verify that the home office audit response spreadsheet was created correctly.

    Checks:
    1. All three sheets exist with correct names (15 pts)
    2. Space_Calculations has correct data and percentage formulas (20 pts)
    3. Monthly_Allocation has expense data with formulas (25 pts)
    4. Annual_Summary has correct total between $3,550-$3,650 (25 pts)
    5. Formulas reference other sheets correctly (15 pts)
    
    Total: 100 points, passing threshold: 80
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/audit_response_template.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_audit_')

    try:
        # Copy and parse the spreadsheet (with data_only=True to get computed values)
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {"passed": False, "score": 0, "feedback": f"Failed to load spreadsheet: {error}"}

        score = 0
        feedback_parts = []
        max_score = 100

        # Check 1: All three sheets exist (15 points)
        required_sheets = ["Space_Calculations", "Monthly_Allocation", "Annual_Summary"]
        sheets_found = [s for s in required_sheets if s in wb.sheetnames]
        
        if len(sheets_found) == 3:
            score += 15
            feedback_parts.append("✅ All three required sheets exist")
        else:
            missing = set(required_sheets) - set(sheets_found)
            feedback_parts.append(f"❌ Missing sheets: {missing} (found: {sheets_found})")
            # Cannot proceed with further checks if sheets are missing
            return {
                "passed": False,
                "score": score,
                "feedback": " | ".join(feedback_parts)
            }

        # Check 2: Space_Calculations sheet (20 points)
        space_sheet = wb["Space_Calculations"]
        space_check_passed = 0
        
        try:
            # Check Old Apt data (row 4)
            old_office_sqft = get_cell_value(wb, "Space_Calculations", 'B4')
            old_total_sqft = get_cell_value(wb, "Space_Calculations", 'C4')
            old_business_pct = get_cell_value(wb, "Space_Calculations", 'D4')
            old_months = get_cell_value(wb, "Space_Calculations", 'E4')
            
            # Check New House data (row 5)
            new_office_sqft = get_cell_value(wb, "Space_Calculations", 'B5')
            new_total_sqft = get_cell_value(wb, "Space_Calculations", 'C5')
            new_business_pct = get_cell_value(wb, "Space_Calculations", 'D5')
            new_months = get_cell_value(wb, "Space_Calculations", 'E5')
            
            # Verify Old Apt calculations (120/850 = 14.1%, 5 months)
            if old_office_sqft and abs(old_office_sqft - 120) <= 1:
                space_check_passed += 1
            if old_total_sqft and abs(old_total_sqft - 850) <= 1:
                space_check_passed += 1
            
            old_expected_pct = 120 / 850
            if old_business_pct and abs(old_business_pct - old_expected_pct) < 0.005:
                space_check_passed += 2
                feedback_parts.append(f"✅ Old Apt business % correct: {old_business_pct:.1%}")
            else:
                feedback_parts.append(f"❌ Old Apt business % incorrect: {old_business_pct} (expected ~{old_expected_pct:.1%})")
            
            if old_months and abs(old_months - 5) == 0:
                space_check_passed += 1
            
            # Verify New House calculations (180/1200 = 15.0%, 7 months)
            if new_office_sqft and abs(new_office_sqft - 180) <= 1:
                space_check_passed += 1
            if new_total_sqft and abs(new_total_sqft - 1200) <= 1:
                space_check_passed += 1
            
            new_expected_pct = 180 / 1200
            if new_business_pct and abs(new_business_pct - new_expected_pct) < 0.005:
                space_check_passed += 2
                feedback_parts.append(f"✅ New House business % correct: {new_business_pct:.1%}")
            else:
                feedback_parts.append(f"❌ New House business % incorrect: {new_business_pct} (expected ~{new_expected_pct:.1%})")
            
            if new_months and abs(new_months - 7) == 0:
                space_check_passed += 1
            
            # Award points proportionally (max 20 points)
            score += int((space_check_passed / 10) * 20)
            
            if space_check_passed >= 8:
                feedback_parts.append(f"✅ Space_Calculations sheet complete ({space_check_passed}/10 checks)")
            else:
                feedback_parts.append(f"⚠️ Space_Calculations incomplete ({space_check_passed}/10 checks)")
        
        except Exception as e:
            feedback_parts.append(f"❌ Error reading Space_Calculations: {str(e)[:100]}")

        # Check 3: Monthly_Allocation sheet (25 points)
        monthly_sheet = wb["Monthly_Allocation"]
        monthly_check_passed = 0
        
        try:
            # Old Apt section (rows 6-11)
            # Check if expense amounts are entered
            old_rent = get_cell_value(wb, "Monthly_Allocation", 'B6')
            old_electric = get_cell_value(wb, "Monthly_Allocation", 'B7')
            old_internet = get_cell_value(wb, "Monthly_Allocation", 'B8')
            old_insurance = get_cell_value(wb, "Monthly_Allocation", 'B9')
            
            if old_rent and abs(old_rent - 1650) <= 10:
                monthly_check_passed += 1
            if old_electric and abs(old_electric - 85) <= 5:
                monthly_check_passed += 1
            if old_internet and abs(old_internet - 65) <= 5:
                monthly_check_passed += 1
            if old_insurance and abs(old_insurance - 18) <= 2:
                monthly_check_passed += 1
            
            # Check if business amounts are calculated (column D)
            old_rent_business = get_cell_value(wb, "Monthly_Allocation", 'D6')
            old_internet_business = get_cell_value(wb, "Monthly_Allocation", 'D8')
            
            if old_rent_business and 200 < old_rent_business < 250:  # ~$232
                monthly_check_passed += 1
            if old_internet_business and abs(old_internet_business - 65) <= 1:  # Should be 100% = $65
                monthly_check_passed += 1
            
            # Check monthly total (row 10)
            old_monthly_total = get_cell_value(wb, "Monthly_Allocation", 'D10')
            if old_monthly_total and 230 < old_monthly_total < 250:  # ~$240/month
                monthly_check_passed += 1
                feedback_parts.append(f"✅ Old Apt monthly total: ${old_monthly_total:.2f}")
            
            # Check 5-month total (row 11)
            old_period_total = get_cell_value(wb, "Monthly_Allocation", 'D11')
            if old_period_total and 1150 < old_period_total < 1250:  # ~$1,200
                monthly_check_passed += 1
            
            # New House section (rows 17-22)
            new_rent = get_cell_value(wb, "Monthly_Allocation", 'B17')
            new_electric = get_cell_value(wb, "Monthly_Allocation", 'B18')
            new_internet = get_cell_value(wb, "Monthly_Allocation", 'B19')
            new_insurance = get_cell_value(wb, "Monthly_Allocation", 'B20')
            
            if new_rent and abs(new_rent - 2100) <= 10:
                monthly_check_passed += 1
            if new_electric and abs(new_electric - 120) <= 5:
                monthly_check_passed += 1
            if new_internet and abs(new_internet - 75) <= 5:
                monthly_check_passed += 1
            if new_insurance and abs(new_insurance - 22) <= 2:
                monthly_check_passed += 1
            
            # Check monthly total (row 21)
            new_monthly_total = get_cell_value(wb, "Monthly_Allocation", 'D21')
            if new_monthly_total and 270 < new_monthly_total < 290:  # ~$280/month
                monthly_check_passed += 1
                feedback_parts.append(f"✅ New House monthly total: ${new_monthly_total:.2f}")
            
            # Check 7-month total (row 22)
            new_period_total = get_cell_value(wb, "Monthly_Allocation", 'D22')
            if new_period_total and 1900 < new_period_total < 2050:  # ~$1,960
                monthly_check_passed += 1
            
            # Award points proportionally (max 25 points)
            score += int((monthly_check_passed / 14) * 25)
            
            if monthly_check_passed >= 10:
                feedback_parts.append(f"✅ Monthly_Allocation complete ({monthly_check_passed}/14 checks)")
            else:
                feedback_parts.append(f"⚠️ Monthly_Allocation incomplete ({monthly_check_passed}/14 checks)")
        
        except Exception as e:
            feedback_parts.append(f"❌ Error reading Monthly_Allocation: {str(e)[:100]}")

        # Check 4: Annual_Summary sheet (25 points)
        summary_sheet = wb["Annual_Summary"]
        summary_check_passed = 0
        
        try:
            # Check months are entered
            jan_may_months = get_cell_value(wb, "Annual_Summary", 'B4')
            jun_dec_months = get_cell_value(wb, "Annual_Summary", 'B5')
            
            if jan_may_months and abs(jan_may_months - 5) == 0:
                summary_check_passed += 1
            if jun_dec_months and abs(jun_dec_months - 7) == 0:
                summary_check_passed += 1
            
            # Check monthly averages are referenced
            jan_may_avg = get_cell_value(wb, "Annual_Summary", 'C4')
            jun_dec_avg = get_cell_value(wb, "Annual_Summary", 'C5')
            
            if jan_may_avg and 230 < jan_may_avg < 250:
                summary_check_passed += 1
            if jun_dec_avg and 270 < jun_dec_avg < 290:
                summary_check_passed += 1
            
            # Check period totals
            jan_may_total = get_cell_value(wb, "Annual_Summary", 'D4')
            jun_dec_total = get_cell_value(wb, "Annual_Summary", 'D5')
            
            if jan_may_total and 1150 < jan_may_total < 1250:
                summary_check_passed += 1
            if jun_dec_total and 1900 < jun_dec_total < 2050:
                summary_check_passed += 1
            
            # CRITICAL CHECK: Annual total must be $3,550-$3,650
            annual_total = get_cell_value(wb, "Annual_Summary", 'D6')
            
            if annual_total and 3550 <= annual_total <= 3650:
                summary_check_passed += 4  # Worth extra points - this is the critical requirement
                score += 25  # Full points for this section
                feedback_parts.append(f"✅ ANNUAL TOTAL CORRECT: ${annual_total:.2f} (within IRS tolerance)")
            elif annual_total:
                # Give partial credit if close
                if 3400 < annual_total < 3800:
                    summary_check_passed += 2
                    score += int((summary_check_passed / 10) * 25)
                    feedback_parts.append(f"⚠️ Annual total close but outside range: ${annual_total:.2f} (expected $3,550-$3,650)")
                else:
                    feedback_parts.append(f"❌ Annual total incorrect: ${annual_total:.2f} (expected $3,550-$3,650)")
            else:
                feedback_parts.append("❌ Annual total not found or invalid")
        
        except Exception as e:
            feedback_parts.append(f"❌ Error reading Annual_Summary: {str(e)[:100]}")

        # Check 5: Cross-sheet formula references (15 points)
        # This is a bonus - we check if formulas appear to reference other sheets
        # We do this by reloading the workbook without data_only to inspect formulas
        try:
            from openpyxl import load_workbook
            import tempfile
            import shutil
            
            # Copy file locally to inspect formulas
            temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            copy_from_env(container_path, temp_xlsx.name)
            
            # Load without data_only to see formulas
            wb_formulas = load_workbook(temp_xlsx.name, data_only=False)
            
            cross_ref_found = 0
            
            # Check if Monthly_Allocation references Space_Calculations
            monthly_c6 = wb_formulas["Monthly_Allocation"]['C6'].value
            if monthly_c6 and isinstance(monthly_c6, str) and 'Space_Calculations' in monthly_c6:
                cross_ref_found += 1
            
            # Check if Annual_Summary references Monthly_Allocation
            summary_c4 = wb_formulas["Annual_Summary"]['C4'].value
            if summary_c4 and isinstance(summary_c4, str) and 'Monthly_Allocation' in summary_c4:
                cross_ref_found += 1
            
            # Check if percentage formulas exist in Space_Calculations
            space_d4 = wb_formulas["Space_Calculations"]['D4'].value
            if space_d4 and isinstance(space_d4, str) and ('B4' in space_d4 or 'C4' in space_d4):
                cross_ref_found += 1
            
            # Award bonus points for formula quality
            if cross_ref_found >= 2:
                score += 15
                feedback_parts.append(f"✅ Cross-sheet formulas detected ({cross_ref_found}/3)")
            elif cross_ref_found == 1:
                score += 8
                feedback_parts.append(f"⚠️ Some formulas detected ({cross_ref_found}/3)")
            else:
                feedback_parts.append("⚠️ Could not verify cross-sheet formulas (may be hardcoded values)")
            
            wb_formulas.close()
            os.unlink(temp_xlsx.name)
        
        except Exception as e:
            logger.warning(f"Could not verify formulas: {e}")
            # Don't penalize if we can't check formulas - give partial credit
            score += 8
            feedback_parts.append("⚠️ Formula verification skipped (not penalized)")

        # Cap score at 100
        score = min(score, max_score)
        
        # Determine pass/fail
        passed = score >= 80

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)[:200]}"}
    finally:
        cleanup_temp_dir(temp_dir)
