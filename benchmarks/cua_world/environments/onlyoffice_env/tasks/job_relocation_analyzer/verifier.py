#!/usr/bin/env python3
"""
Verifier for Job Relocation Analyzer task

Verifies that cost-of-living comparison spreadsheet formulas are correct.
"""

import sys
import os
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    cleanup_temp_dir
)

# Import openpyxl directly for formula checking
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def check_cell_has_formula(filepath, sheet_name, cell_ref):
    """
    Check if a cell contains a formula (not just a value)
    
    Args:
        filepath: Path to XLSX file
        sheet_name: Sheet name
        cell_ref: Cell reference (e.g., 'D15')
    
    Returns:
        Tuple of (has_formula, formula_text)
    """
    if load_workbook is None:
        return False, None
    
    try:
        # Load without data_only to see formulas
        wb = load_workbook(filepath, data_only=False)
        sheet = wb[sheet_name]
        cell = sheet[cell_ref]
        cell_value = cell.value
        
        if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
            return True, cell_value
        return False, None
    except Exception as e:
        logger.error(f"Error checking formula in {cell_ref}: {e}")
        return False, None


def verify_relocation_analyzer(traj, env_info, task_info):
    """
    Verify that relocation cost comparison spreadsheet was completed correctly.

    Checks:
    1. D15 (Denver total) has SUM formula and correct result (~4295)
    2. E15 (Austin total) has SUM formula and correct result (~5325)
    3. F15 (Difference) calculated correctly (~1030)
    4. G15 (Percentage) calculated correctly (~0.24)
    5. D17 (Annual Denver) calculated correctly (~51,540)
    6. E17 (Annual Austin) calculated correctly (~63,900)
    7. E25 (Break-even salary) calculated correctly (~105,331)
    8. E26 (Real increase) calculated correctly (~-5,331)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/relocation_comparison_draft.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_relocation_')
    temp_file = None

    try:
        # Copy file from container
        temp_file = os.path.join(temp_dir, 'relocation_comparison.xlsx')
        copy_from_env(container_path, temp_file)

        if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
            return {"passed": False, "score": 0, "feedback": f"File not found or empty: {container_path}"}

        # Parse the spreadsheet with data_only=True to get calculated values
        wb = load_workbook(temp_file, data_only=True)
        
        if wb is None:
            return {"passed": False, "score": 0, "feedback": "Failed to load spreadsheet"}

        sheet_name = "Cost Comparison"
        
        if sheet_name not in wb.sheetnames:
            return {"passed": False, "score": 0, "feedback": f"Sheet '{sheet_name}' not found"}
        
        sheet = wb[sheet_name]

        criteria_passed = 0
        feedback_parts = []
        total_criteria = 8

        # Get the actual cost data to calculate expected values
        denver_costs = []
        austin_costs = []
        for row_num in range(5, 14):  # D5:D13 and E5:E13
            denver_val = sheet[f'D{row_num}'].value
            austin_val = sheet[f'E{row_num}'].value
            if isinstance(denver_val, (int, float)):
                denver_costs.append(denver_val)
            if isinstance(austin_val, (int, float)):
                austin_costs.append(austin_val)
        
        expected_denver_total = sum(denver_costs) if denver_costs else 4295
        expected_austin_total = sum(austin_costs) if austin_costs else 5325
        expected_difference = expected_austin_total - expected_denver_total
        expected_percentage = expected_difference / expected_denver_total if expected_denver_total > 0 else 0
        
        current_salary = sheet['D20'].value if sheet['D20'].value else 85000
        offered_salary = sheet['E21'].value if sheet['E21'].value else 100000
        expected_breakeven = current_salary * (expected_austin_total / expected_denver_total) if expected_denver_total > 0 else 0
        expected_real_increase = offered_salary - expected_breakeven

        # Criterion 1: Check D15 has SUM formula and correct value
        has_formula_d15, formula_d15 = check_cell_has_formula(temp_file, sheet_name, 'D15')
        denver_total = sheet['D15'].value
        
        if has_formula_d15 and 'SUM' in formula_d15.upper():
            if denver_total and isinstance(denver_total, (int, float)):
                if abs(denver_total - expected_denver_total) <= 10:
                    criteria_passed += 1
                    feedback_parts.append(f"✅ Denver total correct: ${denver_total:,.2f} (SUM formula)")
                else:
                    feedback_parts.append(f"⚠️ Denver total has SUM but value unexpected: ${denver_total:,.2f} (expected ~${expected_denver_total:,.2f})")
                    criteria_passed += 0.5
            else:
                feedback_parts.append(f"❌ Denver total has SUM formula but no valid result")
        else:
            if denver_total and isinstance(denver_total, (int, float)) and abs(denver_total - expected_denver_total) <= 10:
                feedback_parts.append(f"⚠️ Denver total value correct (${denver_total:,.2f}) but no SUM formula detected")
                criteria_passed += 0.5
            else:
                feedback_parts.append(f"❌ Denver total missing SUM formula (found: {formula_d15 if has_formula_d15 else 'no formula'})")

        # Criterion 2: Check E15 has SUM formula and correct value
        has_formula_e15, formula_e15 = check_cell_has_formula(temp_file, sheet_name, 'E15')
        austin_total = sheet['E15'].value
        
        if has_formula_e15 and 'SUM' in formula_e15.upper():
            if austin_total and isinstance(austin_total, (int, float)):
                if abs(austin_total - expected_austin_total) <= 10:
                    criteria_passed += 1
                    feedback_parts.append(f"✅ Austin total correct: ${austin_total:,.2f} (SUM formula)")
                else:
                    feedback_parts.append(f"⚠️ Austin total has SUM but value unexpected: ${austin_total:,.2f} (expected ~${expected_austin_total:,.2f})")
                    criteria_passed += 0.5
            else:
                feedback_parts.append(f"❌ Austin total has SUM formula but no valid result")
        else:
            if austin_total and isinstance(austin_total, (int, float)) and abs(austin_total - expected_austin_total) <= 10:
                feedback_parts.append(f"⚠️ Austin total value correct (${austin_total:,.2f}) but no SUM formula detected")
                criteria_passed += 0.5
            else:
                feedback_parts.append(f"❌ Austin total missing SUM formula (found: {formula_e15 if has_formula_e15 else 'no formula'})")

        # Criterion 3: Check F15 (Difference) is correct
        difference = sheet['F15'].value
        
        if difference and isinstance(difference, (int, float)):
            if abs(difference - expected_difference) <= 10:
                criteria_passed += 1
                feedback_parts.append(f"✅ Difference correct: ${difference:,.2f}")
            else:
                feedback_parts.append(f"❌ Difference incorrect: ${difference:,.2f} (expected ~${expected_difference:,.2f})")
        else:
            feedback_parts.append(f"❌ Difference missing or invalid: {difference}")

        # Criterion 4: Check G15 (Percentage) is correct
        percentage = sheet['G15'].value
        
        if percentage is not None and isinstance(percentage, (int, float)):
            # Percentage could be stored as decimal (0.24) or as percentage (24)
            # Check both possibilities
            if abs(percentage - expected_percentage) <= 0.02 or abs(percentage - (expected_percentage * 100)) <= 2:
                criteria_passed += 1
                if percentage < 1:
                    feedback_parts.append(f"✅ Percentage correct: {percentage:.1%}")
                else:
                    feedback_parts.append(f"✅ Percentage correct: {percentage:.1f}%")
            else:
                feedback_parts.append(f"❌ Percentage incorrect: {percentage} (expected ~{expected_percentage:.1%})")
        else:
            feedback_parts.append(f"❌ Percentage missing or invalid: {percentage}")

        # Criterion 5: Check D17 (Annual Denver) is correct
        annual_denver = sheet['D17'].value
        expected_annual_denver = expected_denver_total * 12
        
        if annual_denver and isinstance(annual_denver, (int, float)):
            if abs(annual_denver - expected_annual_denver) <= 100:
                criteria_passed += 1
                feedback_parts.append(f"✅ Annual Denver costs correct: ${annual_denver:,.2f}")
            else:
                feedback_parts.append(f"❌ Annual Denver costs incorrect: ${annual_denver:,.2f} (expected ~${expected_annual_denver:,.2f})")
        else:
            feedback_parts.append(f"❌ Annual Denver costs missing or invalid: {annual_denver}")

        # Criterion 6: Check E17 (Annual Austin) is correct
        annual_austin = sheet['E17'].value
        expected_annual_austin = expected_austin_total * 12
        
        if annual_austin and isinstance(annual_austin, (int, float)):
            if abs(annual_austin - expected_annual_austin) <= 100:
                criteria_passed += 1
                feedback_parts.append(f"✅ Annual Austin costs correct: ${annual_austin:,.2f}")
            else:
                feedback_parts.append(f"❌ Annual Austin costs incorrect: ${annual_austin:,.2f} (expected ~${expected_annual_austin:,.2f})")
        else:
            feedback_parts.append(f"❌ Annual Austin costs missing or invalid: {annual_austin}")

        # Criterion 7: Check E25 (Break-even salary) is correct
        breakeven_salary = sheet['E25'].value
        
        if breakeven_salary and isinstance(breakeven_salary, (int, float)):
            if abs(breakeven_salary - expected_breakeven) <= 100:
                criteria_passed += 1
                feedback_parts.append(f"✅ Break-even salary correct: ${breakeven_salary:,.2f}")
            else:
                feedback_parts.append(f"❌ Break-even salary incorrect: ${breakeven_salary:,.2f} (expected ~${expected_breakeven:,.2f})")
        else:
            feedback_parts.append(f"❌ Break-even salary missing or invalid: {breakeven_salary}")

        # Criterion 8: Check E26 (Real increase) is correct
        real_increase = sheet['E26'].value
        
        if real_increase is not None and isinstance(real_increase, (int, float)):
            if abs(real_increase - expected_real_increase) <= 100:
                criteria_passed += 1
                if real_increase < 0:
                    feedback_parts.append(f"✅ Real salary increase correct: ${real_increase:,.2f} (actually a loss!)")
                else:
                    feedback_parts.append(f"✅ Real salary increase correct: ${real_increase:,.2f}")
            else:
                feedback_parts.append(f"❌ Real salary increase incorrect: ${real_increase:,.2f} (expected ~${expected_real_increase:,.2f})")
        else:
            feedback_parts.append(f"❌ Real salary increase missing or invalid: {real_increase}")

        # Calculate final score
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 70

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)