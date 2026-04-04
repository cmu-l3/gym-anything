#!/usr/bin/env python3
"""
Verifier for Used Car Comparison task

This verifier checks that the agent created a proper car comparison spreadsheet with:
- Correct headers
- Data for 4 vehicles
- Formulas (not hardcoded values) for cost-per-mile calculation
- Notes for vehicles
"""

import sys
import os
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import cleanup_temp_dir

# Import openpyxl directly for both value and formula checking
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_car_comparison(traj, env_info, task_info):
    """
    Verify that car comparison spreadsheet was created correctly.

    Verification Criteria (8 total, need 5+ to pass):
    1. File exists and is readable
    2. Headers present (Vehicle, Year, Mileage, Price, Cost_Per_Mile, Notes)
    3. Exactly 4 vehicles (rows 2-5)
    4. Price data accurate for all 4 vehicles (±$50)
    5. Mileage data accurate for all 4 vehicles (±500)
    6. Cost_Per_Mile contains formulas (not hardcoded values)
    7. Formula results are accurate (±$0.02/mile)
    8. Notes present for at least 3 vehicles
    """
    if load_workbook is None:
        return {"passed": False, "score": 0, "feedback": "openpyxl not installed"}

    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/car_comparison.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_car_')
    temp_file = os.path.join(temp_dir, 'car_comparison.xlsx')

    try:
        # Copy file from container
        copy_from_env(container_path, temp_file)

        if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
            return {"passed": False, "score": 0, "feedback": f"File not found or empty: {container_path}"}

        # Load workbook twice - once for values, once for formulas
        wb_values = load_workbook(temp_file, data_only=True)
        ws_values = wb_values.active

        wb_formulas = load_workbook(temp_file, data_only=False)
        ws_formulas = wb_formulas.active

        criteria_passed = 0
        total_criteria = 8
        feedback_parts = []

        # Expected vehicle data
        expected_vehicles = [
            {
                "name": "Honda Civic", 
                "year": 2015, 
                "mileage": 67000, 
                "price": 12500, 
                "cost_per_mile": 0.1506,
                "has_note": True
            },
            {
                "name": "Toyota Corolla", 
                "year": 2014, 
                "mileage": 89000, 
                "price": 10200, 
                "cost_per_mile": 0.1672,
                "has_note": True
            },
            {
                "name": "Mazda3", 
                "year": 2016, 
                "mileage": 54000, 
                "price": 13800, 
                "cost_per_mile": 0.1438,
                "has_note": True
            },
            {
                "name": "Ford Focus", 
                "year": 2013, 
                "mileage": 103000, 
                "price": 8900, 
                "cost_per_mile": 0.1894,
                "has_note": True
            }
        ]

        # === Criterion 1: Check headers ===
        header_row = []
        for col in range(1, 12):  # Check first 11 columns
            cell_value = ws_values.cell(row=1, column=col).value
            if cell_value:
                header_row.append(str(cell_value).lower().strip().replace('_', ' ').replace('-', ' '))

        # Flexible header detection
        vehicle_col = None
        year_col = None
        mileage_col = None
        price_col = None
        cost_col = None
        notes_col = None

        for idx, header in enumerate(header_row, start=1):
            h = header.replace(' ', '')
            if 'vehicle' in h or 'car' in h or 'model' in h:
                vehicle_col = idx
            elif 'year' in h:
                year_col = idx
            elif 'mileage' in h or 'miles' in h or 'odometer' in h:
                mileage_col = idx
            elif 'price' in h and not cost_col:
                price_col = idx
            elif ('cost' in h and 'per' in h) or ('cost' in h and 'mile' in h) or 'costpermile' in h:
                cost_col = idx
            elif 'note' in h or 'comment' in h or 'description' in h:
                notes_col = idx

        # If standard order, assume columns even if headers are slightly off
        if len(header_row) >= 5 and not all([vehicle_col, price_col, mileage_col, cost_col]):
            # Assume: A=Vehicle, B=Year, C=Mileage, D=Price, E=Cost, F=Notes
            if not vehicle_col:
                vehicle_col = 1
            if not year_col:
                year_col = 2
            if not mileage_col:
                mileage_col = 3
            if not price_col:
                price_col = 4
            if not cost_col:
                cost_col = 5
            if not notes_col:
                notes_col = 6

        headers_valid = all([vehicle_col, mileage_col, price_col, cost_col])

        if headers_valid:
            criteria_passed += 1
            feedback_parts.append("✅ Headers present and valid")
        else:
            feedback_parts.append(f"❌ Headers missing or incomplete. Found: {', '.join(header_row) if header_row else 'none'}")
            # Cannot proceed without proper structure
            return {
                "passed": False,
                "score": int((criteria_passed / total_criteria) * 100),
                "feedback": " | ".join(feedback_parts) + " | Cannot verify without proper headers"
            }

        # === Criterion 2: Check vehicle count ===
        vehicle_count = 0
        for row in range(2, 10):  # Check rows 2-9
            cell_val = ws_values.cell(row=row, column=vehicle_col).value
            if cell_val and str(cell_val).strip():
                vehicle_count += 1

        if vehicle_count == 4:
            criteria_passed += 1
            feedback_parts.append("✅ Exactly 4 vehicles present")
        else:
            feedback_parts.append(f"❌ Expected 4 vehicles, found {vehicle_count}")

        # === Criterion 3: Check price data ===
        prices_correct = 0
        for i, expected in enumerate(expected_vehicles, start=2):
            price_val = ws_values.cell(row=i, column=price_col).value
            
            if price_val is not None:
                try:
                    price_num = float(price_val)
                    if abs(price_num - expected['price']) <= 50:
                        prices_correct += 1
                except (ValueError, TypeError):
                    pass

        if prices_correct >= 4:
            criteria_passed += 1
            feedback_parts.append(f"✅ Price data accurate ({prices_correct}/4 correct)")
        elif prices_correct >= 3:
            feedback_parts.append(f"⚠️ Price data mostly correct ({prices_correct}/4 correct)")
        else:
            feedback_parts.append(f"❌ Price data incorrect ({prices_correct}/4 correct)")

        # === Criterion 4: Check mileage data ===
        mileages_correct = 0
        for i, expected in enumerate(expected_vehicles, start=2):
            mileage_val = ws_values.cell(row=i, column=mileage_col).value
            
            if mileage_val is not None:
                try:
                    mileage_num = float(mileage_val)
                    if abs(mileage_num - expected['mileage']) <= 500:
                        mileages_correct += 1
                except (ValueError, TypeError):
                    pass

        if mileages_correct >= 4:
            criteria_passed += 1
            feedback_parts.append(f"✅ Mileage data accurate ({mileages_correct}/4 correct)")
        elif mileages_correct >= 3:
            feedback_parts.append(f"⚠️ Mileage data mostly correct ({mileages_correct}/4 correct)")
        else:
            feedback_parts.append(f"❌ Mileage data incorrect ({mileages_correct}/4 correct)")

        # === Criterion 5: Check formulas exist ===
        formulas_found = 0
        for i in range(2, 6):  # Rows 2-5
            cell = ws_formulas.cell(row=i, column=cost_col)
            cell_value = cell.value
            
            # Check if cell contains a formula (starts with =)
            if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
                formulas_found += 1

        if formulas_found >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ Formulas present ({formulas_found}/4 cells contain formulas)")
        else:
            feedback_parts.append(f"❌ Formulas missing ({formulas_found}/4 cells contain formulas, may be hardcoded)")

        # === Criterion 6: Check formula results accuracy ===
        accurate_results = 0
        for i, expected in enumerate(expected_vehicles, start=2):
            cost_val = ws_values.cell(row=i, column=cost_col).value
            
            if cost_val is not None:
                try:
                    cost_num = float(cost_val)
                    # Check if close to expected (within 2 cents per mile)
                    if abs(cost_num - expected['cost_per_mile']) <= 0.02:
                        accurate_results += 1
                except (ValueError, TypeError):
                    pass

        if accurate_results >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ Formula results accurate ({accurate_results}/4 correct)")
        elif accurate_results >= 2:
            feedback_parts.append(f"⚠️ Formula results partially correct ({accurate_results}/4 correct)")
        else:
            feedback_parts.append(f"❌ Formula results incorrect ({accurate_results}/4 correct)")

        # === Criterion 7: Check notes present ===
        notes_present = 0
        if notes_col:
            for i in range(2, 6):  # Rows 2-5
                note_val = ws_values.cell(row=i, column=notes_col).value
                if note_val and str(note_val).strip() and len(str(note_val).strip()) > 5:
                    notes_present += 1

        if notes_present >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ Notes present ({notes_present}/4 vehicles have notes)")
        else:
            feedback_parts.append(f"❌ Notes insufficient ({notes_present}/4 vehicles have notes, need at least 3)")

        # Calculate final score
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 70  # Need at least 5.6/8 criteria, so 6/8 = 75%

        feedback = " | ".join(feedback_parts)

        logger.info(f"Verification complete: {criteria_passed}/{total_criteria} criteria passed, score={score}%")

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)