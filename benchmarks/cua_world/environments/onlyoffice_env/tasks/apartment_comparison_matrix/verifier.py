#!/usr/bin/env python3
"""
Verifier for Apartment Comparison Matrix task

Verifies that the user created a comparison spreadsheet with:
- Header row with required columns (bold formatting)
- 5 apartment rows with complete data
- Formulas for total monthly cost
- Commute times and descriptive features
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_apartment_comparison(traj, env_info, task_info):
    """
    Verify apartment comparison spreadsheet task.
    
    Checks:
    1. Header row contains required columns (2 pts)
    2. Header row has bold formatting (1 pt)
    3. 5 apartment rows with identifiers (2 pts)
    4. Complete numeric data (rent, utilities, parking) (2 pts)
    5. Formulas in Total Cost column (2 pts) - MOST CRITICAL
    6. Commute information present (0.5 pts)
    7. Descriptive features/pros/cons (0.5 pts)
    
    Pass threshold: 7.0/10.0 points
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/apartment_comparison.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_apartment_')

    try:
        # Copy and parse the spreadsheet
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"Failed to load spreadsheet: {error}"
            }

        sheet = wb.active
        feedback_parts = []
        score = 0.0

        # Extract all cell data for analysis
        max_row = min(sheet.max_row, 20)  # Look at first 20 rows
        max_col = min(sheet.max_column, 15)  # Look at first 15 columns

        # ===================================================================
        # CHECK 1: Header row contains required columns (2 points)
        # ===================================================================
        header_row_idx = None
        header_data = []
        
        # Find the header row (look in first 3 rows)
        for row_idx in range(1, min(4, max_row + 1)):
            row_values = [sheet.cell(row_idx, col).value for col in range(1, max_col + 1)]
            row_text = " ".join([str(v).lower() if v else "" for v in row_values])
            
            # Check if this looks like a header row (has multiple key terms)
            has_apartment = any(term in row_text for term in ["apartment", "name", "address", "unit"])
            has_rent = any(term in row_text for term in ["rent", "price", "cost"])
            has_utilities = any(term in row_text for term in ["utilit", "util", "utils"])
            
            if has_apartment and has_rent:
                header_row_idx = row_idx
                header_data = row_values
                break
        
        if header_row_idx:
            header_text = " ".join([str(v).lower() if v else "" for v in header_data])
            
            # Check for required columns (flexible matching)
            required_checks = {
                'apartment': any(term in header_text for term in ["apartment", "name", "address", "unit", "place"]),
                'rent': any(term in header_text for term in ["rent", "price", "base", "monthly"]),
                'utilities': any(term in header_text for term in ["utilit", "util", "utils"]),
                'parking': any(term in header_text for term in ["park", "garage"]),
                'total': any(term in header_text for term in ["total", "sum", "all"]),
                'commute': any(term in header_text for term in ["commute", "travel", "distance", "time", "work"])
            }
            
            headers_found = sum(required_checks.values())
            
            if headers_found >= 5:
                score += 2.0
                feedback_parts.append(f"✅ Header row contains required columns ({headers_found}/6)")
            elif headers_found >= 4:
                score += 1.5
                feedback_parts.append(f"⚠️  Header row mostly complete ({headers_found}/6 columns)")
            else:
                score += 0.5
                feedback_parts.append(f"❌ Header row incomplete ({headers_found}/6 columns)")
        else:
            feedback_parts.append("❌ No clear header row found")
        
        # ===================================================================
        # CHECK 2: Header formatting - bold (1 point)
        # ===================================================================
        if header_row_idx:
            bold_count = 0
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(header_row_idx, col_idx)
                if cell.value and cell.font and cell.font.bold:
                    bold_count += 1
            
            if bold_count >= 4:
                score += 1.0
                feedback_parts.append(f"✅ Header row has bold formatting ({bold_count} cells)")
            elif bold_count >= 2:
                score += 0.5
                feedback_parts.append(f"⚠️  Header formatting partial ({bold_count} bold cells)")
            else:
                feedback_parts.append(f"❌ Header formatting missing (only {bold_count} bold cells)")
        
        # ===================================================================
        # CHECK 3: 5 apartment rows with data (2 points)
        # ===================================================================
        apartment_rows = []
        data_start_row = (header_row_idx + 1) if header_row_idx else 2
        
        for row_idx in range(data_start_row, min(data_start_row + 10, max_row + 1)):
            row_values = [sheet.cell(row_idx, col).value for col in range(1, max_col + 1)]
            # Check if row has meaningful data (at least 3 non-empty cells)
            non_empty = sum(1 for v in row_values if v is not None and str(v).strip())
            if non_empty >= 3:
                apartment_rows.append(row_idx)
        
        apartment_count = len(apartment_rows)
        if apartment_count >= 5:
            score += 2.0
            feedback_parts.append(f"✅ Found {apartment_count} apartment rows")
        elif apartment_count >= 4:
            score += 1.5
            feedback_parts.append(f"⚠️  Found {apartment_count}/5 apartment rows")
        elif apartment_count >= 3:
            score += 1.0
            feedback_parts.append(f"⚠️  Found only {apartment_count}/5 apartment rows")
        else:
            feedback_parts.append(f"❌ Found only {apartment_count}/5 apartment rows")
        
        # ===================================================================
        # CHECK 4: Complete numeric data (rent, utilities, parking) (2 pts)
        # ===================================================================
        rows_with_complete_data = 0
        
        for row_idx in apartment_rows[:5]:
            row_values = [sheet.cell(row_idx, col).value for col in range(1, max_col + 1)]
            
            # Count numeric cells that look like costs (between 0 and 2000)
            numeric_costs = []
            for val in row_values:
                if isinstance(val, (int, float)) and 0 <= val <= 2500:
                    numeric_costs.append(val)
                elif isinstance(val, str):
                    # Try to extract numbers from strings like "$1450"
                    match = re.search(r'\d+', val)
                    if match:
                        num = int(match.group())
                        if 0 <= num <= 2500:
                            numeric_costs.append(num)
            
            # Need at least 3 numeric values (rent, utilities, parking) or 4 (including total)
            if len(numeric_costs) >= 3:
                rows_with_complete_data += 1
        
        if rows_with_complete_data >= 5:
            score += 2.0
            feedback_parts.append(f"✅ {rows_with_complete_data} apartments have complete cost data")
        elif rows_with_complete_data >= 4:
            score += 1.5
            feedback_parts.append(f"⚠️  {rows_with_complete_data}/5 apartments have complete cost data")
        elif rows_with_complete_data >= 3:
            score += 1.0
            feedback_parts.append(f"⚠️  Only {rows_with_complete_data}/5 apartments have complete cost data")
        else:
            feedback_parts.append(f"❌ Only {rows_with_complete_data}/5 apartments have complete cost data")
        
        # ===================================================================
        # CHECK 5: Formulas in Total column (2 points) - MOST CRITICAL
        # ===================================================================
        formula_count = 0
        correct_formula_count = 0
        
        for row_idx in apartment_rows[:5]:
            row_has_formula = False
            
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                
                # Check if cell has a formula (stored in cell.value or cell.data_type)
                # In openpyxl with data_only=True, formulas are evaluated
                # We need to check if it's likely a calculated value
                
                # Heuristic: Look for cells with values between 1400-2000 (total cost range)
                if isinstance(cell.value, (int, float)) and 1400 <= cell.value <= 2200:
                    # This could be a total cost
                    # Try to verify if it's close to sum of other cells in row
                    row_values = [sheet.cell(row_idx, c).value for c in range(1, max_col + 1)]
                    numeric_values = [v for v in row_values if isinstance(v, (int, float)) and 0 < v < 2000]
                    
                    if len(numeric_values) >= 3:
                        # Check if this cell value is approximately sum of 2-3 other values
                        for i in range(len(numeric_values)):
                            for j in range(i+1, len(numeric_values)):
                                for k in range(j+1, len(numeric_values)):
                                    potential_sum = numeric_values[i] + numeric_values[j] + numeric_values[k]
                                    if abs(cell.value - potential_sum) <= 5:
                                        row_has_formula = True
                                        correct_formula_count += 1
                                        break
                                if row_has_formula:
                                    break
                            if row_has_formula:
                                break
                
                if row_has_formula:
                    formula_count += 1
                    break
        
        # Alternative check: Look for any formulas using openpyxl formula detection
        # Reload workbook without data_only to see formulas
        try:
            from openpyxl import load_workbook
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            copy_from_env(container_path, temp_file.name)
            wb_formulas = load_workbook(temp_file.name, data_only=False)
            sheet_formulas = wb_formulas.active
            
            formula_cells = 0
            for row_idx in apartment_rows[:5]:
                for col_idx in range(1, max_col + 1):
                    cell = sheet_formulas.cell(row_idx, col_idx)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_cells += 1
                        if formula_count < apartment_count:  # Update count if we found more
                            formula_count = formula_cells
            
            wb_formulas.close()
            os.unlink(temp_file.name)
        except Exception as e:
            logger.debug(f"Could not check formulas directly: {e}")
        
        if formula_count >= 3:
            score += 2.0
            feedback_parts.append(f"✅ {formula_count} apartments use formulas for total cost")
        elif formula_count >= 2:
            score += 1.5
            feedback_parts.append(f"⚠️  {formula_count} apartments use formulas (expected at least 3)")
        elif formula_count >= 1:
            score += 1.0
            feedback_parts.append(f"⚠️  Only {formula_count} apartment uses formulas")
        else:
            feedback_parts.append(f"❌ No formulas found for total cost calculation")
        
        # ===================================================================
        # CHECK 6: Commute information present (0.5 points)
        # ===================================================================
        commute_count = 0
        
        for row_idx in apartment_rows[:5]:
            row_values = [sheet.cell(row_idx, col).value for col in range(1, max_col + 1)]
            row_text = " ".join([str(v).lower() if v else "" for v in row_values])
            
            # Look for time-related patterns
            has_commute = any([
                "min" in row_text,
                "minute" in row_text,
                "walk" in row_text,
                "drive" in row_text,
                re.search(r'\d+\s*(min|minute)', row_text),
                re.search(r'\d+:\d+', row_text)  # Time format
            ])
            
            if has_commute:
                commute_count += 1
        
        if commute_count >= 4:
            score += 0.5
            feedback_parts.append(f"✅ Commute time included for {commute_count} apartments")
        elif commute_count >= 3:
            score += 0.3
            feedback_parts.append(f"⚠️  Commute time for {commute_count}/5 apartments")
        else:
            feedback_parts.append(f"⚠️  Commute time only for {commute_count}/5 apartments")
        
        # ===================================================================
        # CHECK 7: Descriptive text (pros/cons/features) (0.5 points)
        # ===================================================================
        description_count = 0
        
        for row_idx in apartment_rows[:5]:
            row_values = [sheet.cell(row_idx, col).value for col in range(1, max_col + 1)]
            
            # Look for text cells with descriptive content
            text_cells = [v for v in row_values if isinstance(v, str) and len(v.split()) >= 2]
            
            # Check for descriptive keywords
            descriptive_terms = ["washer", "dryer", "gym", "quiet", "noisy", "new", "modern", 
                                "pet", "kitchen", "landlord", "building", "floor", "park",
                                "downtown", "free", "included", "safe"]
            
            has_description = any(
                any(term in str(v).lower() for term in descriptive_terms)
                for v in text_cells
            )
            
            if has_description or len(text_cells) >= 1:
                description_count += 1
        
        if description_count >= 4:
            score += 0.5
            feedback_parts.append(f"✅ Descriptive features for {description_count} apartments")
        elif description_count >= 3:
            score += 0.3
            feedback_parts.append(f"⚠️  Descriptions for {description_count}/5 apartments")
        else:
            feedback_parts.append(f"⚠️  Limited descriptions ({description_count}/5)")
        
        # ===================================================================
        # Final assessment
        # ===================================================================
        passed = score >= 7.0  # Need 7/10 points to pass
        normalized_score = score / 10.0  # Normalize to 0-1 range
        
        feedback = " | ".join(feedback_parts)
        
        logger.info(f"Apartment comparison verification - Score: {score}/10.0, Passed: {passed}")
        
        return {
            "passed": passed,
            "score": normalized_score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)
