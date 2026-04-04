#!/usr/bin/env python3
"""
Verifier for Instrument Repair Portfolio task (instrument_repair_portfolio@1)

Checks:
1. File exists and is valid XLSX (0.15)
2. Has proper column structure with required headers (0.20)
3. Contains all 12 service records with accurate data (0.25)
4. Uses formulas for calculations (0.20)
5. Has summary section with key metrics (0.15)
6. Flags upcoming maintenance items (0.05)
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_instrument_portfolio(traj, env_info, task_info):
    """
    Verify the instrument maintenance portfolio spreadsheet.
    
    This task requires organizing scattered repair receipts into a structured
    maintenance log with formulas, summary section, and flagged upcoming work.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/instrument_portfolio.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_instrument_')

    try:
        # Copy and parse the spreadsheet
        success, workbook, error = copy_and_parse_document(
            container_path, copy_from_env, 'xlsx'
        )

        if not success:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"Failed to open spreadsheet: {error}"
            }

        score = 0.0
        feedback_parts = []
        max_score = 1.0

        # Get the active sheet
        try:
            sheet_name = workbook.sheetnames[0]
            sheet_data = get_sheet_data(workbook, sheet_name, max_rows=100, max_cols=15)
        except Exception as e:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"Could not read sheet data: {str(e)}"
            }

        # --- Criterion 1: File exists and valid XLSX (0.15) ---
        score += 0.15
        feedback_parts.append("✅ File exists and is valid XLSX format")

        # --- Criterion 2: Data structure - proper columns (0.20) ---
        structure_score = 0.0
        header_row = None
        header_row_idx = None
        
        # Required keywords in headers
        required_keywords = ['date', 'instrument', 'service', 'cost']
        optional_keywords = ['luthier', 'shop', 'notes', 'repair']
        
        # Search for header row in first 20 rows
        for idx, row in enumerate(sheet_data[:20]):
            row_text = ' '.join([str(cell).lower() if cell else '' for cell in row])
            # Check if this row contains most required keywords
            matches = sum(1 for kw in required_keywords if kw in row_text)
            if matches >= 3:  # At least 3 out of 4 required keywords
                header_row = row
                header_row_idx = idx
                break

        if header_row:
            structure_score += 0.10
            feedback_parts.append("✅ Found proper column headers (Date, Instrument, Service, Cost)")
            
            # Count data rows after header
            data_rows = []
            for row in sheet_data[header_row_idx + 1:]:
                # Check if row has meaningful content
                non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
                if len(non_empty) >= 3:  # At least 3 filled cells
                    data_rows.append(row)
            
            if len(data_rows) >= 12:
                structure_score += 0.10
                feedback_parts.append(f"✅ Has {len(data_rows)} data rows (12+ required)")
            elif len(data_rows) >= 8:
                structure_score += 0.05
                feedback_parts.append(f"⚠️ Has {len(data_rows)} data rows (12 expected)")
            else:
                feedback_parts.append(f"❌ Only {len(data_rows)} data rows (12 expected)")
        else:
            feedback_parts.append("❌ Missing proper column headers (need: Date, Instrument, Service, Cost)")

        score += structure_score

        # --- Criterion 3: Data accuracy - key records present (0.25) ---
        accuracy_score = 0.0
        
        # Flatten all cells to searchable text
        all_text = ' '.join([
            str(cell).lower().strip()
            for row in sheet_data
            for cell in row
            if cell is not None
        ])
        
        # Check for key records and details from the receipts
        key_checks = [
            ('antonio', 0.03, "Cello 'Antonio' mentioned"),
            ('daily player', 0.03, "Cello 'Daily Player' mentioned"),
            ('violin', 0.03, "Violin mentioned"),
            ('viola', 0.03, "Viola mentioned"),
            ('maria', 0.02, "Maria's Fine Instruments mentioned"),
            ('450', 0.02, "First receipt cost ($450) present"),
            ('1850', 0.02, "Crack repair cost ($1,850) present"),
            ('crack repair', 0.02, "Crack repair service mentioned"),
            ('bridge', 0.02, "Bridge work mentioned"),
            ('stringed heritage', 0.02, "Stringed Heritage Workshop mentioned"),
            ('bow rehair', 0.01, "Bow rehairing mentioned"),
        ]
        
        found_items = 0
        for keyword, points, description in key_checks:
            if keyword in all_text:
                accuracy_score += points
                found_items += 1
        
        if found_items >= 9:
            feedback_parts.append(f"✅ Found {found_items}/{len(key_checks)} key data points")
        elif found_items >= 6:
            feedback_parts.append(f"⚠️ Found {found_items}/{len(key_checks)} key data points")
        else:
            feedback_parts.append(f"❌ Only found {found_items}/{len(key_checks)} key data points")
        
        score += accuracy_score

        # --- Criterion 4: Formula usage (0.20) ---
        formula_score = 0.0
        formula_count = 0
        has_sum_formula = False
        has_calculation_formulas = False
        
        try:
            # Re-open workbook without data_only to see formulas
            from openpyxl import load_workbook
            
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            copy_from_env(container_path, temp_file.name)
            wb_formulas = load_workbook(temp_file.name, data_only=False)
            sheet_formulas = wb_formulas[sheet_name]
            
            # Scan for formulas
            for row in sheet_formulas.iter_rows(max_row=100, max_col=15):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        formula_upper = cell.value.upper()
                        if 'SUM' in formula_upper or 'SUMIF' in formula_upper:
                            has_sum_formula = True
                        # Check for calculation formulas (addition, references, etc.)
                        if '+' in formula_upper or '*' in formula_upper:
                            has_calculation_formulas = True
            
            if formula_count >= 4 and has_sum_formula:
                formula_score = 0.20
                feedback_parts.append(f"✅ Uses {formula_count} formulas including SUM/SUMIF")
            elif formula_count >= 3:
                formula_score = 0.15
                feedback_parts.append(f"✅ Uses {formula_count} formulas")
            elif formula_count >= 1:
                formula_score = 0.08
                feedback_parts.append(f"⚠️ Only {formula_count} formula(s) found (expected 4+)")
            else:
                feedback_parts.append("❌ No formulas found (should use SUM/SUMIF for totals)")
                
        except Exception as e:
            logger.warning(f"Could not verify formulas: {e}")
            # Give partial credit if we can't check formulas
            feedback_parts.append("⚠️ Could not verify formulas")
            formula_score = 0.05
        
        score += formula_score

        # --- Criterion 5: Summary section (0.15) ---
        summary_score = 0.0
        
        # Look for summary-related keywords
        summary_indicators = [
            'summary', 'total investment', 'purchase price', 
            'total maintenance', 'instrument name', '85000', '12000', '28000', '8500'
        ]
        
        summary_found = sum(1 for indicator in summary_indicators if indicator in all_text)
        
        # Check for presence of purchase prices (strong indicator of summary)
        has_purchase_prices = ('85000' in all_text or '85,000' in all_text) and \
                              ('12000' in all_text or '12,000' in all_text)
        
        if summary_found >= 4 or has_purchase_prices:
            summary_score = 0.15
            feedback_parts.append("✅ Has summary section with key metrics (purchase prices, totals)")
        elif summary_found >= 2:
            summary_score = 0.08
            feedback_parts.append("⚠️ Partial summary section detected")
        else:
            feedback_parts.append("❌ Missing summary section (should show instrument, purchase price, maintenance total, investment)")
        
        score += summary_score

        # --- Criterion 6: Upcoming maintenance flagged (0.05) ---
        upcoming_score = 0.0
        
        upcoming_keywords = [
            'upcoming', 'pending', 'needed', 'future', 
            'bridge replacement', 'recommended', 'planned'
        ]
        
        found_upcoming = sum(1 for kw in upcoming_keywords if kw in all_text)
        
        if found_upcoming >= 2:
            upcoming_score = 0.05
            feedback_parts.append("✅ Upcoming maintenance items noted")
        elif found_upcoming >= 1:
            upcoming_score = 0.03
            feedback_parts.append("⚠️ Some upcoming maintenance noted")
        else:
            feedback_parts.append("❌ No upcoming maintenance flagged")
        
        score += upcoming_score

        # Cap score at maximum
        score = min(score, max_score)
        
        # Determine pass/fail (70% threshold)
        passed = score >= 0.70
        
        # Compile feedback
        feedback = " | ".join(feedback_parts)
        
        return {
            "passed": passed,
            "score": round(score, 2),
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)


# Entry point for gym-anything framework
def verify_task(traj, env_info, task_info):
    """Wrapper function for gym-anything compatibility"""
    return verify_instrument_portfolio(traj, env_info, task_info)