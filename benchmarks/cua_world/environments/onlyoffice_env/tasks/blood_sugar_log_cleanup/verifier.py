#!/usr/bin/env python3
"""
Verifier for Blood Sugar Log Cleanup task

This verifier checks that the agent successfully:
1. Created proper sheet structure with correct columns
2. Standardized Time Category values
3. Sorted data chronologically
4. Applied conditional formatting
5. Created summary analysis with formulas
"""

import sys
import os
import logging
import tempfile
from datetime import datetime
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def parse_date_flexible(date_str):
    """Try to parse date string in various formats"""
    if not date_str:
        return None
    
    # Convert to string if it's already a datetime
    if isinstance(date_str, datetime):
        return date_str
    
    date_str = str(date_str).strip()
    
    # Try common formats
    formats = [
        '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d', '%Y/%m/%d',
        '%d/%m/%Y', '%d-%m-%Y', '%b %d, %Y', '%B %d, %Y'
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except:
            continue
    
    # Try to extract date components with regex
    # Match patterns like "Jan 15", "15-January", etc.
    month_names = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                   'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
    
    date_lower = date_str.lower()
    for idx, month in enumerate(month_names, 1):
        if month in date_lower:
            # Try to find day number
            day_match = re.search(r'\d+', date_str)
            if day_match:
                day = int(day_match.group())
                # Assume current year if not specified
                return datetime(2024, idx, min(day, 28))
    
    return None


def check_time_category_standardized(data_rows, col_idx):
    """Check if Time Category column is standardized"""
    allowed_categories = {'Fasting', 'Post-Meal', 'Bedtime', 
                         'fasting', 'post-meal', 'bedtime',
                         'FASTING', 'POST-MEAL', 'BEDTIME'}
    
    non_standard = []
    for row_idx, row in enumerate(data_rows[1:], start=2):  # Skip header
        if col_idx < len(row):
            value = row[col_idx]
            if value and str(value).strip():
                if str(value).strip() not in allowed_categories:
                    non_standard.append((row_idx, value))
    
    return len(non_standard) == 0, non_standard


def check_chronological_order(data_rows, date_col_idx):
    """Check if data is sorted chronologically"""
    dates = []
    for row in data_rows[1:]:  # Skip header
        if date_col_idx < len(row):
            date_val = row[date_col_idx]
            parsed_date = parse_date_flexible(date_val)
            if parsed_date:
                dates.append(parsed_date)
    
    if len(dates) < 2:
        return True  # Can't verify order with less than 2 dates
    
    # Check if mostly sorted (allow some flexibility)
    sorted_dates = sorted(dates)
    matches = sum(1 for a, b in zip(dates, sorted_dates) if a == b)
    
    # Consider sorted if 80%+ matches
    return matches / len(dates) >= 0.8


def check_conditional_formatting_exists(worksheet):
    """Check if conditional formatting is applied to the worksheet"""
    try:
        # Check if worksheet has any conditional formatting rules
        if hasattr(worksheet, 'conditional_formatting'):
            cf = worksheet.conditional_formatting
            if cf and len(cf._cf_rules) > 0:
                return True, len(cf._cf_rules)
        return False, 0
    except Exception as e:
        logger.warning(f"Could not check conditional formatting: {e}")
        return False, 0


def check_for_formulas(worksheet, sheet_data):
    """Check if Summary Analysis sheet contains formulas"""
    formula_count = 0
    formula_types = set()
    
    try:
        # Check cells in the sheet for formulas
        for row_idx, row in enumerate(sheet_data[:20], start=1):  # Check first 20 rows
            for col_idx, cell_value in enumerate(row[:10], start=1):  # Check first 10 cols
                # Get cell reference
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_idx}"
                
                try:
                    cell = worksheet[cell_ref]
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        # Extract formula type (AVERAGE, SUM, COUNT, etc.)
                        formula_upper = cell.value.upper()
                        for func in ['AVERAGE', 'AVERAGEIF', 'SUM', 'COUNT', 'COUNTIF']:
                            if func in formula_upper:
                                formula_types.add(func)
                except:
                    continue
    except Exception as e:
        logger.warning(f"Error checking formulas: {e}")
    
    return formula_count, formula_types


def verify_blood_sugar_cleanup(traj, env_info, task_info):
    """
    Verify blood sugar log cleanup task.
    
    Success criteria:
    1. File exists and is valid XLSX (10 points)
    2. Has required sheets: "Organized Log" and "Summary Analysis" (15 points)
    3. Organized Log has correct columns (15 points)
    4. Time Category is standardized (15 points)
    5. Data is sorted chronologically (10 points)
    6. At least 80 rows of data (10 points)
    7. Conditional formatting exists (10 points)
    8. Summary Analysis has formulas (10 points)
    9. No excessive blank rows in data (5 points)
    
    Total: 100 points, passing threshold: 70
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/blood_sugar_organized.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_bloodsugar_')

    try:
        # Copy and parse the workbook
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {
                "passed": False, 
                "score": 0, 
                "feedback": f"Failed to load organized spreadsheet: {error}. Did you save the file as blood_sugar_organized.xlsx?"
            }

        score = 0
        feedback_parts = []
        max_score = 100

        # Criterion 1: File exists and is valid (10 points)
        score += 10
        feedback_parts.append("✅ File exists and is valid XLSX")

        # Criterion 2: Check for required sheets (15 points)
        sheet_names = [s.lower() for s in wb.sheetnames]
        has_organized_log = any('organized' in s and 'log' in s for s in sheet_names)
        has_summary_analysis = any('summary' in s and 'analysis' in s for s in sheet_names)

        if has_organized_log and has_summary_analysis:
            score += 15
            feedback_parts.append("✅ Both required sheets present")
        elif has_organized_log or has_summary_analysis:
            score += 8
            feedback_parts.append("⚠️  Only one required sheet found")
        else:
            feedback_parts.append("❌ Missing required sheets: 'Organized Log' and 'Summary Analysis'")

        if not has_organized_log:
            return {
                "passed": False,
                "score": score,
                "feedback": " | ".join(feedback_parts) + " | Cannot verify further without 'Organized Log' sheet"
            }

        # Find the Organized Log sheet
        organized_sheet = None
        for sheet in wb.worksheets:
            if 'organized' in sheet.title.lower() and 'log' in sheet.title.lower():
                organized_sheet = sheet
                break

        if not organized_sheet:
            return {
                "passed": False,
                "score": score,
                "feedback": " | ".join(feedback_parts) + " | Could not find 'Organized Log' sheet"
            }

        # Get data from organized log
        org_data = get_sheet_data(organized_sheet, organized_sheet.title, max_rows=200, max_cols=10)

        if not org_data or len(org_data) < 2:
            return {
                "passed": False,
                "score": score,
                "feedback": " | ".join(feedback_parts) + " | Organized Log sheet is empty"
            }

        # Criterion 3: Check column headers (15 points)
        headers = [str(h).lower().strip() if h else '' for h in org_data[0]]
        
        required_columns = ['date', 'time category', 'reading']
        found_columns = []
        column_indices = {}
        
        for req_col in required_columns:
            for idx, header in enumerate(headers):
                if req_col in header or (req_col == 'time category' and 'category' in header):
                    found_columns.append(req_col)
                    column_indices[req_col] = idx
                    break

        if len(found_columns) >= 3:
            score += 15
            feedback_parts.append(f"✅ Required columns present: {', '.join(found_columns)}")
        elif len(found_columns) >= 2:
            score += 8
            feedback_parts.append(f"⚠️  Some required columns present: {', '.join(found_columns)}")
        else:
            feedback_parts.append(f"❌ Missing required columns. Expected: Date, Time Category, Reading")

        # Criterion 4: Check Time Category standardization (15 points)
        if 'time category' in column_indices:
            is_standardized, non_standard = check_time_category_standardized(
                org_data, column_indices['time category']
            )
            
            if is_standardized:
                score += 15
                feedback_parts.append("✅ Time Category values standardized")
            elif len(non_standard) <= 5:
                score += 8
                feedback_parts.append(f"⚠️  Mostly standardized ({len(non_standard)} non-standard values)")
            else:
                feedback_parts.append(f"❌ Time Category not standardized ({len(non_standard)} non-standard values)")
        else:
            feedback_parts.append("❌ Cannot verify Time Category - column not found")

        # Criterion 5: Check chronological sorting (10 points)
        if 'date' in column_indices:
            is_sorted = check_chronological_order(org_data, column_indices['date'])
            if is_sorted:
                score += 10
                feedback_parts.append("✅ Data sorted chronologically")
            else:
                score += 5
                feedback_parts.append("⚠️  Data not fully sorted chronologically")
        else:
            feedback_parts.append("❌ Cannot verify sorting - Date column not found")

        # Criterion 6: Check data quantity (10 points)
        data_rows = len(org_data) - 1  # Subtract header
        if data_rows >= 80:
            score += 10
            feedback_parts.append(f"✅ Sufficient data: {data_rows} rows")
        elif data_rows >= 60:
            score += 7
            feedback_parts.append(f"⚠️  Adequate data: {data_rows} rows (expected 80+)")
        else:
            feedback_parts.append(f"❌ Insufficient data: {data_rows} rows (expected 80+)")

        # Criterion 7: Check conditional formatting (10 points)
        has_formatting, num_rules = check_conditional_formatting_exists(organized_sheet)
        if has_formatting and num_rules >= 3:
            score += 10
            feedback_parts.append(f"✅ Conditional formatting applied ({num_rules} rules)")
        elif has_formatting:
            score += 5
            feedback_parts.append(f"⚠️  Some conditional formatting present ({num_rules} rules)")
        else:
            feedback_parts.append("❌ No conditional formatting detected")

        # Criterion 8: Check Summary Analysis sheet (10 points)
        if has_summary_analysis:
            summary_sheet = None
            for sheet in wb.worksheets:
                if 'summary' in sheet.title.lower() and 'analysis' in sheet.title.lower():
                    summary_sheet = sheet
                    break
            
            if summary_sheet:
                summary_data = get_sheet_data(summary_sheet, summary_sheet.title, max_rows=50, max_cols=10)
                formula_count, formula_types = check_for_formulas(summary_sheet, summary_data)
                
                if formula_count >= 3 and len(formula_types) >= 2:
                    score += 10
                    feedback_parts.append(f"✅ Summary has formulas ({formula_count} formulas, types: {', '.join(formula_types)})")
                elif formula_count >= 1:
                    score += 5
                    feedback_parts.append(f"⚠️  Summary has some formulas ({formula_count} found)")
                else:
                    feedback_parts.append("❌ Summary Analysis lacks formulas (calculations should use formulas, not hard-coded values)")

        # Criterion 9: Check for excessive blank rows (5 points)
        blank_rows = 0
        for row in org_data[1:]:  # Skip header
            if all(not cell or str(cell).strip() == '' for cell in row):
                blank_rows += 1
        
        if blank_rows <= 2:
            score += 5
            feedback_parts.append("✅ Data properly cleaned (no excessive blank rows)")
        elif blank_rows <= 5:
            score += 3
            feedback_parts.append(f"⚠️  Some blank rows remain ({blank_rows})")
        else:
            feedback_parts.append(f"❌ Too many blank rows in data ({blank_rows})")

        # Determine pass/fail
        passed = score >= 70

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False, 
            "score": 0, 
            "feedback": f"Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)
