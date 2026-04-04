#!/usr/bin/env python3
"""
Verifier for Homebrew Batch Tracker task

Verifies that the agent created a proper brewing log spreadsheet with:
- Batch metadata
- Complete fermentation data (12 days)
- Correct gravity readings
- ABV calculation formula
- Temperature analysis formulas
- Quality assessment formulas
"""

import sys
import os
import logging
import tempfile
import re
from typing import Any, Optional, Tuple

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    parse_xlsx_file,
    get_cell_value,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def copy_file_from_container(container_path: str, copy_from_env, local_path: str) -> bool:
    """Copy file from container to local path"""
    try:
        copy_from_env(container_path, local_path)
        return os.path.exists(local_path) and os.path.getsize(local_path) > 0
    except Exception as e:
        logger.error(f"Error copying file: {e}")
        return False


def is_formula(wb: Any, sheet_name: str, cell_ref: str) -> bool:
    """Check if a cell contains a formula (not a static value)"""
    try:
        sheet = wb[sheet_name]
        cell = sheet[cell_ref]
        # Check if cell has a formula
        if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
            return True
        # For openpyxl, formulas might be stored differently
        if hasattr(cell, 'data_type') and cell.data_type == 'f':
            return True
        return False
    except Exception as e:
        logger.debug(f"Error checking formula: {e}")
        return False


def get_formula_string(wb: Any, sheet_name: str, cell_ref: str) -> Optional[str]:
    """Get the formula string from a cell"""
    try:
        sheet = wb[sheet_name]
        cell = sheet[cell_ref]
        if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
            return cell.value
        return None
    except Exception as e:
        logger.debug(f"Error getting formula: {e}")
        return None


def verify_homebrew_log(traj, env_info, task_info):
    """
    Verify that brewing log spreadsheet was created correctly.

    Checks:
    1. File exists and is valid XLSX
    2. Batch metadata present (name and date)
    3. Data table structure (headers and 12 rows)
    4. Gravity values correct (OG=1.052, FG=1.012)
    5. All 12 temperature readings present
    6. ABV formula exists and calculates correctly (~5.25%)
    7. Temperature statistics formulas (MIN, MAX, AVG)
    8. Quality check formulas (IF/AND logic)
    9. Overall data accuracy
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/brewing_log.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='homebrew_verify_')
    temp_file = os.path.join(temp_dir, 'brewing_log.xlsx')

    try:
        # Copy and parse the spreadsheet
        if not copy_file_from_container(container_path, copy_from_env, temp_file):
            return {"passed": False, "score": 0, "feedback": "Failed to copy brewing log file"}

        # Parse with data_only=True to get calculated values
        wb_values = parse_xlsx_file(temp_file)
        if wb_values is None:
            return {"passed": False, "score": 0, "feedback": "Failed to parse XLSX file"}

        # Also load without data_only to check formulas
        try:
            from openpyxl import load_workbook
            wb_formulas = load_workbook(temp_file, data_only=False)
        except Exception as e:
            logger.warning(f"Could not load formulas: {e}")
            wb_formulas = None

        criteria_passed = 0
        total_criteria = 9
        feedback_parts = []

        # Get the first sheet (should be "Fermentation Log" or similar)
        sheet_name = wb_values.sheetnames[0]

        # ===================================================================
        # CRITERION 1: Batch Metadata Present
        # ===================================================================
        batch_name = get_cell_value(wb_values, sheet_name, 'B1')
        brew_date = get_cell_value(wb_values, sheet_name, 'B2')
        
        metadata_ok = False
        if batch_name and "amber ale" in str(batch_name).lower():
            if brew_date:
                metadata_ok = True
                criteria_passed += 1
                feedback_parts.append(f"✅ Batch metadata: {batch_name}, {brew_date}")
            else:
                feedback_parts.append(f"❌ Brew date missing in B2")
        else:
            feedback_parts.append(f"❌ Batch name missing or incorrect in B1: {batch_name}")

        # ===================================================================
        # CRITERION 2: Data Table Structure (Headers)
        # ===================================================================
        header_day = str(get_cell_value(wb_values, sheet_name, 'A4') or '').lower()
        header_date = str(get_cell_value(wb_values, sheet_name, 'B4') or '').lower()
        header_temp = str(get_cell_value(wb_values, sheet_name, 'C4') or '').lower()
        header_gravity = str(get_cell_value(wb_values, sheet_name, 'D4') or '').lower()

        headers_ok = (
            'day' in header_day and
            'date' in header_date and
            'temp' in header_temp and
            'gravity' in header_gravity
        )

        if headers_ok:
            criteria_passed += 1
            feedback_parts.append("✅ Data table headers correct")
        else:
            feedback_parts.append(f"❌ Headers missing/incorrect: {header_day}, {header_date}, {header_temp}, {header_gravity}")

        # ===================================================================
        # CRITERION 3: Gravity Values (OG and FG)
        # ===================================================================
        og_value = get_cell_value(wb_values, sheet_name, 'D5')  # Day 1 gravity
        fg_value = get_cell_value(wb_values, sheet_name, 'D16')  # Day 12 gravity

        gravity_ok = False
        if og_value and fg_value:
            try:
                og_float = float(og_value)
                fg_float = float(fg_value)
                if abs(og_float - 1.052) <= 0.001 and abs(fg_float - 1.012) <= 0.001:
                    gravity_ok = True
                    criteria_passed += 1
                    feedback_parts.append(f"✅ Gravity values: OG={og_float:.3f}, FG={fg_float:.3f}")
                else:
                    feedback_parts.append(f"❌ Gravity values incorrect: OG={og_float:.3f} (exp 1.052), FG={fg_float:.3f} (exp 1.012)")
            except (ValueError, TypeError):
                feedback_parts.append(f"❌ Gravity values not numeric: OG={og_value}, FG={fg_value}")
        else:
            feedback_parts.append(f"❌ Gravity values missing: OG in D5={og_value}, FG in D16={fg_value}")

        # ===================================================================
        # CRITERION 4: Temperature Data (12 readings)
        # ===================================================================
        temps_present = 0
        temps_valid = 0
        temp_values = []
        
        for row in range(5, 17):  # Rows 5-16 (12 days)
            temp_val = get_cell_value(wb_values, sheet_name, f'C{row}')
            if temp_val is not None:
                temps_present += 1
                try:
                    temp_float = float(temp_val)
                    if 50 <= temp_float <= 85:  # Reasonable temperature range
                        temps_valid += 1
                        temp_values.append(temp_float)
                except (ValueError, TypeError):
                    pass

        if temps_valid >= 12:
            criteria_passed += 1
            feedback_parts.append(f"✅ All 12 temperature readings present (range: {min(temp_values):.0f}-{max(temp_values):.0f}°F)")
        else:
            feedback_parts.append(f"❌ Temperature data incomplete: {temps_valid}/12 valid readings")

        # ===================================================================
        # CRITERION 5: ABV Formula Exists and Calculates Correctly
        # ===================================================================
        abv_value = get_cell_value(wb_values, sheet_name, 'G3')
        abv_formula_exists = False
        
        if wb_formulas:
            abv_formula_exists = is_formula(wb_formulas, sheet_name, 'G3')
            formula_str = get_formula_string(wb_formulas, sheet_name, 'G3')
            if formula_str:
                logger.info(f"ABV formula in G3: {formula_str}")

        expected_abv = (1.052 - 1.012) * 131.25  # = 5.25
        
        if abv_value is not None:
            try:
                abv_float = float(abv_value)
                if abs(abv_float - expected_abv) <= 0.2:
                    if abv_formula_exists:
                        criteria_passed += 1
                        feedback_parts.append(f"✅ ABV formula: {abv_float:.2f}% (formula-based)")
                    else:
                        # Value is correct but might be hardcoded
                        criteria_passed += 0.5
                        feedback_parts.append(f"⚠️ ABV value correct ({abv_float:.2f}%) but formula not detected")
                else:
                    feedback_parts.append(f"❌ ABV calculation wrong: {abv_float:.2f}% (expected ~{expected_abv:.2f}%)")
            except (ValueError, TypeError):
                feedback_parts.append(f"❌ ABV value not numeric: {abv_value}")
        else:
            feedback_parts.append("❌ ABV calculation missing in G3")

        # ===================================================================
        # CRITERION 6: Temperature MIN Formula
        # ===================================================================
        min_temp_value = get_cell_value(wb_values, sheet_name, 'G6')
        min_formula_exists = False
        
        if wb_formulas:
            min_formula_exists = is_formula(wb_formulas, sheet_name, 'G6')

        if min_temp_value is not None and temp_values:
            try:
                min_float = float(min_temp_value)
                expected_min = min(temp_values)
                if abs(min_float - expected_min) <= 1:
                    if min_formula_exists:
                        criteria_passed += 1
                        feedback_parts.append(f"✅ MIN temp formula: {min_float:.0f}°F")
                    else:
                        criteria_passed += 0.5
                        feedback_parts.append(f"⚠️ MIN temp correct ({min_float:.0f}°F) but formula not detected")
                else:
                    feedback_parts.append(f"❌ MIN temp wrong: {min_float:.0f}°F (expected {expected_min:.0f}°F)")
            except (ValueError, TypeError):
                feedback_parts.append(f"❌ MIN temp not numeric: {min_temp_value}")
        else:
            feedback_parts.append("❌ MIN temp missing in G6")

        # ===================================================================
        # CRITERION 7: Temperature MAX Formula
        # ===================================================================
        max_temp_value = get_cell_value(wb_values, sheet_name, 'G7')
        max_formula_exists = False
        
        if wb_formulas:
            max_formula_exists = is_formula(wb_formulas, sheet_name, 'G7')

        if max_temp_value is not None and temp_values:
            try:
                max_float = float(max_temp_value)
                expected_max = max(temp_values)
                if abs(max_float - expected_max) <= 1:
                    if max_formula_exists:
                        criteria_passed += 1
                        feedback_parts.append(f"✅ MAX temp formula: {max_float:.0f}°F")
                    else:
                        criteria_passed += 0.5
                        feedback_parts.append(f"⚠️ MAX temp correct ({max_float:.0f}°F) but formula not detected")
                else:
                    feedback_parts.append(f"❌ MAX temp wrong: {max_float:.0f}°F (expected {expected_max:.0f}°F)")
            except (ValueError, TypeError):
                feedback_parts.append(f"❌ MAX temp not numeric: {max_temp_value}")
        else:
            feedback_parts.append("❌ MAX temp missing in G7")

        # ===================================================================
        # CRITERION 8: Temperature AVERAGE Formula
        # ===================================================================
        avg_temp_value = get_cell_value(wb_values, sheet_name, 'G8')
        avg_formula_exists = False
        
        if wb_formulas:
            avg_formula_exists = is_formula(wb_formulas, sheet_name, 'G8')

        if avg_temp_value is not None and temp_values:
            try:
                avg_float = float(avg_temp_value)
                expected_avg = sum(temp_values) / len(temp_values)
                if abs(avg_float - expected_avg) <= 1:
                    if avg_formula_exists:
                        criteria_passed += 1
                        feedback_parts.append(f"✅ AVG temp formula: {avg_float:.1f}°F")
                    else:
                        criteria_passed += 0.5
                        feedback_parts.append(f"⚠️ AVG temp correct ({avg_float:.1f}°F) but formula not detected")
                else:
                    feedback_parts.append(f"❌ AVG temp wrong: {avg_float:.1f}°F (expected {expected_avg:.1f}°F)")
            except (ValueError, TypeError):
                feedback_parts.append(f"❌ AVG temp not numeric: {avg_temp_value}")
        else:
            feedback_parts.append("❌ AVG temp missing in G8")

        # ===================================================================
        # CRITERION 9: Quality Check Formulas (IF/AND logic)
        # ===================================================================
        abv_check = get_cell_value(wb_values, sheet_name, 'G10')
        temp_check = get_cell_value(wb_values, sheet_name, 'G11')
        
        quality_checks_ok = False
        if abv_check and temp_check:
            abv_check_str = str(abv_check).lower()
            temp_check_str = str(temp_check).lower()
            
            # Should both be "Yes" based on the data
            if 'yes' in abv_check_str and 'yes' in temp_check_str:
                quality_checks_ok = True
                criteria_passed += 1
                feedback_parts.append("✅ Quality checks present: ABV and Temp both OK")
            else:
                feedback_parts.append(f"⚠️ Quality checks present but unexpected values: ABV={abv_check}, Temp={temp_check}")
        else:
            feedback_parts.append(f"❌ Quality checks missing in G10/G11")

        # ===================================================================
        # Calculate Score
        # ===================================================================
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 75

        feedback = " | ".join(feedback_parts)

        logger.info(f"Verification complete: {criteria_passed}/{total_criteria} criteria passed")

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)