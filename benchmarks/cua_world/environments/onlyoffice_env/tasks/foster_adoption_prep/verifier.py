#!/usr/bin/env python3
"""
Verifier for Foster Adoption Prep task
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_foster_adoption_prep(traj, env_info, task_info):
    """
    Verify foster cat adoption readiness spreadsheet.

    Scoring breakdown:
    - 0.15: Column structure (required columns present)
    - 0.30: All 6 cats present
    - 0.40: Data accuracy for each cat (partial credit)
    - 0.10: Evidence of calculation/correct time values
    - 0.05: Presentation quality

    Pass threshold: 0.70 (70%)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/foster_cats_adoption.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_foster_')
    temp_file = os.path.join(temp_dir, "foster_cats.xlsx")

    feedback_parts = []
    score = 0.0

    try:
        # Copy file from container
        copy_from_env(container_path, temp_file)

        if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": "❌ Spreadsheet file not found or empty at expected location"
            }

        # Parse workbook
        wb = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_file, data_only=True)
        except Exception as e:
            logger.error(f"Failed to parse XLSX: {e}")
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"❌ Could not parse spreadsheet file: {str(e)}"
            }

        if not wb:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": "❌ Could not load workbook"
            }

        # Get active sheet data
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(max_row=50, max_col=20, values_only=True):
            data.append(list(row))

        if len(data) < 2:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": "❌ Spreadsheet is empty or has insufficient data"
            }

        # CRITERION 1: Find header row and check column structure (0.15 points)
        header_row_idx = None
        header_row = None

        # Look for header row (first row with multiple non-empty cells containing relevant keywords)
        for idx, row in enumerate(data[:10]):  # Check first 10 rows
            row_text = ' '.join([str(cell).lower() if cell else '' for cell in row])
            keyword_count = sum([
                'name' in row_text,
                'age' in row_text,
                'sex' in row_text or 'gender' in row_text,
                'week' in row_text or 'foster' in row_text or 'time' in row_text,
                'medical' in row_text or 'health' in row_text or 'status' in row_text,
                'behav' in row_text or 'note' in row_text or 'description' in row_text,
                'ready' in row_text or 'adoption' in row_text
            ])
            if keyword_count >= 4:  # At least 4 relevant keywords
                header_row_idx = idx
                header_row = row
                break

        if header_row is None:
            feedback_parts.append("⚠️  No clear header row found with required columns")
        else:
            # Map columns
            col_map = {}
            for idx, cell in enumerate(header_row):
                cell_lower = str(cell).lower() if cell else ''
                if 'name' in cell_lower and 'cat' in cell_lower:
                    col_map['name'] = idx
                elif 'name' in cell_lower and 'name' not in col_map:
                    col_map['name'] = idx
                if 'age' in cell_lower:
                    col_map['age'] = idx
                if 'sex' in cell_lower or 'gender' in cell_lower:
                    col_map['sex'] = idx
                if ('week' in cell_lower or 'time' in cell_lower) and 'foster' in cell_lower:
                    col_map['weeks'] = idx
                elif 'week' in cell_lower and 'weeks' not in col_map:
                    col_map['weeks'] = idx
                if 'medical' in cell_lower or 'health' in cell_lower:
                    col_map['medical'] = idx
                if 'behav' in cell_lower or 'note' in cell_lower or 'description' in cell_lower:
                    col_map['behavior'] = idx
                if 'ready' in cell_lower or 'adoption' in cell_lower or 'status' in cell_lower:
                    col_map['readiness'] = idx

            required_cols = ['name', 'age', 'weeks', 'medical', 'readiness']
            found_cols = sum(1 for col in required_cols if col in col_map)

            if found_cols >= 4:
                feedback_parts.append(f"✅ Column structure present ({found_cols}/5 key columns found)")
                score += 0.15
            elif found_cols >= 3:
                feedback_parts.append(f"⚠️  Weak column structure ({found_cols}/5 key columns found)")
                score += 0.08
            else:
                feedback_parts.append(f"❌ Poor column structure ({found_cols}/5 key columns found)")

        # CRITERION 2 & 3: Check for cat presence and data accuracy
        cat_names = ['whiskers', 'luna', 'mister', 'patches', 'simba', 'shadow']
        
        # Expected data for each cat (for verification)
        cat_criteria = {
            'whiskers': {
                'age_keywords': ['adult', 'mature'],
                'sex_keywords': ['male', 'm'],
                'weeks_range': (7, 9),  # 8 weeks ±1
                'medical_keywords': ['healthy', 'good', 'fine', 'normal', 'none'],
                'readiness_keywords': ['ready', 'now', 'yes', 'available']
            },
            'luna': {
                'age_keywords': ['young', 'kitten', '8 month', 'juvenile'],
                'sex_keywords': ['female', 'f'],
                'weeks_range': (2, 4),  # 3 weeks ±1
                'medical_keywords': ['healthy', 'good', 'fine', 'normal', 'none'],
                'readiness_keywords': ['need', 'time', 'not', 'shy', 'wait']
            },
            'mister': {
                'age_keywords': ['senior', 'old', '12', 'elderly'],
                'sex_keywords': ['male', 'm'],
                'weeks_range': (5, 7),  # 6 weeks ±1
                'medical_keywords': ['special', 'diabetes', 'diabetic', 'insulin', 'needs'],
                'readiness_keywords': ['ready', 'special', 'experienced', 'placement']
            },
            'patches': {
                'age_keywords': ['adult', 'mature'],
                'sex_keywords': ['female', 'f'],
                'weeks_range': (1, 2.5),  # ~1.5 weeks (10 days)
                'medical_keywords': ['treatment', 'sick', 'uri', 'infection', 'pending', 'antibio'],
                'readiness_keywords': ['not', 'no', 'pending', 'wait', 'treatment']
            },
            'simba': {
                'age_keywords': ['kitten', '4 month', 'young', 'baby'],
                'sex_keywords': ['male', 'm'],
                'weeks_range': (7, 9),  # 8 weeks ±1
                'medical_keywords': ['healthy', 'good', 'fine', 'normal', 'healed'],
                'readiness_keywords': ['ready', 'now', 'yes', 'available']
            },
            'shadow': {
                'age_keywords': ['adult', 'mature'],
                'sex_keywords': ['female', 'f'],
                'weeks_range': (4, 6),  # 5 weeks ±1
                'medical_keywords': ['special', 'fiv', 'positive', 'indoor', 'needs'],
                'readiness_keywords': ['ready', 'special', 'placement', 'experienced']
            }
        }

        found_cats = []
        cat_accuracy_scores = {}

        # Search for cats in data rows
        start_row = (header_row_idx + 1) if header_row_idx is not None else 1

        for row_idx, row in enumerate(data[start_row:], start=start_row):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            row_text_full = ' '.join([str(cell).lower() if cell else '' for cell in row])

            for cat in cat_names:
                if cat in row_text_full and cat not in found_cats:
                    found_cats.append(cat)
                    
                    # Evaluate data accuracy for this cat
                    criteria = cat_criteria[cat]
                    checks = {'total': 0, 'passed': 0}

                    # Check age category
                    if 'age' in col_map and col_map['age'] < len(row):
                        age_text = str(row[col_map['age']]).lower() if row[col_map['age']] else ''
                        if any(kw in age_text for kw in criteria['age_keywords']):
                            checks['passed'] += 1
                        checks['total'] += 1

                    # Check sex
                    if 'sex' in col_map and col_map['sex'] < len(row):
                        sex_text = str(row[col_map['sex']]).lower() if row[col_map['sex']] else ''
                        if any(kw in sex_text for kw in criteria['sex_keywords']):
                            checks['passed'] += 1
                        checks['total'] += 1

                    # Check weeks in foster
                    if 'weeks' in col_map and col_map['weeks'] < len(row):
                        weeks_val = row[col_map['weeks']]
                        if weeks_val is not None:
                            try:
                                weeks_num = float(weeks_val)
                                if criteria['weeks_range'][0] <= weeks_num <= criteria['weeks_range'][1]:
                                    checks['passed'] += 1
                                checks['total'] += 1
                            except:
                                pass

                    # Check medical status
                    if 'medical' in col_map and col_map['medical'] < len(row):
                        medical_text = str(row[col_map['medical']]).lower() if row[col_map['medical']] else ''
                        if any(kw in medical_text for kw in criteria['medical_keywords']):
                            checks['passed'] += 1
                        checks['total'] += 1

                    # Check readiness
                    if 'readiness' in col_map and col_map['readiness'] < len(row):
                        readiness_text = str(row[col_map['readiness']]).lower() if row[col_map['readiness']] else ''
                        if any(kw in readiness_text for kw in criteria['readiness_keywords']):
                            checks['passed'] += 1
                        checks['total'] += 1

                    # Calculate accuracy for this cat
                    if checks['total'] > 0:
                        cat_accuracy_scores[cat] = checks['passed'] / checks['total']
                    else:
                        cat_accuracy_scores[cat] = 0.0

        # Score for cat presence (0.30 points)
        cats_found_ratio = len(found_cats) / 6.0
        cats_score = cats_found_ratio * 0.30
        score += cats_score

        if len(found_cats) == 6:
            feedback_parts.append("✅ All 6 cats present in spreadsheet")
        elif len(found_cats) >= 4:
            missing = [c for c in cat_names if c not in found_cats]
            feedback_parts.append(f"⚠️  Found {len(found_cats)}/6 cats (missing: {', '.join(missing)})")
        else:
            missing = [c for c in cat_names if c not in found_cats]
            feedback_parts.append(f"❌ Only {len(found_cats)}/6 cats found (missing: {', '.join(missing)})")

        # Score for data accuracy (0.40 points)
        if cat_accuracy_scores:
            avg_accuracy = sum(cat_accuracy_scores.values()) / len(cat_accuracy_scores)
            accuracy_score = avg_accuracy * 0.40
            score += accuracy_score

            correct_cats = sum(1 for acc in cat_accuracy_scores.values() if acc >= 0.6)
            if correct_cats >= 5:
                feedback_parts.append(f"✅ Cat data mostly accurate ({correct_cats}/6 cats with correct details)")
            elif correct_cats >= 3:
                feedback_parts.append(f"⚠️  Some cat data accurate ({correct_cats}/6 cats with correct details)")
            else:
                feedback_parts.append(f"❌ Cat data largely inaccurate ({correct_cats}/6 cats correct)")
        else:
            feedback_parts.append("❌ No cat data could be validated")

        # CRITERION 4: Check for time calculations (0.10 points)
        has_valid_weeks = False
        if 'weeks' in col_map:
            weeks_values = []
            for row in data[start_row:]:
                if col_map['weeks'] < len(row):
                    val = row[col_map['weeks']]
                    if val is not None and isinstance(val, (int, float)) and 1 <= val <= 10:
                        weeks_values.append(val)
            
            if len(weeks_values) >= 4:  # At least 4 cats have reasonable week values
                has_valid_weeks = True

        if has_valid_weeks:
            feedback_parts.append("✅ Time calculations present (weeks in foster)")
            score += 0.10
        else:
            feedback_parts.append("⚠️  No clear time calculations found")

        # CRITERION 5: Presentation quality (0.05 points)
        quality_checks = 0
        if len(data) >= 7:  # Header + at least 6 cat rows
            quality_checks += 1
        if header_row_idx is not None:
            quality_checks += 1
        if len(found_cats) >= 5:
            quality_checks += 1

        if quality_checks >= 2:
            feedback_parts.append("✅ Good presentation quality")
            score += 0.05
        elif quality_checks == 1:
            score += 0.03

        # Final determination
        passed = score >= 0.70  # 70% threshold

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": round(score, 2),
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"❌ Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)