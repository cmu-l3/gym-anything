#!/usr/bin/env python3
"""
Verifier for Recipe Experiment Tracker task (recipe_experiment_tracker@1)
Checks that the baker's bread experiment data is properly organized in a spreadsheet
"""

import sys
import os
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_recipe_tracker(traj, env_info, task_info):
    """
    Verify that the bread experiment tracker spreadsheet is correctly created.
    
    Requirements:
    1. File exists and is valid XLSX
    2. Header row contains expected columns
    3. At least 3 data rows with experiment information
    4. Numeric data in measurement columns (percentages, rise, time, texture)
    5. Formulas exist in "Total Flour (g)" column
    6. Data values are reasonable for bread experiments
    7. Experiment names are present
    """
    
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}
    
    container_path = "/home/ga/Documents/Spreadsheets/bread_experiments.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_recipe_')
    
    try:
        # Copy and parse the spreadsheet
        success, workbook, error = copy_and_parse_document(
            container_path, 
            copy_from_env, 
            'xlsx'
        )
        
        if not success:
            return {
                "passed": False,
                "score": 0,
                "feedback": f"❌ Could not open spreadsheet: {error}"
            }
        
        # Get the active sheet
        sheet = workbook.active
        sheet_name = sheet.title
        
        # Get all data
        data = get_sheet_data(workbook, sheet_name, max_rows=15, max_cols=10)
        
        if len(data) < 1:
            return {
                "passed": False,
                "score": 0,
                "feedback": "❌ Spreadsheet is empty"
            }
        
        feedback_parts = []
        score = 0.0
        max_score = 7.0
        
        # Check 1: Header row exists and contains expected columns (1.0 points)
        header_row = [str(cell).strip().lower() if cell else "" for cell in data[0]]
        
        expected_headers = [
            "experiment name",
            "bread flour",
            "whole wheat", 
            "rise height",
            "baking time",
            "texture",
            "total flour"
        ]
        
        headers_found = 0
        for expected in expected_headers:
            if any(expected in h for h in header_row):
                headers_found += 1
        
        if headers_found >= 6:
            score += 1.0
            feedback_parts.append(f"✅ Header row complete ({headers_found}/7 columns found)")
        else:
            feedback_parts.append(f"❌ Header row incomplete ({headers_found}/7 expected columns found)")
        
        # Check 2: At least 3 data rows with content (1.0 points)
        data_rows = []
        for row in data[1:]:
            # Skip instruction rows and empty rows
            if row and row[0] and not str(row[0]).startswith("[") and not str(row[0]).lower().startswith("instruction"):
                data_rows.append(row)
        
        if len(data_rows) >= 3:
            score += 1.0
            feedback_parts.append(f"✅ Found {len(data_rows)} experiment entries")
        else:
            feedback_parts.append(f"❌ Only {len(data_rows)} experiment entries (need 3)")
            # If we don't have enough rows, can't check remaining criteria properly
            return {
                "passed": False,
                "score": score / max_score,
                "feedback": " | ".join(feedback_parts)
            }
        
        # Check 3: All experiment entries have names (0.5 points)
        experiments_with_names = 0
        for i, row in enumerate(data_rows[:3], start=1):
            if len(row) > 0 and row[0]:
                name = str(row[0]).strip()
                if name and len(name) > 2 and not name.startswith("["):
                    experiments_with_names += 1
        
        if experiments_with_names >= 3:
            score += 0.5
            feedback_parts.append("✅ All experiments have names")
        else:
            feedback_parts.append(f"⚠️ Only {experiments_with_names}/3 experiments have valid names")
        
        # Check 4: Numeric data in measurement columns (1.5 points)
        numeric_score = 0.0
        total_numeric_checks = 0
        passed_numeric_checks = 0
        
        for i, row in enumerate(data_rows[:3], start=1):
            if len(row) < 6:
                continue
                
            # Check columns: Bread Flour %, Whole Wheat %, Rise Height, Baking Time, Texture
            # Columns indices: 1, 2, 3, 4, 5 (0-indexed)
            for col_idx in [1, 2, 3, 4, 5]:
                total_numeric_checks += 1
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None and isinstance(val, (int, float)):
                        passed_numeric_checks += 1
        
        if total_numeric_checks > 0:
            numeric_score = (passed_numeric_checks / total_numeric_checks) * 1.5
            score += numeric_score
            if numeric_score >= 1.3:
                feedback_parts.append(f"✅ Measurement data is numeric ({passed_numeric_checks}/{total_numeric_checks} cells)")
            else:
                feedback_parts.append(f"⚠️ Some measurement data missing ({passed_numeric_checks}/{total_numeric_checks} numeric cells)")
        else:
            feedback_parts.append("❌ Could not verify numeric data")
        
        # Check 5: Formula exists in Total Flour column (2.0 points)
        # Need to reload without data_only to check formulas
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=temp_dir)
        formula_score = 0.0
        
        try:
            copy_from_env(container_path, temp_file.name)
            
            from openpyxl import load_workbook
            wb_formulas = load_workbook(temp_file.name, data_only=False)
            sheet_formulas = wb_formulas.active
            
            formulas_found = 0
            formulas_correct = 0
            
            # Check rows 2-4 (1-indexed), column G (7)
            for row_idx in range(2, 5):
                cell = sheet_formulas.cell(row=row_idx, column=7)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas_found += 1
                    # Check if formula contains addition and multiplication
                    formula_lower = cell.value.lower()
                    if ('+' in formula_lower or 'sum' in formula_lower) and '*' in formula_lower:
                        formulas_correct += 1
            
            if formulas_found >= 3 and formulas_correct >= 2:
                formula_score = 2.0
                feedback_parts.append(f"✅ Formulas found in Total Flour column ({formulas_found} formulas, {formulas_correct} appear correct)")
            elif formulas_found >= 2:
                formula_score = 1.0
                feedback_parts.append(f"⚠️ Some formulas found ({formulas_found}/3), but may be incomplete")
            elif formulas_found >= 1:
                formula_score = 0.5
                feedback_parts.append(f"⚠️ Only {formulas_found} formula found in Total Flour column")
            else:
                feedback_parts.append("❌ No formulas found in Total Flour column")
            
            score += formula_score
            
        except Exception as e:
            logger.warning(f"Could not check formulas: {e}")
            feedback_parts.append("⚠️ Could not verify formula presence")
        finally:
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
        
        # Check 6: Data values are reasonable for bread experiments (1.0 points)
        reasonable_score = 0.0
        checks_passed = 0
        total_checks = 0
        
        for i, row in enumerate(data_rows[:3], start=1):
            if len(row) < 6:
                continue
            
            # Bread Flour % (should be 0-100)
            if len(row) > 1 and row[1] is not None:
                total_checks += 1
                if isinstance(row[1], (int, float)) and 0 <= row[1] <= 100:
                    checks_passed += 1
            
            # Whole Wheat % (should be 0-100)
            if len(row) > 2 and row[2] is not None:
                total_checks += 1
                if isinstance(row[2], (int, float)) and 0 <= row[2] <= 100:
                    checks_passed += 1
            
            # Rise height (should be 3-30 cm)
            if len(row) > 3 and row[3] is not None:
                total_checks += 1
                if isinstance(row[3], (int, float)) and 3 <= row[3] <= 30:
                    checks_passed += 1
            
            # Baking time (should be 20-90 min)
            if len(row) > 4 and row[4] is not None:
                total_checks += 1
                if isinstance(row[4], (int, float)) and 20 <= row[4] <= 90:
                    checks_passed += 1
            
            # Texture rating (should be 1-5)
            if len(row) > 5 and row[5] is not None:
                total_checks += 1
                if isinstance(row[5], (int, float)) and 1 <= row[5] <= 5:
                    checks_passed += 1
        
        if total_checks > 0:
            reasonable_score = (checks_passed / total_checks) * 1.0
            score += reasonable_score
            if reasonable_score >= 0.8:
                feedback_parts.append(f"✅ Data values are reasonable ({checks_passed}/{total_checks} checks passed)")
            else:
                feedback_parts.append(f"⚠️ Some data values may be unreasonable ({checks_passed}/{total_checks} checks passed)")
        else:
            feedback_parts.append("⚠️ Could not verify data reasonableness")
        
        # Check 7: Total Flour calculated values are reasonable (if formulas present) (0.5 points)
        if len(row) >= 7:
            total_flour_values = []
            for row in data_rows[:3]:
                if len(row) > 6 and row[6] is not None:
                    total_flour_values.append(row[6])
            
            if len(total_flour_values) >= 2:
                reasonable_totals = sum(1 for v in total_flour_values if isinstance(v, (int, float)) and 400 <= v <= 600)
                if reasonable_totals >= 2:
                    score += 0.5
                    feedback_parts.append(f"✅ Total Flour calculations are reasonable")
                else:
                    feedback_parts.append(f"⚠️ Total Flour values may be incorrect")
        
        # Final evaluation
        normalized_score = score / max_score
        passed = score >= 5.0  # Need at least ~71% to pass
        
        feedback = " | ".join(feedback_parts)
        
        return {
            "passed": passed,
            "score": int(normalized_score * 100),
            "feedback": feedback
        }
    
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0,
            "feedback": f"❌ Verification failed: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)
