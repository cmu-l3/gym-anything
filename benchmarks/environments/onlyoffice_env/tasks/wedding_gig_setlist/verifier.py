#!/usr/bin/env python3
"""
Verifier for Wedding Gig Setlist task

This verifier checks:
1. Spreadsheet exists and is valid XLSX
2. Headers are present and correct
3. Song data is intact (8 songs with readiness ratings)
4. Band Avg column uses AVERAGE formulas with correct results
5. Gap to Target column uses subtraction formulas with correct results
6. Priority labels are correctly assigned based on readiness levels
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_wedding_setlist(traj, env_info, task_info):
    """
    Verify that wedding setlist spreadsheet was completed correctly.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/wedding_setlist_readiness.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_setlist_')

    try:
        # Copy and parse the spreadsheet (first with data_only=True for values)
        success, wb_values, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {"passed": False, "score": 0, "feedback": f"Failed to load spreadsheet: {error}"}

        # Also load without data_only to check formulas
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=temp_dir)
        copy_from_env(container_path, temp_file.name)
        
        try:
            from openpyxl import load_workbook
            wb_formulas = load_workbook(temp_file.name, data_only=False)
        except Exception as e:
            logger.error(f"Error loading workbook for formula check: {e}")
            wb_formulas = None

        criteria_passed = 0
        total_criteria = 6
        feedback_parts = []

        sheet_name = wb_values.active.title
        ws_values = wb_values.active
        ws_formulas = wb_formulas.active if wb_formulas else None

        # Criterion 1: Check headers are present
        expected_headers = ["song title", "guitar", "bass", "drums", "vocals", "band avg", "target", "gap", "priority"]
        actual_headers = []
        for col_idx in range(1, 10):
            header_val = ws_values.cell(row=1, column=col_idx).value
            if header_val:
                actual_headers.append(str(header_val).lower())
        
        headers_match = all(any(exp in actual for actual in actual_headers) for exp in expected_headers[:6])
        
        if headers_match:
            criteria_passed += 1
            feedback_parts.append("✅ Headers present and correct")
        else:
            feedback_parts.append(f"❌ Headers missing or incorrect. Found: {actual_headers[:6]}")

        # Criterion 2: Check song data is intact (8 songs with valid readiness values)
        song_count = 0
        data_valid = True
        for row_idx in range(2, 10):
            song_title = ws_values.cell(row=row_idx, column=1).value
            if song_title and not str(song_title).startswith("["):
                song_count += 1
            
            # Check readiness values are numeric and in valid range
            for col_idx in range(2, 6):  # Columns B-E (Guitar, Bass, Drums, Vocals)
                val = ws_values.cell(row=row_idx, column=col_idx).value
                if not isinstance(val, (int, float)) or not (0 <= val <= 100):
                    data_valid = False
                    break
        
        if song_count >= 8 and data_valid:
            criteria_passed += 1
            feedback_parts.append(f"✅ Song data intact ({song_count} songs with valid readiness ratings)")
        else:
            feedback_parts.append(f"❌ Song data incomplete or invalid (found {song_count} songs, data_valid={data_valid})")

        # Criterion 3: Check Band Avg column (F) has AVERAGE formulas with correct results
        avg_formulas_correct = 0
        avg_values_correct = 0
        
        for row_idx in range(2, 10):
            # Check formula if available
            if ws_formulas:
                formula_cell = ws_formulas.cell(row=row_idx, column=6)
                formula_str = str(formula_cell.value) if formula_cell.value else ""
                
                if formula_str.startswith('=') and 'AVERAGE' in formula_str.upper():
                    # Check if it references the correct range (B:E for that row)
                    expected_refs = [f"B{row_idx}", f"E{row_idx}"]
                    if all(ref in formula_str for ref in expected_refs):
                        avg_formulas_correct += 1
            
            # Check calculated value
            band_avg = ws_values.cell(row=row_idx, column=6).value
            guitar = ws_values.cell(row=row_idx, column=2).value
            bass = ws_values.cell(row=row_idx, column=3).value
            drums = ws_values.cell(row=row_idx, column=4).value
            vocals = ws_values.cell(row=row_idx, column=5).value
            
            if all(isinstance(v, (int, float)) for v in [guitar, bass, drums, vocals, band_avg]):
                expected_avg = (guitar + bass + drums + vocals) / 4
                if abs(band_avg - expected_avg) <= 1.0:  # Allow 1% tolerance
                    avg_values_correct += 1

        if avg_formulas_correct >= 6 or avg_values_correct >= 7:
            criteria_passed += 1
            feedback_parts.append(f"✅ Band Avg formulas correct ({avg_formulas_correct} formulas, {avg_values_correct}/8 values correct)")
        else:
            feedback_parts.append(f"❌ Band Avg formulas incomplete ({avg_formulas_correct} formulas, {avg_values_correct}/8 values correct)")

        # Criterion 4: Check Gap to Target column (H) has subtraction formulas with correct results
        gap_formulas_correct = 0
        gap_values_correct = 0
        
        for row_idx in range(2, 10):
            # Check formula if available
            if ws_formulas:
                formula_cell = ws_formulas.cell(row=row_idx, column=8)
                formula_str = str(formula_cell.value) if formula_cell.value else ""
                
                # Check for subtraction formula (G - F)
                if formula_str.startswith('='):
                    # Accept various formats: =G2-F2, =G2 - F2, etc.
                    if f"G{row_idx}" in formula_str and f"F{row_idx}" in formula_str and '-' in formula_str:
                        gap_formulas_correct += 1
            
            # Check calculated value
            gap = ws_values.cell(row=row_idx, column=8).value
            target = ws_values.cell(row=row_idx, column=7).value
            band_avg = ws_values.cell(row=row_idx, column=6).value
            
            if all(isinstance(v, (int, float)) for v in [gap, target, band_avg]):
                expected_gap = target - band_avg
                if abs(gap - expected_gap) <= 1.0:  # Allow 1 tolerance
                    gap_values_correct += 1

        if gap_formulas_correct >= 6 or gap_values_correct >= 7:
            criteria_passed += 1
            feedback_parts.append(f"✅ Gap formulas correct ({gap_formulas_correct} formulas, {gap_values_correct}/8 values correct)")
        else:
            feedback_parts.append(f"❌ Gap formulas incomplete ({gap_formulas_correct} formulas, {gap_values_correct}/8 values correct)")

        # Criterion 5: Check Priority column (I) has correct labels based on Band Avg
        priority_correct = 0
        
        for row_idx in range(2, 10):
            band_avg = ws_values.cell(row=row_idx, column=6).value
            priority = ws_values.cell(row=row_idx, column=9).value
            
            if not isinstance(band_avg, (int, float)):
                continue
            
            priority_str = str(priority).upper() if priority else ""
            
            # Determine expected priority
            if band_avg < 50:
                expected = "HIGH"
            elif band_avg < 80:
                expected = "MEDIUM"
            else:
                expected = "LOW"
            
            # Check if priority matches (flexible matching)
            if expected in priority_str:
                priority_correct += 1

        if priority_correct >= 7:
            criteria_passed += 1
            feedback_parts.append(f"✅ Priority labels correct ({priority_correct}/8 songs)")
        else:
            feedback_parts.append(f"❌ Priority labels incorrect or missing ({priority_correct}/8 songs)")

        # Criterion 6: Check Target values are all 100
        targets_correct = all(
            ws_values.cell(row=row_idx, column=7).value == 100
            for row_idx in range(2, 10)
        )
        
        if targets_correct:
            criteria_passed += 1
            feedback_parts.append("✅ Target values correct (all 100%)")
        else:
            feedback_parts.append("❌ Target values modified or incorrect")

        # Calculate score
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 70  # Pass threshold is 70%

        feedback = " | ".join(feedback_parts)

        # Add summary feedback
        if passed:
            summary = f"✅ Task completed successfully! Band readiness tracker is functional. "
        else:
            summary = f"❌ Task incomplete. Missing {total_criteria - criteria_passed} criteria. "
        
        feedback = summary + feedback

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)
