#!/usr/bin/env python3
"""
Verifier for Music Practice Log task (music_practice_log@1)

Verifies that a piano practice log spreadsheet was created correctly with:
- Proper column headers
- 5 practice session entries
- Correct duration values
- SUM formula calculating total practice time
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_music_practice_log(traj, env_info, task_info):
    """
    Verify that the music practice log spreadsheet is correctly structured.
    
    Scoring criteria (100 points total):
    1. File exists and is valid (10 points)
    2. Column headers present (15 points)
    3. Five data rows present (20 points)
    4. Duration values correct (20 points)
    5. SUM formula exists in B8 (20 points)
    6. SUM formula result correct (15 points)
    
    Pass threshold: 75 points
    
    Returns:
        dict: {"passed": bool, "score": float, "feedback": str}
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/practice_log.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_practice_')
    
    score = 0
    max_score = 100
    feedback_parts = []

    try:
        # Parse the workbook with data_only=True to get calculated values
        success, wb, error = copy_and_parse_document(
            container_path, copy_from_env, file_format='xlsx'
        )

        if not success:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"❌ Could not open practice log file: {error}"
            }

        # Criterion 1: File exists and is valid (10 points)
        # Already passed if we got here
        score += 10
        feedback_parts.append("✅ File exists and is valid XLSX (10/10)")

        # Get the active sheet
        sheet = wb.active
        sheet_name = sheet.title

        # Criterion 2: Check column headers (15 points)
        # Headers should be in Row 1
        header_keywords = ['date', 'duration', 'piece', 'technique', 'tempo', 'note']
        headers_found = 0
        
        # Check first 6 columns for headers
        for col_idx in range(1, 8):  # Check A1 through G1 (in case of extra columns)
            try:
                cell_value = str(sheet.cell(1, col_idx).value or '').lower()
                if any(keyword in cell_value for keyword in header_keywords):
                    headers_found += 1
            except:
                pass
        
        # Need at least 5 out of 6 headers for full points
        if headers_found >= 6:
            header_score = 15
            feedback_parts.append(f"✅ All column headers present (15/15)")
        elif headers_found >= 5:
            header_score = 12
            feedback_parts.append(f"✅ Most column headers present ({header_score}/15)")
        elif headers_found >= 4:
            header_score = 8
            feedback_parts.append(f"⚠️ Some column headers present ({header_score}/15)")
        else:
            header_score = 0
            feedback_parts.append(f"❌ Column headers missing or incomplete (0/15)")
        
        score += header_score

        # Criterion 3: Check five data rows present (20 points)
        # Rows 2-6 should contain practice session data
        data_rows_found = 0
        
        for row_idx in range(2, 7):  # Rows 2-6
            # Check if row has at least some data (first few columns)
            has_data = False
            for col_idx in range(1, 4):  # Check first 3 columns
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            
            if has_data:
                data_rows_found += 1
        
        data_row_score = int((data_rows_found / 5) * 20)
        score += data_row_score
        
        if data_rows_found == 5:
            feedback_parts.append(f"✅ Five practice sessions entered (20/20)")
        elif data_rows_found >= 3:
            feedback_parts.append(f"⚠️ {data_rows_found}/5 practice sessions found ({data_row_score}/20)")
        else:
            feedback_parts.append(f"❌ Only {data_rows_found}/5 practice sessions found ({data_row_score}/20)")

        # Criterion 4: Check duration values (20 points)
        # Expected durations in column B (or wherever Duration column is)
        expected_durations = [45, 60, 30, 15, 90]
        
        # Try to find the duration column (usually B, but could vary)
        duration_col_idx = 2  # Default to column B
        
        # Check if column B has numeric values that look like durations
        b_values = [sheet.cell(row_idx, 2).value for row_idx in range(2, 7)]
        if all(isinstance(v, (int, float)) or v is None for v in b_values):
            duration_col_idx = 2
        
        actual_durations = []
        duration_matches = 0
        
        for i, row_idx in enumerate(range(2, 7)):
            val = sheet.cell(row_idx, duration_col_idx).value
            actual_durations.append(val)
            
            if val is not None and isinstance(val, (int, float)):
                expected = expected_durations[i]
                # Allow ±2 tolerance for typos
                if abs(val - expected) <= 2:
                    duration_matches += 1
        
        duration_score = int((duration_matches / 5) * 20)
        score += duration_score
        
        if duration_matches >= 5:
            feedback_parts.append(f"✅ Duration values correct (20/20)")
        elif duration_matches >= 4:
            feedback_parts.append(f"✅ Most duration values correct ({duration_score}/20)")
        elif duration_matches >= 3:
            feedback_parts.append(f"⚠️ Duration values: {duration_matches}/5 correct ({duration_score}/20)")
        else:
            feedback_parts.append(f"❌ Duration values incorrect: found {actual_durations}, expected {expected_durations} ({duration_score}/20)")

        # Criterion 5: Check for SUM formula (20 points)
        # Formula should be in B8 (or row 8 of duration column)
        b8_cell = sheet.cell(8, duration_col_idx)
        
        # To detect formulas, we need to check if it's a formula type
        # In openpyxl with data_only=True, we can't see the formula directly
        # We need to load again with data_only=False
        has_formula = False
        formula_score = 0
        
        try:
            # Load workbook again without data_only to check for formulas
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=temp_dir)
            copy_from_env(container_path, temp_file.name)
            
            from openpyxl import load_workbook
            wb_formula_check = load_workbook(temp_file.name, data_only=False)
            sheet_formula_check = wb_formula_check.active
            
            b8_cell_check = sheet_formula_check.cell(8, duration_col_idx)
            
            # Check if cell contains a formula
            if b8_cell_check.value and isinstance(b8_cell_check.value, str):
                if b8_cell_check.value.startswith('='):
                    has_formula = True
                    # Check if it's a SUM-like formula
                    formula_text = b8_cell_check.value.upper()
                    if 'SUM' in formula_text or '+' in formula_text:
                        formula_score = 20
                        feedback_parts.append(f"✅ SUM formula exists in row 8 (20/20)")
                    else:
                        formula_score = 10
                        feedback_parts.append(f"⚠️ Formula exists but not clearly a SUM (10/20)")
            
            os.unlink(temp_file.name)
            
        except Exception as e:
            logger.warning(f"Could not check for formula: {e}")
            # Fallback: if the value looks calculated (exactly 240), give partial credit
            b8_value = b8_cell.value
            if b8_value == sum(expected_durations):
                formula_score = 10  # Partial credit, might be hardcoded
                feedback_parts.append(f"⚠️ Total value correct but could not verify formula (10/20)")
        
        if not has_formula and formula_score == 0:
            feedback_parts.append(f"❌ No formula found in row 8 - expected =SUM(...) (0/20)")
        
        score += formula_score

        # Criterion 6: Check SUM result (15 points)
        # The sum should be 240 (45+60+30+15+90)
        sum_value = b8_cell.value
        expected_sum = sum(expected_durations)  # 240
        
        sum_correct = False
        sum_score = 0
        
        if sum_value is not None and isinstance(sum_value, (int, float)):
            # Allow small tolerance for rounding or if durations were slightly off
            if abs(sum_value - expected_sum) <= 5:
                sum_correct = True
                sum_score = 15
                feedback_parts.append(f"✅ Total calculated correctly: {sum_value} minutes (15/15)")
            else:
                # Partial credit if it's summing something
                if sum_value > 100 and sum_value < 400:
                    sum_score = 5
                    feedback_parts.append(f"⚠️ Total calculation: got {sum_value}, expected ~{expected_sum} (5/15)")
                else:
                    feedback_parts.append(f"❌ Total calculation wrong: got {sum_value}, expected {expected_sum} (0/15)")
        else:
            feedback_parts.append(f"❌ Total value missing or invalid: {sum_value} (0/15)")
        
        score += sum_score

        # Final assessment
        passed = score >= 75
        normalized_score = score / max_score
        
        feedback = " | ".join(feedback_parts)
        feedback += f" || TOTAL SCORE: {score}/{max_score}"
        
        logger.info(f"Verification complete: passed={passed}, score={score}/{max_score}")
        
        return {
            "passed": passed,
            "score": normalized_score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"❌ Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)
