#!/usr/bin/env python3
"""
Verifier for garden_planting_schedule@1
Checks that the agent created a structured planting schedule with calculated dates
"""

import sys
import os
import logging
import tempfile
from datetime import datetime
from typing import Dict, Any, List, Tuple

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def find_header_row(data: List[List], max_search_rows: int = 10) -> Tuple[int, List[str]]:
    """
    Find the row containing column headers.
    Returns: (row_index, normalized_headers_list)
    """
    expected_keywords = ['crop', 'frost', 'date', 'spacing', 'companion']
    
    for row_idx in range(min(max_search_rows, len(data))):
        row = data[row_idx]
        if not row:
            continue
            
        # Normalize row values to lowercase strings
        normalized_row = [str(cell).lower().strip() if cell else "" for cell in row]
        
        # Count how many expected keywords appear in this row
        keyword_matches = sum(1 for keyword in expected_keywords 
                            if any(keyword in cell for cell in normalized_row))
        
        # If we find at least 3 of our expected keywords, consider it the header row
        if keyword_matches >= 3:
            return row_idx, normalized_row
    
    return -1, []


def find_column_indices(headers: List[str]) -> Dict[str, int]:
    """
    Map expected column names to their indices in the header row.
    Returns dict with keys: crop, days_from_frost, planting_date, spacing, companions
    """
    column_map = {}
    
    for idx, header in enumerate(headers):
        if 'crop' in header and 'companion' not in header:
            column_map['crop'] = idx
        elif 'day' in header and 'frost' in header:
            column_map['days_from_frost'] = idx
        elif ('plant' in header or 'date' in header) and 'frost' not in header:
            column_map['planting_date'] = idx
        elif 'spacing' in header or 'space' in header:
            column_map['spacing'] = idx
        elif 'companion' in header or 'pair' in header:
            column_map['companions'] = idx
    
    return column_map


def check_formula_references_cell(formula_str: str, target_cell: str) -> bool:
    """
    Check if a formula string references a specific cell (case-insensitive).
    Handles variations like C1, $C$1, c1, etc.
    """
    if not formula_str or not isinstance(formula_str, str):
        return False
    
    formula_upper = formula_str.upper()
    target_upper = target_cell.upper()
    
    # Check for various reference formats
    variations = [
        target_upper,           # C1
        f'${target_upper}',     # $C1
        f'{target_upper[:1]}${target_upper[1:]}',  # C$1
        f'${target_upper[:1]}${target_upper[1:]}'  # $C$1
    ]
    
    return any(var in formula_upper for var in variations)


def verify_garden_planting_schedule(traj, env_info, task_info):
    """
    Verify that agent created a proper garden planting schedule.
    
    Criteria:
    1. Column headers are present (flexible row detection)
    2. At least 6 crops are listed
    3. At least 3 planting date formulas exist that reference C1 (frost date)
    4. Spacing information is filled for at least 70% of crops
    5. Companion planting info is present for at least 3 crops
    """
    
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}
    
    SPREADSHEET_PATH = "/home/ga/Documents/Spreadsheets/planting_schedule.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_garden_')
    temp_file_path = None
    
    feedback_parts = []
    score = 0
    max_score = 100
    
    try:
        # Copy file to temp location
        temp_file_path = os.path.join(temp_dir, 'planting_schedule.xlsx')
        copy_from_env(SPREADSHEET_PATH, temp_file_path)
        
        if not os.path.exists(temp_file_path) or os.path.getsize(temp_file_path) == 0:
            return {
                "passed": False,
                "score": 0,
                "feedback": f"File not found or empty: {SPREADSHEET_PATH}"
            }
        
        # Parse the spreadsheet with data_only=True first for values
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "passed": False,
                "score": 0,
                "feedback": "openpyxl library not available"
            }
        
        wb = load_workbook(temp_file_path, data_only=True)
        sheet = wb.active
        
        # Get data from the sheet
        data = []
        for row in sheet.iter_rows(max_row=25, max_col=10, values_only=True):
            data.append(list(row))
        
        # Criterion 1: Find and validate headers
        header_row_idx, headers = find_header_row(data)
        
        if header_row_idx >= 0:
            feedback_parts.append(f"✅ Column headers found in row {header_row_idx + 1}")
            score += 20
            
            # Map columns
            col_map = find_column_indices(headers)
            required_cols = ['crop', 'planting_date']
            found_required = all(col in col_map for col in required_cols)
            
            if not found_required:
                feedback_parts.append(f"⚠️ Some expected columns missing")
        else:
            feedback_parts.append("❌ Column headers not found")
            return {
                "passed": False,
                "score": score,
                "feedback": " | ".join(feedback_parts)
            }
        
        # Criterion 2: Count crops (rows with data after headers)
        col_map = find_column_indices(headers)
        crop_col = col_map.get('crop', 0)
        
        crops = []
        crop_rows = []  # Track which rows have crops
        data_start_row = header_row_idx + 1
        
        for row_idx in range(data_start_row, min(data_start_row + 15, len(data))):
            row = data[row_idx]
            if row and len(row) > crop_col and row[crop_col]:
                crop_name = str(row[crop_col]).strip().lower()
                # Filter out non-crop entries
                if len(crop_name) > 2 and crop_name not in ['crop', 'instructions']:
                    crops.append(crop_name)
                    crop_rows.append(row_idx)
        
        num_crops = len(crops)
        if num_crops >= 6:
            feedback_parts.append(f"✅ {num_crops} crops listed")
            score += 25
        elif num_crops >= 4:
            feedback_parts.append(f"⚠️ Only {num_crops} crops listed (expected 6+)")
            score += 12
        else:
            feedback_parts.append(f"❌ Only {num_crops} crops listed (expected 6+)")
        
        if num_crops == 0:
            feedback_parts.append("❌ No crops found in spreadsheet")
            return {
                "passed": False,
                "score": score,
                "feedback": " | ".join(feedback_parts)
            }
        
        # Criterion 3: Check for date formulas that reference C1
        # Need to reload without data_only to see formulas
        wb_formula = load_workbook(temp_file_path, data_only=False)
        sheet_formula = wb_formula.active
        
        date_col = col_map.get('planting_date', 2)  # Default to column C (index 2)
        formula_count = 0
        formula_rows_found = []
        
        for row_idx in crop_rows:
            excel_row = row_idx + 1  # Convert to 1-based Excel row
            excel_col = date_col + 1  # Convert to 1-based Excel column
            
            cell = sheet_formula.cell(row=excel_row, column=excel_col)
            
            if cell.value:
                # Check if it's a formula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    # Check if formula references C1
                    if check_formula_references_cell(cell.value, 'C1'):
                        formula_count += 1
                        formula_rows_found.append(excel_row)
                # Also check for date values that might be calculated
                elif isinstance(cell.value, datetime):
                    # If we have a date value, check if it was calculated from a formula
                    # (it might have been evaluated already)
                    formula_count += 0.5  # Partial credit
        
        if formula_count >= 3:
            feedback_parts.append(f"✅ {int(formula_count)} planting date formulas reference frost date")
            score += 30
        elif formula_count >= 2:
            feedback_parts.append(f"⚠️ Only {int(formula_count)} formulas reference frost date (expected 3+)")
            score += 15
        else:
            feedback_parts.append(f"❌ Few or no date formulas reference frost date ({int(formula_count)} found)")
        
        # Criterion 4: Check spacing information
        spacing_col = col_map.get('spacing')
        spacing_count = 0
        
        if spacing_col is not None:
            for row_idx in crop_rows:
                row = data[row_idx]
                if len(row) > spacing_col and row[spacing_col]:
                    spacing_value = str(row[spacing_col]).strip()
                    if len(spacing_value) > 0 and spacing_value.lower() not in ['none', 'n/a', '-']:
                        spacing_count += 1
        
        spacing_ratio = spacing_count / max(num_crops, 1)
        if spacing_ratio >= 0.7:
            feedback_parts.append(f"✅ Spacing info for {spacing_count}/{num_crops} crops")
            score += 15
        elif spacing_ratio >= 0.4:
            feedback_parts.append(f"⚠️ Spacing info incomplete ({spacing_count}/{num_crops})")
            score += 8
        else:
            feedback_parts.append(f"❌ Spacing info mostly missing ({spacing_count}/{num_crops})")
        
        # Criterion 5: Check companion planting info
        companion_col = col_map.get('companions')
        companion_count = 0
        
        if companion_col is not None:
            for row_idx in crop_rows:
                row = data[row_idx]
                if len(row) > companion_col and row[companion_col]:
                    companion_value = str(row[companion_col]).strip()
                    if len(companion_value) > 2 and companion_value.lower() not in ['none', 'n/a', '-']:
                        companion_count += 1
        
        if companion_count >= 3:
            feedback_parts.append(f"✅ Companion info for {companion_count} crops")
            score += 10
        elif companion_count >= 2:
            feedback_parts.append(f"⚠️ Limited companion info ({companion_count} crops)")
            score += 5
        else:
            feedback_parts.append(f"❌ Companion info mostly missing ({companion_count} crops)")
        
        # Compile feedback
        feedback = " | ".join(feedback_parts)
        passed = score >= 70  # Pass threshold
        
        return {
            "passed": passed,
            "score": score / max_score,
            "feedback": feedback
        }
        
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0,
            "feedback": f"Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)


# Entry point for gym-anything
def verify_task(traj, env_info, task_info):
    return verify_garden_planting_schedule(traj, env_info, task_info)