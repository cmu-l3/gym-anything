#!/usr/bin/env python3
"""
Verifier for Crowdfunding Calculator task
Checks reward tier structure, scenario data, financial formulas, and accuracy
"""

import sys
import os
import logging
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import cleanup_temp_dir

# Direct imports for spreadsheet parsing
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_crowdfund_calculator(traj, env_info, task_info):
    """
    Verify crowdfunding calculator spreadsheet.
    
    Checks:
    1. Reward tier structure with correct data and formulas (30 points)
    2. Scenario backer count data (20 points)
    3. Financial calculation formulas (35 points)
    4. Accuracy of calculated values (15 points)
    
    Pass threshold: 75/100 points
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    if load_workbook is None:
        return {"passed": False, "score": 0, "feedback": "openpyxl not installed"}

    container_path = "/home/ga/Documents/Spreadsheets/crowdfund_calculator.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_crowdfund_')
    temp_file = None

    try:
        # Copy file from container
        temp_file = os.path.join(temp_dir, 'crowdfund_calculator.xlsx')
        copy_from_env(container_path, temp_file)

        if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
            return {"passed": False, "score": 0, "feedback": f"File not found or empty: {container_path}"}

        # Load workbook twice: once for formulas, once for values
        wb_formulas = load_workbook(temp_file, data_only=False)
        wb_values = load_workbook(temp_file, data_only=True)
        
        ws_formulas = wb_formulas.active
        ws_values = wb_values.active

        score = 0
        feedback_parts = []

        # ===== CRITERION 1: Reward Tier Structure (30 points) =====
        criterion1_score = 0
        criterion1_max = 30
        
        # Expected tier data
        expected_tiers = {
            2: {"name": "Digital Thank You", "pledge": 5, "material": 0, "shipping": 0},
            3: {"name": "Bookplate Pack", "pledge": 15, "material": 3, "shipping": 2},
            4: {"name": "Graphic Novel Bundle", "pledge": 40, "material": 18, "shipping": 8},
            5: {"name": "Class Visit + Books", "pledge": 150, "material": 50, "shipping": 12},
            6: {"name": "Full Library Sponsor", "pledge": 500, "material": 100, "shipping": 0}
        }
        
        tier_data_correct = 0
        tier_formulas_correct = 0
        
        for row_num, expected in expected_tiers.items():
            # Check tier name (column A)
            tier_name = ws_values[f'A{row_num}'].value
            if tier_name and expected["name"].lower() in str(tier_name).lower():
                tier_data_correct += 1
            
            # Check pledge amount (column B)
            pledge = ws_values[f'B{row_num}'].value
            if pledge and abs(float(pledge) - expected["pledge"]) < 0.5:
                tier_data_correct += 1
            
            # Check material cost (column C)
            material = ws_values[f'C{row_num}'].value
            if material is not None and abs(float(material) - expected["material"]) < 0.5:
                tier_data_correct += 1
            
            # Check shipping cost (column D)
            shipping = ws_values[f'D{row_num}'].value
            if shipping is not None and abs(float(shipping) - expected["shipping"]) < 0.5:
                tier_data_correct += 1
            
            # Check Net Per Backer formula (column E)
            net_cell = ws_formulas[f'E{row_num}']
            net_value = ws_values[f'E{row_num}'].value
            expected_net = expected["pledge"] - expected["material"] - expected["shipping"]
            
            # Check if it's a formula
            is_formula = (net_cell.data_type == 'f' or 
                         (isinstance(net_cell.value, str) and net_cell.value.startswith('=')))
            
            # Check if value is correct
            value_correct = (net_value is not None and 
                           abs(float(net_value) - expected_net) < 0.5)
            
            if is_formula and value_correct:
                tier_formulas_correct += 1

        # Tier data: 20 data points (5 tiers × 4 values), each worth 0.6 points = 12 points
        criterion1_score += min(12, int((tier_data_correct / 20) * 12))
        
        # Net formulas: 5 formulas, each worth 3.6 points = 18 points
        criterion1_score += min(18, int((tier_formulas_correct / 5) * 18))
        
        if tier_data_correct >= 18:
            feedback_parts.append(f"✅ Tier data: {tier_data_correct}/20 correct")
        else:
            feedback_parts.append(f"❌ Tier data: {tier_data_correct}/20 correct (need 18+)")
        
        if tier_formulas_correct >= 4:
            feedback_parts.append(f"✅ Net formulas: {tier_formulas_correct}/5 correct")
        else:
            feedback_parts.append(f"❌ Net formulas: {tier_formulas_correct}/5 correct (need 4+)")

        score += criterion1_score

        # ===== CRITERION 2: Scenario Data (20 points) =====
        criterion2_score = 0
        criterion2_max = 20
        
        expected_scenarios = {
            "conservative": [20, 15, 8, 2, 1],
            "optimistic": [40, 25, 15, 4, 2]
        }
        
        scenario_correct = 0
        
        # Check Conservative counts (column B, rows 10-14)
        for i, expected_count in enumerate(expected_scenarios["conservative"]):
            row = 10 + i
            value = ws_values[f'B{row}'].value
            if value and abs(float(value) - expected_count) < 0.5:
                scenario_correct += 1
        
        # Check Optimistic counts (column C, rows 10-14)
        for i, expected_count in enumerate(expected_scenarios["optimistic"]):
            row = 10 + i
            value = ws_values[f'C{row}'].value
            if value and abs(float(value) - expected_count) < 0.5:
                scenario_correct += 1
        
        # 10 data points, each worth 2 points
        criterion2_score = min(20, int((scenario_correct / 10) * 20))
        
        if scenario_correct >= 9:
            feedback_parts.append(f"✅ Scenario data: {scenario_correct}/10 correct")
        else:
            feedback_parts.append(f"❌ Scenario data: {scenario_correct}/10 correct (need 9+)")
        
        score += criterion2_score

        # ===== CRITERION 3: Financial Calculations (35 points) =====
        criterion3_score = 0
        criterion3_max = 35
        
        financial_formulas_correct = 0
        financial_cells = ['B18', 'C18', 'B19', 'C19', 'B20', 'C20', 'B22', 'C22', 
                          'B23', 'C23', 'B24', 'C24', 'B27']
        
        for cell_ref in financial_cells:
            cell_formula = ws_formulas[cell_ref]
            is_formula = (cell_formula.data_type == 'f' or 
                         (isinstance(cell_formula.value, str) and 
                          cell_formula.value.startswith('=')))
            
            # Exception: B21 and C21 (Campaign Goal) can be hardcoded as 2500
            if cell_ref in ['B21', 'C21']:
                continue
            
            if is_formula:
                financial_formulas_correct += 1
        
        # 13 formula cells, worth ~2.7 points each = 35 points
        criterion3_score = min(35, int((financial_formulas_correct / 13) * 35))
        
        if financial_formulas_correct >= 11:
            feedback_parts.append(f"✅ Financial formulas: {financial_formulas_correct}/13 present")
        else:
            feedback_parts.append(f"❌ Financial formulas: {financial_formulas_correct}/13 present (need 11+)")
        
        score += criterion3_score

        # ===== CRITERION 4: Accuracy & Realism (15 points) =====
        criterion4_score = 0
        criterion4_max = 15
        
        accuracy_checks = 0
        
        # Check Conservative Net Funding (should be ~1485)
        conservative_net = ws_values['B20'].value
        if conservative_net and abs(float(conservative_net) - 1485) <= 50:
            accuracy_checks += 1
            feedback_parts.append(f"✅ Conservative net funding: ${conservative_net:.0f} (expected ~$1,485)")
        else:
            feedback_parts.append(f"❌ Conservative net funding: ${conservative_net if conservative_net else 0:.0f} (expected ~$1,485)")
        
        # Check Optimistic Net Funding (should be ~3290)
        optimistic_net = ws_values['C20'].value
        if optimistic_net and abs(float(optimistic_net) - 3290) <= 50:
            accuracy_checks += 1
            feedback_parts.append(f"✅ Optimistic net funding: ${optimistic_net:.0f} (expected ~$3,290)")
        else:
            feedback_parts.append(f"❌ Optimistic net funding: ${optimistic_net if optimistic_net else 0:.0f} (expected ~$3,290)")
        
        # Check Goal Achieved? for Conservative (should be "NO")
        conservative_achieved = ws_values['B22'].value
        if conservative_achieved and str(conservative_achieved).strip().upper() == "NO":
            accuracy_checks += 1
            feedback_parts.append("✅ Conservative goal status: NO")
        else:
            feedback_parts.append(f"❌ Conservative goal status: {conservative_achieved} (expected NO)")
        
        # Check Goal Achieved? for Optimistic (should be "YES")
        optimistic_achieved = ws_values['C22'].value
        if optimistic_achieved and str(optimistic_achieved).strip().upper() == "YES":
            accuracy_checks += 1
            feedback_parts.append("✅ Optimistic goal status: YES")
        else:
            feedback_parts.append(f"❌ Optimistic goal status: {optimistic_achieved} (expected YES)")
        
        # 4 accuracy checks, each worth 3.75 points
        criterion4_score = int((accuracy_checks / 4) * 15)
        score += criterion4_score

        # ===== FINAL EVALUATION =====
        passed = score >= 75

        feedback = " | ".join(feedback_parts)
        
        logger.info(f"Verification complete: Score={score}/100, Passed={passed}")

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        if temp_file and os.path.exists(temp_file):
            try:
                os.unlink(temp_file)
            except:
                pass
        cleanup_temp_dir(temp_dir)