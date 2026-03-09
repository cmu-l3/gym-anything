#!/usr/bin/env python3
"""
Verifier for Internet Troubleshooting Log task

Verifies that the user created a systematic troubleshooting log for documenting
intermittent internet connectivity issues, suitable for sending to ISP support.
"""

import sys
import os
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    parse_xlsx_file,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_internet_troubleshooting_log(traj, env_info, task_info):
    """
    Verify the internet troubleshooting log spreadsheet.
    
    Scoring breakdown:
    1. File exists and is valid XLSX (10 points)
    2. Correct column headers in row 1 (20 points)
    3. Sufficient incident data - at least 8 rows (15 points)
    4. Realistic data values (15 points)
    5. Summary statistics present (20 points)
    6. Correct formulas in summary (15 points)
    7. Conditional formatting applied (5 points)
    
    Total: 100 points
    Passing threshold: 75 points
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/internet_log.xlsx"
    temp_file = None
    
    score = 0
    max_score = 100
    feedback_parts = []
    
    try:
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        # 1. File exists and valid (10 points)
        try:
            copy_from_env(container_path, temp_file.name)
        except Exception as e:
            feedback_parts.append(f"❌ File not found: {container_path}")
            return {"passed": False, "score": 0.0, "feedback": " | ".join(feedback_parts)}
        
        if not os.path.exists(temp_file.name) or os.path.getsize(temp_file.name) == 0:
            feedback_parts.append("❌ File is empty or doesn't exist")
            return {"passed": False, "score": 0.0, "feedback": " | ".join(feedback_parts)}
        
        wb = parse_xlsx_file(temp_file.name)
        if wb is None:
            feedback_parts.append("❌ Could not parse XLSX file")
            return {"passed": False, "score": 0.0, "feedback": " | ".join(feedback_parts)}
        
        score += 10
        feedback_parts.append("✅ Valid XLSX file (10/10)")
        
        sheet = wb.active
        
        # 2. Check column headers (20 points)
        required_headers = [
            'date', 'time', 'issue type', 'download speed', 'upload speed',
            'expected speed', 'duration', 'devices affected', 'activity', 'troubleshooting'
        ]
        
        actual_headers = []
        for col in range(1, 11):  # A through J
            cell_value = sheet.cell(1, col).value
            if cell_value:
                actual_headers.append(str(cell_value).lower().strip())
            else:
                actual_headers.append('')
        
        headers_correct = 0
        for req in required_headers:
            for actual in actual_headers:
                if req in actual:
                    headers_correct += 1
                    break
        
        header_score = int((headers_correct / len(required_headers)) * 20)
        score += header_score
        if header_score >= 15:
            feedback_parts.append(f"✅ Column headers ({header_score}/20) - found {headers_correct}/{len(required_headers)} required headers")
        else:
            feedback_parts.append(f"⚠️ Column headers ({header_score}/20) - only found {headers_correct}/{len(required_headers)} required headers")
        
        # 3. Sufficient incident data (15 points)
        data_rows = 0
        issue_types = set()
        
        for row in range(2, 20):  # Check rows 2-19 for data
            issue_val = sheet.cell(row, 3).value  # Column C: Issue Type
            if issue_val and str(issue_val).strip():
                data_rows += 1
                issue_types.add(str(issue_val).lower().strip())
        
        data_score = 0
        if data_rows >= 8:
            data_score += 10
            feedback_parts.append(f"✅ Sufficient data rows (10/10): {data_rows} incidents logged")
        else:
            partial = max(0, int((data_rows / 8) * 10))
            data_score += partial
            feedback_parts.append(f"⚠️ Insufficient data rows ({partial}/10): only {data_rows} incidents (need 8+)")
        
        if len(issue_types) >= 3:
            data_score += 5
            feedback_parts.append(f"✅ Good variety (5/5): {len(issue_types)} different issue types")
        else:
            feedback_parts.append(f"⚠️ Limited variety (0/5): only {len(issue_types)} issue types")
        
        score += data_score
        
        # 4. Realistic data values (15 points)
        realistic_score = 0
        download_speeds = []
        durations = []
        expected_speeds = []
        
        for row in range(2, 2 + data_rows):
            dl_speed = sheet.cell(row, 4).value  # Column D
            expected_speed = sheet.cell(row, 6).value  # Column F
            duration = sheet.cell(row, 7).value  # Column G
            
            if dl_speed is not None and isinstance(dl_speed, (int, float)):
                download_speeds.append(float(dl_speed))
            if duration is not None and isinstance(duration, (int, float)):
                durations.append(float(duration))
            if expected_speed is not None and isinstance(expected_speed, (int, float)):
                expected_speeds.append(float(expected_speed))
        
        # Check download speeds are reasonable
        if download_speeds and all(0 <= s <= 600 for s in download_speeds):
            realistic_score += 5
        
        # Check durations are reasonable
        if durations and all(1 <= d <= 500 for d in durations):
            realistic_score += 5
        
        # Check expected speed includes 500
        if any(abs(s - 500) < 50 for s in expected_speeds):
            realistic_score += 2
        
        # Check for variety in speeds (at least one very slow)
        if any(s < 250 for s in download_speeds):
            realistic_score += 2
        
        # Check for at least one long outage
        if any(d > 30 for d in durations):
            realistic_score += 1
        
        score += realistic_score
        if realistic_score >= 12:
            feedback_parts.append(f"✅ Realistic data values ({realistic_score}/15)")
        else:
            feedback_parts.append(f"⚠️ Data values need improvement ({realistic_score}/15)")
        
        # 5. Summary statistics present (20 points)
        summary_score = 0
        summary_found = {}
        
        # Search rows 18-26 for summary labels
        for row in range(18, 27):
            label = sheet.cell(row, 1).value  # Column A
            value = sheet.cell(row, 2).value  # Column B
            
            if label:
                label_lower = str(label).lower()
                if 'total' in label_lower and 'incident' in label_lower:
                    summary_found['total'] = (row, value)
                elif 'average' in label_lower and ('downtime' in label_lower or 'duration' in label_lower):
                    summary_found['avg_duration'] = (row, value)
                elif 'average' in label_lower and 'download' in label_lower:
                    summary_found['avg_speed'] = (row, value)
                elif 'reliability' in label_lower or 'uptime' in label_lower:
                    summary_found['reliability'] = (row, value)
        
        for key in ['total', 'avg_duration', 'avg_speed', 'reliability']:
            if key in summary_found:
                _, value = summary_found[key]
                if value is not None and isinstance(value, (int, float)):
                    summary_score += 5
        
        score += summary_score
        if summary_score >= 15:
            feedback_parts.append(f"✅ Summary statistics ({summary_score}/20) - found {len(summary_found)}/4 required metrics")
        else:
            feedback_parts.append(f"⚠️ Summary statistics ({summary_score}/20) - only found {len([k for k in summary_found if isinstance(summary_found[k][1], (int, float))])}/4 metrics with values")
        
        # 6. Correct formulas (15 points)
        formula_score = 0
        
        # Re-open without data_only to check formulas
        from openpyxl import load_workbook
        try:
            wb_formula = load_workbook(temp_file.name, data_only=False)
            sheet_formula = wb_formula.active
            
            if 'total' in summary_found:
                row_num = summary_found['total'][0]
                cell = sheet_formula.cell(row_num, 2)
                formula_str = str(cell.value) if cell.value else ""
                if formula_str.startswith('=') and 'COUNT' in formula_str.upper():
                    formula_score += 4
                    feedback_parts.append("✅ Total Incidents uses COUNT formula")
                elif isinstance(summary_found['total'][1], (int, float)) and summary_found['total'][1] >= 8:
                    # At least they got the right count value
                    formula_score += 2
                    feedback_parts.append("⚠️ Total Incidents has correct value but should use COUNT formula")
            
            if 'avg_duration' in summary_found:
                row_num = summary_found['avg_duration'][0]
                cell = sheet_formula.cell(row_num, 2)
                formula_str = str(cell.value) if cell.value else ""
                if formula_str.startswith('=') and 'AVERAGE' in formula_str.upper():
                    formula_score += 4
                    feedback_parts.append("✅ Average Duration uses AVERAGE formula")
                elif isinstance(summary_found['avg_duration'][1], (int, float)):
                    formula_score += 2
                    feedback_parts.append("⚠️ Average Duration has value but should use AVERAGE formula")
            
            if 'avg_speed' in summary_found:
                row_num = summary_found['avg_speed'][0]
                cell = sheet_formula.cell(row_num, 2)
                formula_str = str(cell.value) if cell.value else ""
                if formula_str.startswith('=') and 'AVERAGE' in formula_str.upper():
                    formula_score += 4
                    feedback_parts.append("✅ Average Speed uses AVERAGE formula")
                elif isinstance(summary_found['avg_speed'][1], (int, float)):
                    formula_score += 2
                    feedback_parts.append("⚠️ Average Speed has value but should use AVERAGE formula")
            
            if 'reliability' in summary_found:
                row_num = summary_found['reliability'][0]
                cell = sheet_formula.cell(row_num, 2)
                formula_str = str(cell.value) if cell.value else ""
                if formula_str.startswith('=') and ('43200' in formula_str or '30' in formula_str):
                    formula_score += 3
                    feedback_parts.append("✅ Reliability uses time calculation formula")
                elif isinstance(summary_found['reliability'][1], (int, float)):
                    formula_score += 1
                    feedback_parts.append("⚠️ Reliability has value but should use formula with time period")
        except Exception as e:
            logger.warning(f"Could not check formulas: {e}")
        
        score += formula_score
        
        # 7. Conditional formatting (5 points) - Basic check
        # Check if any cells in columns D and G have fill color
        conditional_score = 0
        
        has_d_formatting = False
        has_g_formatting = False
        
        for row in range(2, 2 + min(data_rows, 15)):
            cell_d = sheet.cell(row, 4)
            cell_g = sheet.cell(row, 7)
            
            # Check if cell has any fill color (not default)
            if cell_d.fill and cell_d.fill.start_color:
                color_idx = str(cell_d.fill.start_color.index) if hasattr(cell_d.fill.start_color, 'index') else ""
                if color_idx and color_idx not in ['00000000', '00FFFFFF', 'FFFFFFFF', '']:
                    has_d_formatting = True
            
            if cell_g.fill and cell_g.fill.start_color:
                color_idx = str(cell_g.fill.start_color.index) if hasattr(cell_g.fill.start_color, 'index') else ""
                if color_idx and color_idx not in ['00000000', '00FFFFFF', 'FFFFFFFF', '']:
                    has_g_formatting = True
        
        if has_d_formatting:
            conditional_score += 2.5
        if has_g_formatting:
            conditional_score += 2.5
        
        score += int(conditional_score)
        if conditional_score >= 4:
            feedback_parts.append(f"✅ Conditional formatting ({int(conditional_score)}/5)")
        else:
            feedback_parts.append(f"⚠️ Conditional formatting ({int(conditional_score)}/5) - apply colors to highlight issues")
        
        # Final assessment
        passed = score >= 75
        feedback = " | ".join(feedback_parts)
        
        return {
            "passed": passed,
            "score": score / max_score,
            "feedback": f"Score: {score}/{max_score}. {feedback}"
        }
        
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"Verification error: {str(e)}"
        }
    finally:
        if temp_file and os.path.exists(temp_file.name):
            try:
                os.unlink(temp_file.name)
            except:
                pass