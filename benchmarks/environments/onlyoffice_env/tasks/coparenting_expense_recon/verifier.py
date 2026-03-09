#!/usr/bin/env python3
"""
Verifier for Co-Parenting Expense Reconciliation task
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_coparenting_expense_recon(traj, env_info, task_info):
    """
    Verify the co-parenting expense reconciliation spreadsheet.
    
    Checks:
    1. Three sheets exist: "Custody Days", "Expenses", "Summary"
    2. Custody Days sheet has proper date tracking and calculations
    3. Expenses sheet has all 8 expenses with proper categorization
    4. Summary sheet has formulas and correct calculations
    5. Data integrity and realistic values
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/Custody_Recon_March_2024.xlsx"
    temp_dir = None
    temp_file = None
    
    try:
        # Create temporary file for the spreadsheet
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        # Copy the file from container
        try:
            copy_from_env(container_path, temp_file.name)
        except Exception as e:
            return {
                "passed": False,
                "score": 0,
                "feedback": f"Failed to copy spreadsheet from container: {str(e)}"
            }
        
        if not os.path.exists(temp_file.name) or os.path.getsize(temp_file.name) == 0:
            return {
                "passed": False,
                "score": 0,
                "feedback": f"Spreadsheet file not found or empty: {container_path}"
            }
        
        # Parse the spreadsheet
        success, wb, error = copy_and_parse_document(
            temp_file.name,
            lambda src, dst: __import__('shutil').copy(src, dst),
            file_format='xlsx'
        )
        
        if not success:
            return {
                "passed": False,
                "score": 0,
                "feedback": f"Failed to parse spreadsheet: {error}"
            }
        
        criteria_passed = 0
        max_criteria = 10
        feedback_parts = []
        
        # Check 1: Verify all three sheets exist (1 point)
        sheet_names = wb.sheetnames
        required_sheets = ["Custody Days", "Expenses", "Summary"]
        missing_sheets = [s for s in required_sheets if s not in sheet_names]
        
        if not missing_sheets:
            criteria_passed += 1
            feedback_parts.append("✅ All three required sheets exist")
        else:
            feedback_parts.append(f"❌ Missing sheets: {', '.join(missing_sheets)}")
            return {
                "passed": False,
                "score": 0,
                "feedback": " | ".join(feedback_parts) + " | Cannot verify without all sheets"
            }
        
        # Check 2: Custody Days sheet structure (2 points)
        custody_sheet = wb["Custody Days"]
        custody_data = get_sheet_data(wb, "Custody Days", max_rows=55, max_cols=10)
        
        # Look for custody entries (You/Ex/Split)
        custody_entries = []
        for row_idx, row in enumerate(custody_data[8:40], start=9):  # Days should be rows 9-39
            if row and len(row) > 2 and row[2]:
                entry = str(row[2]).strip().lower()
                if entry in ['you', 'ex', 'split']:
                    custody_entries.append(entry)
        
        if len(custody_entries) >= 30:  # Should have entries for all 31 days
            criteria_passed += 1
            feedback_parts.append(f"✅ Custody tracking complete ({len(custody_entries)} entries)")
        elif len(custody_entries) >= 20:
            criteria_passed += 0.5
            feedback_parts.append(f"⚠️ Partial custody tracking ({len(custody_entries)}/31 entries)")
        else:
            feedback_parts.append(f"❌ Insufficient custody tracking ({len(custody_entries)}/31 entries)")
        
        # Look for custody totals
        has_totals = False
        your_days = None
        for row in custody_data[40:]:  # Summary section
            if row and len(row) > 1:
                row_str = ' '.join([str(cell).lower() for cell in row if cell])
                if 'your total days' in row_str or 'your days' in row_str:
                    # Look for a number in this row or next cells
                    for cell in row[1:5]:
                        if cell and isinstance(cell, (int, float)) and 10 <= cell <= 20:
                            your_days = cell
                            has_totals = True
                            break
        
        if has_totals:
            criteria_passed += 1
            feedback_parts.append(f"✅ Custody day totals calculated (Your days: {your_days})")
        else:
            feedback_parts.append("❌ Custody day totals not calculated")
        
        # Check 3: Expenses sheet structure and data (3 points)
        expense_sheet = wb["Expenses"]
        expense_data = get_sheet_data(wb, "Expenses", max_rows=50, max_cols=10)
        
        # Look for expense entries (starting around row 14)
        expense_rows = []
        expense_keywords = ['medical', 'educational', 'extracurricular', 'disputed', 'routine',
                           'copay', 'emma', 'liam', 'soccer', 'museum', 'calculator', 'haircut']
        
        for row_idx, row in enumerate(expense_data[13:30], start=14):  # Check rows 14-30
            if row and len(row) >= 4:
                row_str = ' '.join([str(cell).lower() for cell in row if cell])
                # Check if this row contains expense-related keywords
                if any(keyword in row_str for keyword in expense_keywords):
                    # Check if there's an amount (numeric value in typical range)
                    has_amount = False
                    for cell in row:
                        if isinstance(cell, (int, float)) and 5 <= cell <= 200:
                            has_amount = True
                            break
                    if has_amount:
                        expense_rows.append((row_idx, row))
        
        num_expenses = len(expense_rows)
        if num_expenses >= 8:
            criteria_passed += 1.5
            feedback_parts.append(f"✅ All {num_expenses} expense entries found")
        elif num_expenses >= 6:
            criteria_passed += 1
            feedback_parts.append(f"⚠️ Found {num_expenses}/8 expense entries")
        elif num_expenses >= 4:
            criteria_passed += 0.5
            feedback_parts.append(f"⚠️ Found only {num_expenses}/8 expense entries")
        else:
            feedback_parts.append(f"❌ Found only {num_expenses}/8 expense entries")
        
        # Check for categories
        has_categories = False
        category_keywords = ['medical', 'educational', 'extracurricular', 'disputed', 'routine']
        category_count = 0
        
        for row in expense_rows:
            row_str = ' '.join([str(cell).lower() for cell in row[1] if cell])
            if any(cat in row_str for cat in category_keywords):
                category_count += 1
        
        if category_count >= 6:
            criteria_passed += 0.75
            has_categories = True
            feedback_parts.append(f"✅ Expense categories properly assigned ({category_count} categorized)")
        elif category_count >= 4:
            criteria_passed += 0.5
            feedback_parts.append(f"⚠️ Some expense categories assigned ({category_count}/8)")
        else:
            feedback_parts.append(f"❌ Insufficient categorization ({category_count}/8)")
        
        # Check for "Your Share" calculations
        has_share_calc = False
        share_values = []
        for row in expense_rows:
            if len(row) >= 5:
                # Check column E (index 4) for share amounts
                if isinstance(row[4], (int, float)) and 0 < row[4] <= 200:
                    share_values.append(row[4])
                    has_share_calc = True
        
        if len(share_values) >= 6:
            criteria_passed += 0.75
            feedback_parts.append(f"✅ Share calculations present ({len(share_values)} entries)")
        elif len(share_values) >= 4:
            criteria_passed += 0.5
            feedback_parts.append(f"⚠️ Some share calculations ({len(share_values)}/8)")
        else:
            feedback_parts.append(f"❌ Missing share calculations")
        
        # Check 4: Summary sheet with calculations (3 points)
        summary_sheet = wb["Summary"]
        summary_data = get_sheet_data(wb, "Summary", max_rows=30, max_cols=10)
        
        # Load workbook without data_only to check for formulas
        from openpyxl import load_workbook as load_wb_raw
        has_formulas = False
        formula_count = 0
        
        try:
            wb_formulas = load_wb_raw(temp_file.name, data_only=False)
            summary_formula_sheet = wb_formulas["Summary"]
            
            # Check for formulas in summary section
            for row in summary_formula_sheet.iter_rows(min_row=5, max_row=20, max_col=5):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        has_formulas = True
                        formula_count += 1
        except Exception as e:
            feedback_parts.append(f"⚠️ Could not verify formulas: {e}")
        
        if formula_count >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ Formulas used in Summary ({formula_count} formulas)")
        elif formula_count >= 1:
            criteria_passed += 0.5
            feedback_parts.append(f"⚠️ Some formulas in Summary ({formula_count} formulas)")
        else:
            feedback_parts.append("❌ No formulas detected in Summary (hard-coded values?)")
        
        # Check for key summary calculations
        total_found = False
        extraordinary_found = False
        reimbursement_found = False
        
        for row_idx, row in enumerate(summary_data[4:20], start=5):
            if row and len(row) >= 2:
                label = str(row[0]).lower() if row[0] else ""
                value = row[1] if len(row) > 1 else None
                
                # Look for total expenses
                if 'total' in label and 'expense' in label:
                    if isinstance(value, (int, float)) and 200 <= value <= 400:
                        total_found = True
                
                # Look for extraordinary expenses
                if 'extraordinary' in label:
                    if isinstance(value, (int, float)) and 200 <= value <= 300:
                        extraordinary_found = True
                
                # Look for reimbursement/owed amount
                if 'owed' in label or 'owes' in label or 'reimburs' in label:
                    if isinstance(value, (int, float)) and 100 <= value <= 200:
                        reimbursement_found = True
        
        summary_calcs = sum([total_found, extraordinary_found, reimbursement_found])
        if summary_calcs >= 3:
            criteria_passed += 1.5
            feedback_parts.append("✅ All key summary calculations present")
        elif summary_calcs >= 2:
            criteria_passed += 1
            feedback_parts.append(f"⚠️ Some summary calculations present ({summary_calcs}/3)")
        elif summary_calcs >= 1:
            criteria_passed += 0.5
            feedback_parts.append(f"⚠️ Limited summary calculations ({summary_calcs}/3)")
        else:
            feedback_parts.append("❌ Summary calculations missing or incorrect")
        
        # Check for custody reference in summary
        has_custody_ref = False
        for row in summary_data[10:20]:
            if row:
                row_str = ' '.join([str(cell).lower() for cell in row if cell])
                if 'custody' in row_str or 'days' in row_str:
                    has_custody_ref = True
                    break
        
        if has_custody_ref:
            criteria_passed += 0.5
            feedback_parts.append("✅ Custody day verification included in Summary")
        else:
            feedback_parts.append("⚠️ No custody day reference in Summary")
        
        # Calculate final score
        score = int((criteria_passed / max_criteria) * 100)
        passed = score >= 70  # Need 70% to pass
        
        feedback = " | ".join(feedback_parts)
        feedback += f" | Final score: {criteria_passed:.1f}/{max_criteria} ({score}%)"
        
        if passed:
            feedback += " | ✅ Reconciliation spreadsheet is well-structured and suitable for co-parenting documentation"
        else:
            feedback += " | ❌ Spreadsheet needs improvements - ensure all sheets are complete with proper formulas"
        
        return {
            "passed": passed,
            "score": score / 100,
            "feedback": feedback
        }
        
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0,
            "feedback": f"Verification error: {str(e)}"
        }
    
    finally:
        # Cleanup
        if temp_file and os.path.exists(temp_file.name):
            try:
                os.unlink(temp_file.name)
            except Exception as e:
                logger.warning(f"Failed to cleanup temp file: {e}")
        
        if temp_dir:
            cleanup_temp_dir(temp_dir)
