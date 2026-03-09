#!/usr/bin/env python3
"""
Verifier for House Hunting Comparison task
"""

import sys
import os
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_house_hunting_comparison(traj, env_info, task_info):
    """
    Verify that house hunting comparison spreadsheet was created correctly.

    Checks:
    1. File exists and can be opened
    2. Headers are present and correct (Row 1)
    3. Property data is entered correctly (Rows 2-5)
    4. Formulas exist in column F (not hardcoded values)
    5. Formula calculations are correct (price/sqft)
    6. Currency formatting applied to columns B and F
    7. Summary label exists in A7
    8. MIN formula exists in B7
    9. Summary calculation is correct
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/house_comparison.xlsx"
    temp_file = None

    try:
        # Copy and parse the spreadsheet
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {"passed": False, "score": 0, "feedback": f"Failed to load spreadsheet: {error}"}

        criteria_passed = 0
        total_criteria = 8
        feedback_parts = []

        sheet = wb.active
        sheet_name = sheet.title

        # Expected data structure
        expected_headers = ["Address", "Price", "Square Feet", "Bedrooms", "Bathrooms", "Price per Sq Ft"]
        expected_properties = {
            2: {
                "address": "742 Evergreen Terrace",
                "price": 385000,
                "sqft": 2100,
                "bed": 3,
                "bath": 2,
                "price_per_sqft": 385000 / 2100  # ~183.33
            },
            3: {
                "address": "1640 Riverside Drive",
                "price": 425000,
                "sqft": 2450,
                "bed": 4,
                "bath": 2.5,
                "price_per_sqft": 425000 / 2450  # ~173.47
            },
            4: {
                "address": "344 Clinton Way",
                "price": 310000,
                "sqft": 1650,
                "bed": 2,
                "bath": 2,
                "price_per_sqft": 310000 / 1650  # ~187.88
            },
            5: {
                "address": "2311 North Los Robles",
                "price": 468000,
                "sqft": 2800,
                "bed": 4,
                "bath": 3,
                "price_per_sqft": 468000 / 2800  # ~167.14
            }
        }

        # ===== CRITERION 1: Check headers (Row 1) =====
        try:
            actual_headers = []
            for col in range(1, 7):  # A1 to F1
                cell_value = sheet.cell(row=1, column=col).value
                actual_headers.append(cell_value)

            headers_correct = True
            for i, expected_header in enumerate(expected_headers):
                actual = actual_headers[i] if i < len(actual_headers) else None
                if actual is None:
                    headers_correct = False
                    feedback_parts.append(f"❌ Missing header in column {chr(65+i)}1")
                    break
                # Flexible matching (case-insensitive, contains check)
                if expected_header.lower() not in str(actual).lower():
                    headers_correct = False
                    feedback_parts.append(f"❌ Header mismatch in {chr(65+i)}1: expected '{expected_header}', got '{actual}'")
                    break

            if headers_correct:
                criteria_passed += 1
                feedback_parts.append("✅ All headers correct")
        except Exception as e:
            feedback_parts.append(f"❌ Error checking headers: {str(e)}")

        # ===== CRITERION 2: Check property data accuracy (Rows 2-5, Columns A-E) =====
        try:
            data_correct = True
            data_errors = []

            for row_num, expected in expected_properties.items():
                # Check address (column A)
                address_cell = sheet.cell(row=row_num, column=1).value
                if address_cell is None or expected["address"].lower() not in str(address_cell).lower():
                    data_errors.append(f"A{row_num}: expected '{expected['address']}', got '{address_cell}'")
                    data_correct = False

                # Check price (column B)
                price_cell = sheet.cell(row=row_num, column=2).value
                if price_cell is None or abs(float(price_cell) - expected["price"]) > 1:
                    data_errors.append(f"B{row_num}: expected {expected['price']}, got {price_cell}")
                    data_correct = False

                # Check square feet (column C)
                sqft_cell = sheet.cell(row=row_num, column=3).value
                if sqft_cell is None or abs(float(sqft_cell) - expected["sqft"]) > 1:
                    data_errors.append(f"C{row_num}: expected {expected['sqft']}, got {sqft_cell}")
                    data_correct = False

                # Check bedrooms (column D)
                bed_cell = sheet.cell(row=row_num, column=4).value
                if bed_cell is None or abs(float(bed_cell) - expected["bed"]) > 0.1:
                    data_errors.append(f"D{row_num}: expected {expected['bed']}, got {bed_cell}")
                    data_correct = False

                # Check bathrooms (column E)
                bath_cell = sheet.cell(row=row_num, column=5).value
                if bath_cell is None or abs(float(bath_cell) - expected["bath"]) > 0.1:
                    data_errors.append(f"E{row_num}: expected {expected['bath']}, got {bath_cell}")
                    data_correct = False

            if data_correct:
                criteria_passed += 1
                feedback_parts.append("✅ All property data correct")
            else:
                feedback_parts.append(f"❌ Property data errors: {'; '.join(data_errors[:3])}")  # Show first 3 errors
        except Exception as e:
            feedback_parts.append(f"❌ Error checking property data: {str(e)}")

        # ===== CRITERION 3 & 4: Check formulas exist and are correct (Column F, Rows 2-5) =====
        try:
            # Need to reload workbook without data_only to see formulas
            temp_file_obj = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file = temp_file_obj.name
            temp_file_obj.close()
            
            copy_from_env(container_path, temp_file)
            
            from openpyxl import load_workbook
            wb_formulas = load_workbook(temp_file, data_only=False)
            sheet_formulas = wb_formulas.active

            formulas_exist = True
            formula_errors = []

            for row_num in range(2, 6):  # Rows 2-5
                cell_ref = f'F{row_num}'
                cell = sheet_formulas[cell_ref]
                cell_value_str = str(cell.value) if cell.value else ""

                # Check if cell contains a formula (starts with =)
                if not cell_value_str.startswith('='):
                    formulas_exist = False
                    formula_errors.append(f"{cell_ref} missing formula (contains: '{cell_value_str[:20]}')")

            if formulas_exist:
                criteria_passed += 1
                feedback_parts.append("✅ Formulas exist in F2-F5")
            else:
                feedback_parts.append(f"❌ Missing formulas: {'; '.join(formula_errors)}")

            # Check formula calculations are correct
            calculations_correct = True
            calc_errors = []

            for row_num, expected in expected_properties.items():
                cell_ref = f'F{row_num}'
                actual_value = sheet[cell_ref].value

                if actual_value is None:
                    calculations_correct = False
                    calc_errors.append(f"{cell_ref} is empty")
                elif abs(float(actual_value) - expected["price_per_sqft"]) > 1.0:
                    calculations_correct = False
                    calc_errors.append(f"{cell_ref}: expected ~${expected['price_per_sqft']:.2f}, got ${actual_value:.2f}")

            if calculations_correct:
                criteria_passed += 1
                feedback_parts.append("✅ All price/sqft calculations correct")
            else:
                feedback_parts.append(f"❌ Calculation errors: {'; '.join(calc_errors[:2])}")

        except Exception as e:
            feedback_parts.append(f"❌ Error checking formulas: {str(e)}")

        # ===== CRITERION 5: Check currency formatting (Columns B and F) =====
        try:
            formatting_correct = True
            format_issues = []

            # Check column B (Price) - sample B2
            b2_format = sheet['B2'].number_format
            # Common currency formats
            currency_formats = ['$#,##0.00', '_($* #,##0.00_)', '[$$-409]#,##0.00', '#,##0.00', '$ #,##0.00', 
                              '"$"#,##0.00', '_-$* #,##0.00_-', '0.00', '$#,##0.00_);[Red]($#,##0.00)']
            
            if '$' not in b2_format and 'Currency' not in b2_format and b2_format not in currency_formats:
                formatting_correct = False
                format_issues.append(f"Column B not currency formatted (format: {b2_format})")

            # Check column F (Price per Sq Ft) - sample F2
            f2_format = sheet['F2'].number_format
            if '$' not in f2_format and 'Currency' not in f2_format and f2_format not in currency_formats:
                formatting_correct = False
                format_issues.append(f"Column F not currency formatted (format: {f2_format})")

            if formatting_correct:
                criteria_passed += 1
                feedback_parts.append("✅ Currency formatting applied")
            else:
                feedback_parts.append(f"❌ Formatting issues: {'; '.join(format_issues)}")
        except Exception as e:
            feedback_parts.append(f"❌ Error checking formatting: {str(e)}")

        # ===== CRITERION 6: Check summary label in A7 =====
        try:
            a7_value = sheet['A7'].value
            if a7_value and ("best value" in str(a7_value).lower() or "lowest" in str(a7_value).lower()):
                criteria_passed += 1
                feedback_parts.append("✅ Summary label present in A7")
            else:
                feedback_parts.append(f"❌ A7 missing proper label (has: '{a7_value}')")
        except Exception as e:
            feedback_parts.append(f"❌ Error checking A7 label: {str(e)}")

        # ===== CRITERION 7: Check MIN formula exists in B7 =====
        try:
            if temp_file and os.path.exists(temp_file):
                wb_formulas_b7 = load_workbook(temp_file, data_only=False)
                sheet_formulas_b7 = wb_formulas_b7.active
                
                b7_cell = sheet_formulas_b7['B7']
                b7_formula = str(b7_cell.value) if b7_cell.value else ""

                if '=MIN' in b7_formula.upper() or '=min' in b7_formula:
                    criteria_passed += 1
                    feedback_parts.append("✅ MIN formula exists in B7")
                else:
                    feedback_parts.append(f"❌ B7 missing MIN formula (contains: '{b7_formula[:30]}')")
            else:
                feedback_parts.append("❌ Cannot verify B7 formula (temp file issue)")
        except Exception as e:
            feedback_parts.append(f"❌ Error checking B7 formula: {str(e)}")

        # ===== CRITERION 8: Check summary calculation is correct =====
        try:
            b7_result = sheet['B7'].value
            # Expected minimum is 2311 North Los Robles at ~$167.14/sqft
            expected_min = min([prop["price_per_sqft"] for prop in expected_properties.values()])

            if b7_result is not None and abs(float(b7_result) - expected_min) < 1.5:
                criteria_passed += 1
                feedback_parts.append(f"✅ Summary MIN result correct: ${b7_result:.2f}/sqft")
            else:
                feedback_parts.append(f"❌ B7 result incorrect: expected ~${expected_min:.2f}, got {b7_result}")
        except Exception as e:
            feedback_parts.append(f"❌ Error checking B7 result: {str(e)}")

        # Calculate final score
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 75

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        # Clean up temp file
        if temp_file and os.path.exists(temp_file):
            try:
                os.unlink(temp_file)
            except:
                pass
        cleanup_temp_dir(None)