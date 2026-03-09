#!/usr/bin/env python3
"""
Verifier for Craft Fair Pricing Sheet task

Checks:
1. File exists and is parseable
2. Headers are present and correct
3. Product data is entered correctly
4. Total Cost formulas (column E) are correct
5. Selling Price formulas (column G) are correct
6. Potential Revenue formulas (column I) are correct
7. Summary section formulas are correct
8. Currency formatting is applied
9. Percentage formatting is applied
"""

import sys
import os
import logging
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    parse_xlsx_file,
    get_cell_value,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_craft_fair_pricing(traj, env_info, task_info):
    """
    Verify that craft fair pricing sheet was created correctly.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/craft_fair_pricing.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_pricing_')

    try:
        # Copy file from container
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=temp_dir)
        temp_file.close()
        
        try:
            copy_from_env(container_path, temp_file.name)
        except Exception as e:
            logger.error(f"Failed to copy file: {e}")
            return {"passed": False, "score": 0.0, "feedback": f"File not found or not accessible: {container_path}"}

        if not os.path.exists(temp_file.name) or os.path.getsize(temp_file.name) == 0:
            return {"passed": False, "score": 0.0, "feedback": f"File not found or empty: {container_path}"}

        # Parse spreadsheet with data_only=True to get calculated values
        wb_values = parse_xlsx_file(temp_file.name)
        if wb_values is None:
            return {"passed": False, "score": 0.0, "feedback": "Failed to parse spreadsheet"}

        # Also load with data_only=False to check formulas
        try:
            from openpyxl import load_workbook
            wb_formulas = load_workbook(temp_file.name, data_only=False)
        except Exception as e:
            logger.warning(f"Could not load formulas: {e}")
            wb_formulas = None

        score = 0
        max_score = 100
        feedback_parts = []

        # Get the active sheet
        ws_values = wb_values.active
        ws_formulas = wb_formulas.active if wb_formulas else None

        # ===== Criterion 1: File exists and parseable (5 points) =====
        score += 5
        feedback_parts.append("✅ File exists and is parseable")

        # ===== Criterion 2: Headers present (10 points) =====
        expected_headers = [
            'product', 'material', 'time', 'hourly', 'total', 'markup', 'selling', 'quantity', 'revenue'
        ]
        headers_found = 0
        for col in range(1, 10):
            cell_val = str(ws_values.cell(1, col).value or "").lower()
            if any(exp in cell_val for exp in expected_headers):
                headers_found += 1

        if headers_found >= 7:
            score += 10
            feedback_parts.append(f"✅ Headers present ({headers_found}/9)")
        elif headers_found >= 5:
            score += 5
            feedback_parts.append(f"⚠️ Some headers present ({headers_found}/9)")
        else:
            feedback_parts.append(f"❌ Headers missing ({headers_found}/9)")

        # ===== Criterion 3: Product data entered (15 points) =====
        # Expected data (approximate values)
        expected_products = [
            {'material': 4.5, 'time': 0.5, 'markup': 1.5, 'qty': 12},
            {'material': 2.0, 'time': 0.25, 'markup': 2.0, 'qty': 20},
            {'material': 8.0, 'time': 2.0, 'markup': 1.2, 'qty': 8},
            {'material': 5.5, 'time': 1.0, 'markup': 1.8, 'qty': 15},
            {'material': 6.0, 'time': 1.5, 'markup': 1.6, 'qty': 10},
        ]

        products_valid = 0
        for row_idx in range(2, 7):  # Rows 2-6
            material = ws_values.cell(row_idx, 2).value  # Column B
            time_hrs = ws_values.cell(row_idx, 3).value  # Column C
            markup = ws_values.cell(row_idx, 6).value    # Column F
            qty = ws_values.cell(row_idx, 8).value       # Column H

            # Check if data is present and numeric
            if all(val is not None and isinstance(val, (int, float)) for val in [material, time_hrs, qty]):
                products_valid += 1

        if products_valid >= 4:
            score += 15
            feedback_parts.append(f"✅ Product data entered ({products_valid}/5 products)")
        elif products_valid >= 3:
            score += 10
            feedback_parts.append(f"⚠️ Most product data entered ({products_valid}/5 products)")
        elif products_valid >= 1:
            score += 5
            feedback_parts.append(f"⚠️ Some product data entered ({products_valid}/5 products)")
        else:
            feedback_parts.append(f"❌ Product data missing ({products_valid}/5 products)")

        # ===== Criterion 4: Total Cost formulas correct (15 points) =====
        # Formula: Material + (Time × Hourly Rate)
        total_cost_correct = 0
        total_cost_has_formula = 0

        for row_idx in range(2, 7):
            material = ws_values.cell(row_idx, 2).value
            time_hrs = ws_values.cell(row_idx, 3).value
            hourly = ws_values.cell(row_idx, 4).value
            total_cost = ws_values.cell(row_idx, 5).value

            # Check if formula exists
            if ws_formulas:
                formula_cell = ws_formulas.cell(row_idx, 5).value
                if isinstance(formula_cell, str) and formula_cell.startswith('='):
                    total_cost_has_formula += 1

            # Verify calculation
            if all(val is not None and isinstance(val, (int, float)) for val in [material, time_hrs, hourly, total_cost]):
                expected = material + (time_hrs * hourly)
                if abs(total_cost - expected) < 0.50:  # $0.50 tolerance
                    total_cost_correct += 1

        if total_cost_correct >= 4:
            score += 15
            feedback_parts.append(f"✅ Total Cost formulas correct ({total_cost_correct}/5)")
        elif total_cost_correct >= 3:
            score += 10
            feedback_parts.append(f"⚠️ Most Total Cost formulas correct ({total_cost_correct}/5)")
        elif total_cost_correct >= 1:
            score += 5
            feedback_parts.append(f"⚠️ Some Total Cost formulas correct ({total_cost_correct}/5)")
        else:
            feedback_parts.append(f"❌ Total Cost formulas incorrect ({total_cost_correct}/5)")

        # ===== Criterion 5: Selling Price formulas correct (15 points) =====
        # Formula: Total Cost × (1 + Markup%)
        selling_price_correct = 0

        for row_idx in range(2, 7):
            total_cost = ws_values.cell(row_idx, 5).value
            markup = ws_values.cell(row_idx, 6).value
            selling_price = ws_values.cell(row_idx, 7).value

            if all(val is not None and isinstance(val, (int, float)) for val in [total_cost, markup, selling_price]):
                expected = total_cost * (1 + markup)
                if abs(selling_price - expected) < 1.0:  # $1.00 tolerance
                    selling_price_correct += 1

        if selling_price_correct >= 4:
            score += 15
            feedback_parts.append(f"✅ Selling Price formulas correct ({selling_price_correct}/5)")
        elif selling_price_correct >= 3:
            score += 10
            feedback_parts.append(f"⚠️ Most Selling Price formulas correct ({selling_price_correct}/5)")
        elif selling_price_correct >= 1:
            score += 5
            feedback_parts.append(f"⚠️ Some Selling Price formulas correct ({selling_price_correct}/5)")
        else:
            feedback_parts.append(f"❌ Selling Price formulas incorrect ({selling_price_correct}/5)")

        # ===== Criterion 6: Potential Revenue formulas correct (10 points) =====
        # Formula: Selling Price × Quantity
        revenue_correct = 0

        for row_idx in range(2, 7):
            selling_price = ws_values.cell(row_idx, 7).value
            qty = ws_values.cell(row_idx, 8).value
            revenue = ws_values.cell(row_idx, 9).value

            if all(val is not None and isinstance(val, (int, float)) for val in [selling_price, qty, revenue]):
                expected = selling_price * qty
                if abs(revenue - expected) < 2.0:  # $2.00 tolerance
                    revenue_correct += 1

        if revenue_correct >= 4:
            score += 10
            feedback_parts.append(f"✅ Potential Revenue formulas correct ({revenue_correct}/5)")
        elif revenue_correct >= 3:
            score += 7
            feedback_parts.append(f"⚠️ Most Potential Revenue formulas correct ({revenue_correct}/5)")
        elif revenue_correct >= 1:
            score += 3
            feedback_parts.append(f"⚠️ Some Potential Revenue formulas correct ({revenue_correct}/5)")
        else:
            feedback_parts.append(f"❌ Potential Revenue formulas incorrect ({revenue_correct}/5)")

        # ===== Criterion 7: Summary section formulas (15 points) =====
        # Look for summary section (might be in rows 8-12 or nearby)
        summary_correct = 0
        summary_found = False

        # Try to find summary section by looking for non-empty cells in column B around row 9-12
        for start_row in range(8, 13):
            val = ws_values.cell(start_row, 2).value
            if val is not None and isinstance(val, (int, float)):
                summary_found = True
                break

        if summary_found:
            # Check B9 or nearby: Total Inventory Value (sum of column E)
            total_costs = []
            for row_idx in range(2, 7):
                val = ws_values.cell(row_idx, 5).value
                if val and isinstance(val, (int, float)):
                    total_costs.append(val)

            expected_inventory = sum(total_costs) if total_costs else 0

            # Check B10 or nearby: 100% sellthrough (sum of column I)
            revenues = []
            for row_idx in range(2, 7):
                val = ws_values.cell(row_idx, 9).value
                if val and isinstance(val, (int, float)):
                    revenues.append(val)

            expected_revenue_100 = sum(revenues) if revenues else 0
            expected_revenue_60 = expected_revenue_100 * 0.6
            expected_profit = expected_revenue_60 - expected_inventory

            # Check if summary values are approximately correct
            for check_row in range(8, 14):
                summary_val = ws_values.cell(check_row, 2).value
                if summary_val and isinstance(summary_val, (int, float)):
                    # Check if it matches any expected value
                    if abs(summary_val - expected_inventory) < 5:
                        summary_correct += 1
                    elif abs(summary_val - expected_revenue_100) < 10:
                        summary_correct += 1
                    elif abs(summary_val - expected_revenue_60) < 10:
                        summary_correct += 1
                    elif abs(summary_val - expected_profit) < 10:
                        summary_correct += 1

            # Cap at 4 (since there are 4 summary calculations)
            summary_correct = min(summary_correct, 4)

            if summary_correct >= 3:
                score += 15
                feedback_parts.append(f"✅ Summary section present ({summary_correct}/4 calculations)")
            elif summary_correct >= 2:
                score += 10
                feedback_parts.append(f"⚠️ Summary section partial ({summary_correct}/4 calculations)")
            elif summary_correct >= 1:
                score += 5
                feedback_parts.append(f"⚠️ Summary section incomplete ({summary_correct}/4 calculations)")
            else:
                feedback_parts.append(f"❌ Summary section missing or incorrect")
        else:
            feedback_parts.append("❌ Summary section not found")

        # ===== Criterion 8: Currency formatting (10 points) =====
        # Check if columns B, E, G, I have currency formatting
        currency_cols = [2, 5, 7, 9]  # B, E, G, I
        currency_formatted = 0

        for col in currency_cols:
            # Check formatting on row 2 as sample
            cell = ws_values.cell(2, col)
            number_format = cell.number_format if hasattr(cell, 'number_format') else ''
            
            # Check if format contains currency symbols
            if '$' in number_format or 'currency' in number_format.lower():
                currency_formatted += 1

        if currency_formatted >= 3:
            score += 10
            feedback_parts.append(f"✅ Currency formatting applied ({currency_formatted}/4 columns)")
        elif currency_formatted >= 2:
            score += 6
            feedback_parts.append(f"⚠️ Some currency formatting ({currency_formatted}/4 columns)")
        elif currency_formatted >= 1:
            score += 3
            feedback_parts.append(f"⚠️ Minimal currency formatting ({currency_formatted}/4 columns)")
        else:
            feedback_parts.append("❌ Currency formatting not applied")

        # ===== Criterion 9: Percentage formatting (5 points) =====
        # Check if column F has percentage formatting
        cell = ws_values.cell(2, 6)  # F2
        number_format = cell.number_format if hasattr(cell, 'number_format') else ''
        
        if '%' in number_format or 'percent' in number_format.lower():
            score += 5
            feedback_parts.append("✅ Percentage formatting applied")
        else:
            # Check if values look like percentages (e.g., 1.5 for 150%)
            # This is still valid, just not formatted as percent display
            val = ws_values.cell(2, 6).value
            if val and isinstance(val, (int, float)) and 1.0 <= val <= 3.0:
                score += 3
                feedback_parts.append("⚠️ Percentage values present but not formatted as %")
            else:
                feedback_parts.append("❌ Percentage formatting not applied")

        # Calculate final score and pass/fail
        final_score = score / max_score
        passed = score >= 75

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": final_score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0.0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)
