#!/usr/bin/env python3
"""
Verifier for Backyard Flock Production Tracker task

Verifies that the user created a complete egg production tracking spreadsheet with:
- Proper headers (hen names)
- Date column (7 days)
- Production data (28 data points: 4 hens × 7 days)
- Weekly total formulas (SUM for each hen)
- Average formulas (AVERAGE for each hen)
- Daily total column (SUM for each day)
- Conditional formatting (red highlighting for <3 eggs/week)
- Correct calculated values
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

# Need openpyxl for detailed verification
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def find_header_row(sheet_data, keywords):
    """
    Find the row containing header keywords.
    Returns row index (0-based) or None.
    """
    for row_idx, row in enumerate(sheet_data):
        row_text = ' '.join([str(cell).lower() if cell else '' for cell in row])
        if any(kw.lower() in row_text for kw in keywords):
            return row_idx
    return None


def find_column_by_header(row, header_name):
    """
    Find column index (0-based) by header name in a row.
    Returns column index or None.
    """
    for col_idx, cell in enumerate(row):
        if cell and header_name.lower() in str(cell).lower():
            return col_idx
    return None


def check_cell_has_formula(ws, cell_ref, formula_keyword):
    """
    Check if a cell contains a formula with a specific keyword (e.g., 'SUM', 'AVERAGE').
    Requires workbook loaded with data_only=False.
    """
    try:
        cell = ws[cell_ref]
        cell_value = cell.value
        if cell_value and isinstance(cell_value, str):
            if formula_keyword.upper() in cell_value.upper():
                return True
        return False
    except Exception as e:
        logger.debug(f"Error checking formula in {cell_ref}: {e}")
        return False


def check_conditional_formatting(ws, cell_range):
    """
    Check if conditional formatting rules exist for the given range.
    Returns True if any conditional formatting is found.
    """
    try:
        if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
            # Check if any rules apply to cells in our range
            for cf in ws.conditional_formatting:
                # cf is a ConditionalFormattingRule object
                if hasattr(cf, 'sqref'):
                    # sqref contains the cell range(s) this rule applies to
                    return True
        return False
    except Exception as e:
        logger.debug(f"Error checking conditional formatting: {e}")
        return False


def check_cell_has_red_fill(ws, cell_ref):
    """
    Check if a cell has red/orange background fill.
    """
    try:
        cell = ws[cell_ref]
        if cell.fill and hasattr(cell.fill, 'fgColor'):
            if cell.fill.patternType == 'solid':
                color = cell.fill.fgColor
                if hasattr(color, 'rgb') and color.rgb:
                    # Check if color is reddish (high R value)
                    rgb_str = str(color.rgb)
                    if len(rgb_str) >= 8:  # AARRGGBB format
                        r_channel = rgb_str[2:4]
                        # Red channel should be high (>= 0xDD = 221)
                        if int(r_channel, 16) >= 0xDD:
                            return True
        return False
    except Exception as e:
        logger.debug(f"Error checking cell fill color: {e}")
        return False


def verify_egg_production_tracker(traj, env_info, task_info):
    """
    Verify the egg production tracker spreadsheet.
    
    Checks (10 criteria, need 8+ to pass at 80%):
    1. File exists and is readable
    2. Headers present (hen names: Henrietta, Nugget, Pepper, Goldie)
    3. Date column has 7 entries
    4. Production data is complete (28 data points)
    5. Weekly total formulas exist (B12:E12 with SUM)
    6. Weekly totals are correct (5, 4, 1, 6)
    7. Average formulas exist (B13:E13 with AVERAGE)
    8. Daily total column has formulas (F5:F11 with SUM)
    9. Conditional formatting applied or red highlighting exists
    10. Problem hen (Pepper with 1 egg) is highlighted
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    if load_workbook is None:
        return {"passed": False, "score": 0, "feedback": "openpyxl library not available"}

    container_path = "/home/ga/Documents/Spreadsheets/egg_production_log.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_egg_')
    temp_file = os.path.join(temp_dir, 'egg_production_log.xlsx')

    try:
        # Copy file from container
        copy_from_env(container_path, temp_file)

        if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
            return {
                "passed": False,
                "score": 0,
                "feedback": f"File not found or empty: {container_path}"
            }

        criteria_passed = 0
        feedback_parts = []

        # Load workbook with data_only=True first to check values
        wb_data = load_workbook(temp_file, data_only=True)
        ws_data = wb_data.active

        # Load workbook with data_only=False to check formulas
        wb_formulas = load_workbook(temp_file, data_only=False)
        ws_formulas = wb_formulas.active

        # Get all data for analysis
        sheet_data = []
        for row in ws_data.iter_rows(max_row=20, max_col=10, values_only=True):
            sheet_data.append(row)

        # Criterion 1: File exists and is readable (already passed if we got here)
        criteria_passed += 1
        feedback_parts.append("✅ File exists and is readable")

        # Criterion 2: Check for hen name headers
        hen_names = ['henrietta', 'nugget', 'pepper', 'goldie']
        headers_found = []
        header_row_idx = None

        for row_idx, row in enumerate(sheet_data):
            row_text = ' '.join([str(cell).lower() if cell else '' for cell in row])
            found_count = sum(1 for name in hen_names if name in row_text)
            if found_count >= 3:  # At least 3 hen names found
                header_row_idx = row_idx
                headers_found = [name for name in hen_names if name in row_text]
                break

        if len(headers_found) >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ Headers found: {', '.join(headers_found)} (row {header_row_idx + 1})")
        else:
            feedback_parts.append(f"❌ Hen name headers missing (found {len(headers_found)}/4)")

        if header_row_idx is None:
            # Can't verify further without finding header row
            score = int((criteria_passed / 10) * 100)
            feedback = " | ".join(feedback_parts)
            return {
                "passed": False,
                "score": score,
                "feedback": feedback + " | Cannot verify further without headers"
            }

        # Criterion 3: Check for date column (7 date entries)
        date_col_idx = 0  # Assume column A (index 0)
        data_start_row = header_row_idx + 1
        date_count = 0

        for row_idx in range(data_start_row, min(data_start_row + 10, len(sheet_data))):
            if row_idx < len(sheet_data):
                cell_value = sheet_data[row_idx][date_col_idx] if date_col_idx < len(sheet_data[row_idx]) else None
                if cell_value:
                    # Check if it looks like a date (contains number or day name)
                    cell_str = str(cell_value).lower()
                    if any(day in cell_str for day in ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']) or \
                       any(char.isdigit() for char in cell_str):
                        date_count += 1

        if date_count >= 7:
            criteria_passed += 1
            feedback_parts.append(f"✅ Date column has {date_count} entries")
        else:
            feedback_parts.append(f"❌ Date column incomplete ({date_count}/7 dates found)")

        # Criterion 4: Check production data completeness (count numeric entries in data area)
        data_cell_count = 0
        for row_idx in range(data_start_row, data_start_row + 7):
            if row_idx < len(sheet_data):
                for col_idx in range(1, 5):  # Columns B-E (indices 1-4)
                    if col_idx < len(sheet_data[row_idx]):
                        cell_value = sheet_data[row_idx][col_idx]
                        if cell_value is not None and isinstance(cell_value, (int, float)):
                            data_cell_count += 1

        if data_cell_count >= 20:  # At least 20 of 28 cells filled
            criteria_passed += 1
            feedback_parts.append(f"✅ Production data complete ({data_cell_count}/28 cells)")
        else:
            feedback_parts.append(f"❌ Production data incomplete ({data_cell_count}/28 cells)")

        # Criterion 5: Check for weekly total formulas (row 12 or nearby)
        # Look for row with "weekly" or "total" label
        total_row_idx = None
        for row_idx in range(data_start_row + 7, min(data_start_row + 12, len(sheet_data))):
            if row_idx < len(sheet_data):
                first_cell = sheet_data[row_idx][0] if len(sheet_data[row_idx]) > 0 else None
                if first_cell and ('total' in str(first_cell).lower() or 'weekly' in str(first_cell).lower()):
                    total_row_idx = row_idx
                    break

        formula_found = False
        if total_row_idx is not None:
            # Check if cells in this row contain SUM formulas
            total_row_num = total_row_idx + 1  # Convert to 1-based for openpyxl
            formula_count = 0
            for col_letter in ['B', 'C', 'D', 'E']:
                cell_ref = f'{col_letter}{total_row_num}'
                if check_cell_has_formula(ws_formulas, cell_ref, 'SUM'):
                    formula_count += 1

            if formula_count >= 3:
                criteria_passed += 1
                formula_found = True
                feedback_parts.append(f"✅ Weekly total SUM formulas found ({formula_count}/4)")
            else:
                feedback_parts.append(f"❌ Weekly total formulas missing ({formula_count}/4)")
        else:
            feedback_parts.append("❌ Weekly total row not found")

        # Criterion 6: Check weekly total values (5, 4, 1, 6 expected)
        expected_totals = [5, 4, 1, 6]
        totals_correct = False

        if total_row_idx is not None:
            actual_totals = []
            for col_idx in range(1, 5):  # Columns B-E
                if col_idx < len(sheet_data[total_row_idx]):
                    val = sheet_data[total_row_idx][col_idx]
                    if isinstance(val, (int, float)):
                        actual_totals.append(int(val))

            # Check if totals match expected (with some tolerance)
            if len(actual_totals) >= 4:
                matches = sum(1 for i, val in enumerate(actual_totals[:4]) 
                             if abs(val - expected_totals[i]) <= 1)
                if matches >= 3:
                    criteria_passed += 1
                    totals_correct = True
                    feedback_parts.append(f"✅ Weekly totals correct: {actual_totals[:4]}")
                else:
                    feedback_parts.append(f"❌ Weekly totals incorrect: {actual_totals[:4]} (expected {expected_totals})")
            else:
                feedback_parts.append(f"❌ Weekly totals missing (found {len(actual_totals)}/4)")
        else:
            feedback_parts.append("❌ Cannot verify weekly totals without total row")

        # Criterion 7: Check for average formulas
        avg_row_idx = None
        for row_idx in range(data_start_row + 7, min(data_start_row + 14, len(sheet_data))):
            if row_idx < len(sheet_data):
                first_cell = sheet_data[row_idx][0] if len(sheet_data[row_idx]) > 0 else None
                if first_cell and 'avg' in str(first_cell).lower():
                    avg_row_idx = row_idx
                    break

        if avg_row_idx is not None:
            avg_row_num = avg_row_idx + 1
            avg_formula_count = 0
            for col_letter in ['B', 'C', 'D', 'E']:
                cell_ref = f'{col_letter}{avg_row_num}'
                if check_cell_has_formula(ws_formulas, cell_ref, 'AVERAGE'):
                    avg_formula_count += 1

            if avg_formula_count >= 3:
                criteria_passed += 1
                feedback_parts.append(f"✅ Average formulas found ({avg_formula_count}/4)")
            else:
                feedback_parts.append(f"❌ Average formulas missing ({avg_formula_count}/4)")
        else:
            feedback_parts.append("❌ Average row not found")

        # Criterion 8: Check for daily total column (column F with SUM formulas)
        daily_total_formulas = 0
        for row_num in range(data_start_row + 1, data_start_row + 8):  # 7 days
            cell_ref = f'F{row_num}'
            if check_cell_has_formula(ws_formulas, cell_ref, 'SUM'):
                daily_total_formulas += 1

        if daily_total_formulas >= 5:
            criteria_passed += 1
            feedback_parts.append(f"✅ Daily total formulas found ({daily_total_formulas}/7)")
        else:
            feedback_parts.append(f"❌ Daily total formulas missing ({daily_total_formulas}/7)")

        # Criterion 9: Check for conditional formatting
        has_conditional_formatting = check_conditional_formatting(ws_formulas, 'B12:E12')

        if has_conditional_formatting:
            criteria_passed += 1
            feedback_parts.append("✅ Conditional formatting applied")
        else:
            # Check if at least one cell has red fill (manual highlighting)
            red_cells_found = 0
            if total_row_idx is not None:
                total_row_num = total_row_idx + 1
                for col_letter in ['B', 'C', 'D', 'E']:
                    cell_ref = f'{col_letter}{total_row_num}'
                    if check_cell_has_red_fill(ws_formulas, cell_ref):
                        red_cells_found += 1

            if red_cells_found > 0:
                criteria_passed += 1
                feedback_parts.append(f"✅ Red highlighting found ({red_cells_found} cells)")
            else:
                feedback_parts.append("❌ No conditional formatting or red highlighting found")

        # Criterion 10: Check if Pepper's cell (should have value 1) is highlighted
        pepper_highlighted = False
        if total_row_idx is not None:
            total_row_num = total_row_idx + 1
            # Pepper is typically in column D
            pepper_cell_ref = f'D{total_row_num}'
            pepper_value = get_cell_value(wb_data, ws_data.title, pepper_cell_ref)

            if pepper_value is not None and abs(pepper_value - 1) <= 1:
                # Check if this cell is highlighted red
                if check_cell_has_red_fill(ws_formulas, pepper_cell_ref):
                    criteria_passed += 1
                    pepper_highlighted = True
                    feedback_parts.append(f"✅ Problem hen (Pepper: {pepper_value} eggs) is highlighted")
                else:
                    # Even without red fill, if conditional formatting exists and value is correct, give partial credit
                    if has_conditional_formatting:
                        criteria_passed += 1
                        pepper_highlighted = True
                        feedback_parts.append(f"✅ Problem hen identified (Pepper: {pepper_value} eggs, conditional format set)")
                    else:
                        feedback_parts.append(f"⚠️ Pepper has low production ({pepper_value} eggs) but not highlighted")
            else:
                feedback_parts.append(f"❌ Pepper's weekly total incorrect or missing: {pepper_value}")
        else:
            feedback_parts.append("❌ Cannot verify Pepper highlighting without total row")

        # Calculate final score
        score = int((criteria_passed / 10) * 100)
        passed = score >= 80

        feedback = " | ".join(feedback_parts)

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0,
            "feedback": f"Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)