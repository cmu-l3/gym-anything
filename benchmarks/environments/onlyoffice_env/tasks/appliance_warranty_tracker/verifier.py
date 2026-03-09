#!/usr/bin/env python3
"""
Verifier for Appliance Warranty Tracker task

This verifier checks:
1. Spreadsheet structure (headers, columns)
2. Data completeness (at least 3 appliances)
3. Formula presence and correctness in 'Warranty Expires' column
4. Conditional formula logic in 'Status' column
5. Accuracy of calculated values
"""

import sys
import os
import logging
import tempfile
import re
from datetime import datetime, timedelta

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    parse_xlsx_file,
    get_cell_value,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def find_header_row(sheet, max_rows=5):
    """
    Find the row containing headers by looking for key terms.
    Returns (row_number, column_mapping) or (None, None)
    
    column_mapping is a dict like:
    {'name': 1, 'model': 2, 'purchase': 3, 'warranty': 4, 'expires': 5, 'status': 6}
    """
    keywords = {
        'name': ['appliance', 'name', 'item', 'device', 'equipment'],
        'model': ['brand', 'model', 'manufacturer', 'make'],
        'purchase': ['purchase', 'bought', 'date', 'acquired', 'purchased'],
        'warranty': ['warranty', 'year', 'period', 'duration', 'term'],
        'expires': ['expire', 'expiration', 'end', 'expiry'],
        'status': ['status', 'state', 'condition', 'active']
    }
    
    for row_idx in range(1, max_rows + 1):
        row_values = []
        for col_idx in range(1, 12):  # Check first 11 columns
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = str(cell.value or "").lower().strip()
            row_values.append((col_idx, value))
        
        # Check if this row contains header-like terms
        matches = {}
        for col_idx, value in row_values:
            if len(value) == 0 or value.startswith('('):
                continue
                
            for key, terms in keywords.items():
                if any(term in value for term in terms):
                    if key not in matches:  # First match wins
                        matches[key] = col_idx
        
        # If we found most of the required columns, this is likely the header row
        if len(matches) >= 4:  # Need at least 4 of the 6 key columns
            logger.info(f"Found header row {row_idx} with columns: {matches}")
            return row_idx, matches
    
    return None, None


def verify_warranty_tracker(traj, env_info, task_info):
    """
    Verify that the appliance warranty tracker was created correctly.
    
    Checks:
    1. File exists and is readable
    2. Headers are present with appropriate columns  
    3. At least 3 appliance entries exist
    4. Warranty Expires column has formulas that calculate correctly
    5. Status column has conditional formulas with correct logic
    6. Data types are appropriate (dates, numbers)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/appliance_warranty_tracker.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_warranty_')

    try:
        # Copy file from container
        temp_file = os.path.join(temp_dir, "appliance_warranty_tracker.xlsx")
        
        try:
            copy_from_env(container_path, temp_file)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to copy file: {str(e)}"}
        
        if not os.path.exists(temp_file) or os.path.getsize(temp_file) == 0:
            return {"passed": False, "score": 0, "feedback": f"Spreadsheet not found or empty at {container_path}"}

        # Parse the spreadsheet (with data_only=True to see calculated values)
        wb_values = parse_xlsx_file(temp_file)
        if wb_values is None:
            return {"passed": False, "score": 0, "feedback": "Failed to parse spreadsheet"}

        # Also load with data_only=False to see formulas
        try:
            from openpyxl import load_workbook
            wb_formulas = load_workbook(temp_file, data_only=False)
        except Exception as e:
            logger.warning(f"Could not load formulas: {e}")
            wb_formulas = None

        sheet_values = wb_values.active
        sheet_formulas = wb_formulas.active if wb_formulas else None

        criteria_passed = 0
        total_criteria = 8
        feedback_parts = []

        # Criterion 1: Find header row and column mapping
        header_row, col_map = find_header_row(sheet_values)
        
        if header_row is None:
            feedback_parts.append("❌ Could not find header row with required columns (need: name, purchase date, warranty, expires, status)")
            return {
                "passed": False,
                "score": 0,
                "feedback": " | ".join(feedback_parts)
            }
        
        criteria_passed += 1
        feedback_parts.append(f"✅ Headers found at row {header_row}")
        
        # Check we have the essential columns
        required_cols = ['name', 'purchase', 'warranty', 'expires', 'status']
        missing_cols = [col for col in required_cols if col not in col_map]
        
        if len(missing_cols) > 0:
            feedback_parts.append(f"❌ Missing required columns: {', '.join(missing_cols)}")
        else:
            criteria_passed += 1
            feedback_parts.append("✅ All required columns present")

        # Criterion 2: Check for data rows (at least 3 appliances)
        data_row_start = header_row + 1
        data_rows = []
        
        for row_idx in range(data_row_start, data_row_start + 15):  # Check up to 15 rows
            name_col = col_map.get('name', 1)
            cell_value = sheet_values.cell(row=row_idx, column=name_col).value
            
            # Valid data row: has content, not empty, not an instruction
            if cell_value and str(cell_value).strip():
                text = str(cell_value).strip().lower()
                # Skip instruction rows
                if not (text.startswith('(') or text.startswith('create') or text.startswith('suggested')):
                    data_rows.append(row_idx)
        
        logger.info(f"Found {len(data_rows)} data rows: {data_rows}")
        
        if len(data_rows) >= 3:
            criteria_passed += 1
            feedback_parts.append(f"✅ {len(data_rows)} appliance entries found")
        else:
            feedback_parts.append(f"❌ Only {len(data_rows)} appliance entries found (need at least 3)")

        if len(data_rows) == 0:
            feedback_parts.append("❌ No data rows found - cannot verify formulas")
            return {
                "passed": False,
                "score": int((criteria_passed / total_criteria) * 100),
                "feedback": " | ".join(feedback_parts)
            }

        # Criterion 3: Check purchase dates are actual dates
        purchase_col = col_map.get('purchase')
        warranty_col = col_map.get('warranty')
        expires_col = col_map.get('expires')
        status_col = col_map.get('status')
        
        date_count = 0
        if purchase_col:
            for row_idx in data_rows[:5]:  # Check first 5 data rows
                cell_value = sheet_values.cell(row=row_idx, column=purchase_col).value
                if isinstance(cell_value, datetime):
                    date_count += 1
        
        checked_rows = min(len(data_rows), 5)
        if date_count >= 2:
            criteria_passed += 1
            feedback_parts.append(f"✅ Purchase dates in date format ({date_count}/{checked_rows})")
        else:
            feedback_parts.append(f"❌ Purchase dates not in proper date format ({date_count}/{checked_rows})")

        # Criterion 4: Check warranty periods are numbers
        number_count = 0
        if warranty_col:
            for row_idx in data_rows[:5]:
                cell_value = sheet_values.cell(row=row_idx, column=warranty_col).value
                if isinstance(cell_value, (int, float)) and 0 < cell_value <= 10:
                    number_count += 1
        
        if number_count >= 2:
            criteria_passed += 1
            feedback_parts.append(f"✅ Warranty periods are numeric ({number_count}/{checked_rows})")
        else:
            feedback_parts.append(f"❌ Warranty periods not numeric ({number_count}/{checked_rows})")

        # Criterion 5: Check Warranty Expires column has formulas
        formula_count = 0
        correct_calculation_count = 0
        
        if expires_col and sheet_formulas and purchase_col and warranty_col:
            for row_idx in data_rows[:5]:
                formula_cell = sheet_formulas.cell(row=row_idx, column=expires_col)
                formula = str(formula_cell.value or "")
                
                if formula.startswith('='):
                    formula_count += 1
                    logger.info(f"Row {row_idx} expiration formula: {formula}")
                    
                    # Verify calculated value is correct
                    purchase_date = sheet_values.cell(row=row_idx, column=purchase_col).value
                    warranty_years = sheet_values.cell(row=row_idx, column=warranty_col).value
                    expires_date = sheet_values.cell(row=row_idx, column=expires_col).value
                    
                    if isinstance(purchase_date, datetime) and isinstance(warranty_years, (int, float)) and isinstance(expires_date, datetime):
                        # Calculate expected expiration (allow for leap year differences)
                        expected_expires = purchase_date + timedelta(days=int(warranty_years * 365.25))
                        days_diff = abs((expires_date - expected_expires).days)
                        
                        logger.info(f"Row {row_idx}: purchase={purchase_date.date()}, warranty={warranty_years}y, expires={expires_date.date()}, expected={expected_expires.date()}, diff={days_diff}d")
                        
                        if days_diff <= 5:  # Allow 5-day tolerance for different calculation methods
                            correct_calculation_count += 1
        
        if formula_count >= 2:
            criteria_passed += 1
            feedback_parts.append(f"✅ Warranty Expires formulas found ({formula_count}/{checked_rows})")
        else:
            feedback_parts.append(f"❌ Missing formulas in Warranty Expires column ({formula_count}/{checked_rows})")

        if correct_calculation_count >= 2:
            criteria_passed += 1
            feedback_parts.append(f"✅ Expiration dates calculated correctly ({correct_calculation_count}/{checked_rows})")
        else:
            feedback_parts.append(f"⚠️ Expiration calculations may be incorrect ({correct_calculation_count}/{checked_rows})")

        # Criterion 6: Check Status column has conditional formulas
        status_formula_count = 0
        correct_status_count = 0
        
        if status_col and expires_col and sheet_formulas:
            today = datetime.now()
            
            for row_idx in data_rows[:5]:
                formula_cell = sheet_formulas.cell(row=row_idx, column=status_col)
                formula = str(formula_cell.value or "")
                
                if formula.startswith('='):
                    formula_upper = formula.upper()
                    
                    # Check if formula uses IF and TODAY
                    if 'IF' in formula_upper and 'TODAY' in formula_upper:
                        status_formula_count += 1
                        logger.info(f"Row {row_idx} status formula: {formula}")
                        
                        # Verify the status value is correct
                        expires_date = sheet_values.cell(row=row_idx, column=expires_col).value
                        status_value = str(sheet_values.cell(row=row_idx, column=status_col).value or "").upper()
                        
                        if isinstance(expires_date, datetime):
                            days_until_expiration = (expires_date - today).days
                            
                            # Determine expected status (allow some flexibility in threshold)
                            if days_until_expiration < 0:
                                expected_keywords = ["EXPIRED", "PAST"]
                            elif days_until_expiration <= 95:  # Allow some threshold variation
                                expected_keywords = ["EXPIRING", "SOON", "WARNING"]
                            else:
                                expected_keywords = ["ACTIVE", "VALID", "OK", "GOOD"]
                            
                            logger.info(f"Row {row_idx}: expires in {days_until_expiration}d, status='{status_value}', expected={expected_keywords}")
                            
                            if any(keyword in status_value for keyword in expected_keywords):
                                correct_status_count += 1
        
        if status_formula_count >= 2:
            criteria_passed += 1
            feedback_parts.append(f"✅ Status conditional formulas found ({status_formula_count}/{checked_rows})")
        else:
            feedback_parts.append(f"❌ Missing conditional formulas in Status column ({status_formula_count}/{checked_rows})")

        if correct_status_count >= 2:
            criteria_passed += 1
            feedback_parts.append(f"✅ Status values are correct ({correct_status_count}/{checked_rows})")
        else:
            feedback_parts.append(f"⚠️ Status values may be incorrect ({correct_status_count}/{checked_rows})")

        # Calculate final score
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 70

        feedback = " | ".join(feedback_parts)
        
        logger.info(f"Verification complete: {criteria_passed}/{total_criteria} criteria passed, score={score}%")

        return {
            "passed": passed,
            "score": score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_temp_dir(temp_dir)