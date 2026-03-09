#!/usr/bin/env python3
"""
Verifier for Credit Report Dispute Tracker task

This verifier checks that the agent has:
1. Created a structured tracking system in the "Dispute Tracker" sheet
2. Added appropriate column headers (flexible matching)
3. Transformed the 4 errors from messy notes into organized rows
4. Included all three credit bureaus (Experian, TransUnion, Equifax)
5. Added realistic USPS tracking numbers
6. Used formulas for deadline calculations
7. Included status tracking

Edge cases handled:
- Column headers may be worded differently
- "All 3 bureaus" entry might be split into separate rows
- Tracking numbers may have varied spacing/formatting
- Formulas might be in different columns than expected
- Different organizational approaches are acceptable
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    get_cell_value,
    get_sheet_data,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_credit_dispute_tracker(traj, env_info, task_info):
    """
    Verify that the credit dispute tracking spreadsheet was created correctly.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/credit_disputes_template.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_disputes_')

    try:
        # Copy and parse the spreadsheet
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"❌ Failed to load spreadsheet: {error}"
            }

        # Check if "Dispute Tracker" sheet exists
        if "Dispute Tracker" not in wb.sheetnames:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": "❌ Sheet 'Dispute Tracker' not found. Please create this sheet."
            }

        sheet = wb["Dispute Tracker"]
        
        # Get all data from the sheet
        data = get_sheet_data(wb, "Dispute Tracker", max_rows=50, max_cols=15)
        
        if not data or len(data) < 2:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": "❌ Dispute Tracker sheet is empty or has insufficient data. Please build the tracking system."
            }

        # Initialize scoring
        score = 0.0
        max_score = 100.0
        feedback_parts = []

        # ====================
        # CRITERION 1: Column Headers (15 points)
        # ====================
        headers = [str(cell).lower() if cell else "" for cell in data[0]]
        headers_text = " ".join(headers)
        
        non_empty_headers = [h for h in headers if h and h != "none" and len(h) > 1]
        
        # Required column concepts with flexible keyword matching
        required_concepts = [
            ("bureau", ["bureau", "credit bureau", "agency", "company"]),
            ("error/description", ["error", "account", "description", "issue", "dispute", "problem"]),
            ("tracking", ["tracking", "certified", "receipt", "mail", "usps", "confirmation"]),
            ("deadline", ["deadline", "due", "response", "30 day", "30-day", "30day"]),
            ("status", ["status", "outcome", "result", "state", "progress"]),
        ]
        
        found_concepts = []
        for concept_name, keywords in required_concepts:
            found = any(keyword in headers_text for keyword in keywords)
            if found:
                found_concepts.append(concept_name)
        
        if len(non_empty_headers) >= 8 and len(found_concepts) >= 4:
            score += 15.0
            feedback_parts.append(f"✅ Well-structured headers ({len(non_empty_headers)} columns, {len(found_concepts)}/5 key concepts)")
        elif len(non_empty_headers) >= 6 and len(found_concepts) >= 3:
            score += 10.0
            feedback_parts.append(f"✅ Good header structure ({len(non_empty_headers)} columns, {len(found_concepts)}/5 concepts)")
        elif len(non_empty_headers) >= 4:
            score += 5.0
            feedback_parts.append(f"⚠️ Basic headers present but incomplete ({len(non_empty_headers)} columns)")
        else:
            feedback_parts.append(f"❌ Insufficient column headers ({len(non_empty_headers)} columns)")

        # ====================
        # CRITERION 2: Number of Dispute Entries (15 points)
        # ====================
        data_rows = []
        for row in data[1:]:
            has_content = any(cell for cell in row if cell is not None and str(cell).strip() and str(cell).lower() != "none")
            if has_content:
                data_rows.append(row)
        
        num_entries = len(data_rows)
        
        if num_entries >= 4:
            score += 15.0
            feedback_parts.append(f"✅ Contains {num_entries} dispute entries (required: 4+)")
        elif num_entries == 3:
            score += 10.0
            feedback_parts.append(f"⚠️ Contains 3 dispute entries (expected 4+)")
        elif num_entries >= 1:
            score += 5.0
            feedback_parts.append(f"⚠️ Only {num_entries} dispute entry/entries (need 4+)")
        else:
            feedback_parts.append(f"❌ No dispute entries found")

        # ====================
        # CRITERION 3: All Three Credit Bureaus (15 points)
        # ====================
        all_text = " ".join([
            str(cell).lower() 
            for row in data_rows 
            for cell in row 
            if cell is not None
        ])
        
        bureaus_found = []
        if "experian" in all_text:
            bureaus_found.append("Experian")
        if "transunion" in all_text or "trans union" in all_text:
            bureaus_found.append("TransUnion")
        if "equifax" in all_text:
            bureaus_found.append("Equifax")
        
        if len(bureaus_found) >= 3:
            score += 15.0
            feedback_parts.append(f"✅ All three credit bureaus: {', '.join(bureaus_found)}")
        elif len(bureaus_found) == 2:
            score += 10.0
            feedback_parts.append(f"⚠️ Only two bureaus: {', '.join(bureaus_found)} (need all three)")
        elif len(bureaus_found) == 1:
            score += 5.0
            feedback_parts.append(f"⚠️ Only one bureau: {bureaus_found[0]}")
        else:
            feedback_parts.append(f"❌ No credit bureau names found")

        # ====================
        # CRITERION 4: Specific Errors from Scenario (20 points)
        # ====================
        required_errors = [
            (["capital one", "5412"], "Capital One fraudulent card"),
            (["navient", "student"], "Navient student loan"),
            (["midwest medical", "midwest", "collection", "890"], "Midwest Medical collection"),
            (["address", "456 oak", "wrong address", "oak st"], "Wrong address issue")
        ]
        
        errors_found = 0
        errors_details = []
        
        for keywords, error_name in required_errors:
            if any(keyword in all_text for keyword in keywords):
                errors_found += 1
                errors_details.append(error_name)
        
        if errors_found >= 4:
            score += 20.0
            feedback_parts.append(f"✅ All 4 specific errors from scenario included")
        elif errors_found == 3:
            score += 15.0
            feedback_parts.append(f"✅ Found 3/4 specific errors")
        elif errors_found == 2:
            score += 10.0
            feedback_parts.append(f"⚠️ Only 2/4 specific errors found")
        elif errors_found == 1:
            score += 5.0
            feedback_parts.append(f"⚠️ Only 1/4 specific errors found")
        else:
            feedback_parts.append(f"❌ Specific errors from scenario not found")

        # ====================
        # CRITERION 5: Tracking Numbers (10 points)
        # ====================
        tracking_pattern = re.compile(r'\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{2,4}')
        
        tracking_numbers_found = []
        for row in data_rows:
            for cell in row:
                if cell and tracking_pattern.search(str(cell)):
                    tracking_numbers_found.append(str(cell))
        
        if len(tracking_numbers_found) >= 3:
            score += 10.0
            feedback_parts.append(f"✅ Found {len(tracking_numbers_found)} properly formatted tracking numbers")
        elif len(tracking_numbers_found) >= 1:
            score += 5.0
            feedback_parts.append(f"⚠️ Found {len(tracking_numbers_found)} tracking number(s), expected more")
        else:
            feedback_parts.append(f"⚠️ No USPS tracking numbers found")

        # ====================
        # CRITERION 6: Formulas for Deadline Calculations (15 points)
        # ====================
        has_formulas = False
        formula_count = 0
        deadline_formula_count = 0
        
        try:
            from openpyxl import load_workbook
            
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=temp_dir)
            copy_from_env(container_path, temp_file.name)
            
            wb_formula = load_workbook(temp_file.name, data_only=False)
            sheet_formula = wb_formula["Dispute Tracker"]
            
            for row in sheet_formula.iter_rows(min_row=2, max_row=20, max_col=15):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        has_formulas = True
                        formula_count += 1
                        # Check for date arithmetic (deadline calculations)
                        if '+30' in cell.value or '+ 30' in cell.value or '+' in cell.value:
                            deadline_formula_count += 1
            
            os.unlink(temp_file.name)
            
            if deadline_formula_count >= 3:
                score += 15.0
                feedback_parts.append(f"✅ Contains deadline calculation formulas ({deadline_formula_count} found)")
            elif deadline_formula_count >= 1:
                score += 10.0
                feedback_parts.append(f"✅ Some deadline formulas present ({deadline_formula_count} found)")
            elif formula_count >= 1:
                score += 5.0
                feedback_parts.append(f"⚠️ Found formulas but may not be for deadlines")
            else:
                feedback_parts.append(f"⚠️ No formulas detected (use =DateSent+30 for deadlines)")
                
        except Exception as e:
            logger.warning(f"Could not check for formulas: {e}")
            score += 5.0
            feedback_parts.append("⚠️ Could not verify formulas")

        # ====================
        # CRITERION 7: Status/Progress Tracking (10 points)
        # ====================
        status_keywords = [
            "pending", "resolved", "investigating", "follow", "sent", 
            "received", "deleted", "disputed", "waiting", "responded",
            "need to send", "mailed", "in progress", "complete"
        ]
        
        status_count = sum(1 for keyword in status_keywords if keyword in all_text)
        
        if status_count >= 2:
            score += 10.0
            feedback_parts.append(f"✅ Status tracking included ({status_count} indicators)")
        elif status_count >= 1:
            score += 5.0
            feedback_parts.append(f"⚠️ Some status tracking present")
        else:
            feedback_parts.append(f"⚠️ No status indicators found")

        # ====================
        # BONUS: Date Formatting
        # ====================
        date_pattern = re.compile(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}')
        dates_found = sum(1 for row in data_rows for cell in row 
                         if cell and date_pattern.search(str(cell)))
        
        if dates_found >= 3 and score < 100.0:
            bonus = min(5.0, 100.0 - score)
            score += bonus
            if bonus > 0:
                feedback_parts.append(f"✅ Dates properly formatted")

        # ====================
        # FINAL SCORING
        # ====================
        score = min(score, max_score)
        passed = score >= 70.0
        
        feedback = " | ".join(feedback_parts)
        
        return {
            "passed": passed,
            "score": round(score, 1),
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"❌ Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)