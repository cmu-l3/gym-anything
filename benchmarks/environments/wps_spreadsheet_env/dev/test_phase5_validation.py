#!/usr/bin/env python3
"""
Phase 5 Validation: Wrong-target and Partial-completion tests for all 5 very_hard tasks.

Tests:
1. Wrong-target: Provide a workbook with irrelevant sheet names → expect score=0
2. Partial-completion: Provide a workbook with only 1-2 of the required sheets → expect 0 < score < pass_threshold
"""

import sys
import os
import tempfile
import shutil
import json

# Add utils and task directories to path
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(BASE, 'utils'))
sys.path.insert(0, os.path.join(BASE, 'tasks', 'financial_consolidation_analysis'))
sys.path.insert(0, os.path.join(BASE, 'tasks', 'production_capacity_planning'))
sys.path.insert(0, os.path.join(BASE, 'tasks', 'compensation_equity_analysis'))
sys.path.insert(0, os.path.join(BASE, 'tasks', 'budget_variance_dashboard'))
sys.path.insert(0, os.path.join(BASE, 'tasks', 'loan_portfolio_amortization'))

from openpyxl import Workbook

def import_verifier(task_name, func_name):
    """Import a verifier function from a task directory."""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        f"verifier_{task_name}",
        os.path.join(BASE, 'tasks', task_name, 'verifier.py')
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return getattr(mod, func_name)


def make_copy_from_env(xlsx_path):
    """Create a mock copy_from_env that copies a local XLSX file."""
    def copy_from_env(src, dst):
        if src.endswith('.xlsx'):
            shutil.copy2(xlsx_path, dst)
        else:
            raise FileNotFoundError(f"Mock: only serves xlsx, not {src}")
    return copy_from_env


def create_wrong_target_wb(path):
    """Create a workbook with completely irrelevant content."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RandomData"
    ws['A1'] = "Name"
    ws['B1'] = "Score"
    ws['A2'] = "Alice"
    ws['B2'] = 95
    ws['A3'] = "Bob"
    ws['B3'] = 87
    wb.save(path)


def create_partial_financial(path):
    """Create a workbook with starter sheets + only a Consolidated sheet (partial)."""
    wb = Workbook()
    # Starter sheets
    for name in ["Alpha_Inc", "Beta_Corp", "Gamma_LLC", "Intercompany", "Prior_Year"]:
        ws = wb.create_sheet(name)
        ws['A1'] = "Account"
        ws['B1'] = "Amount"
        ws['A2'] = "Revenue"
        ws['B2'] = 49200000 if name == "Alpha_Inc" else 34300000

    # Only add Consolidated (1 of 4 required new sheets)
    consol = wb.create_sheet("Consolidated")
    consol['A1'] = "Account"
    consol['B1'] = "Amount"
    consol['A2'] = "Revenue"
    consol['B2'] = "=Alpha_Inc!B2+Beta_Corp!B2+Gamma_LLC!B2"
    consol['A3'] = "Net Income"
    consol['B3'] = 8493000

    # Remove default sheet
    del wb['Sheet']
    wb.save(path)


def create_partial_production(path):
    """Create workbook with starter sheets + only a Schedule sheet."""
    wb = Workbook()
    for name in ["Production_Lines", "Orders", "Calendar"]:
        ws = wb.create_sheet(name)
        ws['A1'] = "ID"
        ws['B1'] = "Name"

    sched = wb.create_sheet("Schedule")
    sched['A1'] = "Order"
    sched['B1'] = "Line"
    sched['C1'] = "Start"
    sched['D1'] = "End"
    sched['E1'] = "Slack"
    for i in range(15):
        sched[f'A{i+2}'] = f"ORD-{i+1:03d}"
        sched[f'B{i+2}'] = f"L{(i % 4) + 1}"

    del wb['Sheet']
    wb.save(path)


def create_partial_compensation(path):
    """Create workbook with starter sheets + only a Compa_Ratio sheet."""
    wb = Workbook()
    for name in ["Employees", "Market_Benchmarks"]:
        ws = wb.create_sheet(name)
        ws['A1'] = "ID"
        ws['B1'] = "Name"

    compa = wb.create_sheet("Compa_Ratio")
    compa['A1'] = "Employee"
    compa['B1'] = "Salary"
    compa['C1'] = "Market"
    compa['D1'] = "Compa Ratio"
    for i in range(36):
        compa[f'A{i+2}'] = f"E{i+1:04d}"
        compa[f'B{i+2}'] = 80000 + i * 1000
        compa[f'C{i+2}'] = f"=VLOOKUP(A{i+2},Employees!A:B,2,FALSE)"
        compa[f'D{i+2}'] = 0.85 + (i * 0.01)

    del wb['Sheet']
    wb.save(path)


def create_partial_budget(path):
    """Create workbook with starter sheets + only a Monthly_Variance sheet."""
    wb = Workbook()
    for name in ["Budget", "Actuals"]:
        ws = wb.create_sheet(name)
        ws['A1'] = "Cost Center"
        ws['B1'] = "Jan"

    var = wb.create_sheet("Monthly_Variance")
    var['A1'] = "Cost Center"
    var['B1'] = "$ Variance"
    var['C1'] = "% Variance"
    var['C1'].number_format = '0%'
    for i in range(5):
        var[f'A{i+2}'] = f"CC-{i+1:03d}"
        var[f'B{i+2}'] = f"=Budget!B{i+2}-Actuals!B{i+2}"
        var[f'C{i+2}'] = 0.05

    del wb['Sheet']
    wb.save(path)


def create_partial_loan(path):
    """Create workbook with starter sheets + only amortization schedules (no summary/covenant)."""
    wb = Workbook()
    for name in ["Loan_Terms", "Rate_Curve", "Property_NOI"]:
        ws = wb.create_sheet(name)
        ws['A1'] = "ID"

    # Add 4 amortization sheets
    for i in range(4):
        lid = f"LOAN-{i+1:03d}"
        ws = wb.create_sheet(f"Amort_{lid}")
        ws['A1'] = "Period"
        ws['B1'] = "Payment"
        ws['C1'] = "Interest"
        ws['D1'] = "Principal"
        ws['E1'] = "Balance"
        ws['B2'] = f"=PMT(0.05/12, 360, -1000000)"
        ws['C2'] = f"=IPMT(0.05/12, 1, 360, -1000000)"
        ws['D2'] = f"=PPMT(0.05/12, 1, 360, -1000000)"

    del wb['Sheet']
    wb.save(path)


def run_test(test_name, verifier_fn, xlsx_path, task_info, expect_pass, expect_score_range):
    """Run a single test case."""
    env_info = {
        'copy_from_env': make_copy_from_env(xlsx_path),
    }
    result = verifier_fn([], env_info, task_info)
    score = result.get('score', 0)
    passed = result.get('passed', False)
    feedback = result.get('feedback', '')

    lo, hi = expect_score_range
    score_ok = lo <= score <= hi
    pass_ok = passed == expect_pass

    status = "PASS" if (score_ok and pass_ok) else "FAIL"
    print(f"  [{status}] {test_name}: score={score} (expected {lo}-{hi}), passed={passed} (expected {expect_pass})")
    if not score_ok or not pass_ok:
        print(f"         Feedback: {feedback}")
    return status == "PASS"


def main():
    tmp_dir = tempfile.mkdtemp(prefix='phase5_test_')
    all_passed = True

    try:
        # ====================================================================
        # 1. FINANCIAL CONSOLIDATION
        # ====================================================================
        print("\n=== financial_consolidation_analysis ===")
        verify_fn = import_verifier('financial_consolidation_analysis', 'verify_financial_consolidation')
        task_info = {"metadata": {"ground_truth": {"consolidated_revenue": 96500000, "consolidated_net_income": 8493000}}}

        # Wrong-target
        wrong_path = os.path.join(tmp_dir, 'wrong_financial.xlsx')
        create_wrong_target_wb(wrong_path)
        ok = run_test("Wrong-target", verify_fn, wrong_path, task_info,
                       expect_pass=False, expect_score_range=(0, 0))
        all_passed &= ok

        # Partial (only Consolidated sheet, missing Ratios/Variance/Dashboard)
        partial_path = os.path.join(tmp_dir, 'partial_financial.xlsx')
        create_partial_financial(partial_path)
        ok = run_test("Partial (Consolidated only)", verify_fn, partial_path, task_info,
                       expect_pass=False, expect_score_range=(1, 54))
        all_passed &= ok

        # ====================================================================
        # 2. PRODUCTION CAPACITY PLANNING
        # ====================================================================
        print("\n=== production_capacity_planning ===")
        verify_fn = import_verifier('production_capacity_planning', 'verify_production_capacity')
        task_info = {"metadata": {"ground_truth": {"total_orders": 15, "num_lines": 4}}}

        wrong_path = os.path.join(tmp_dir, 'wrong_production.xlsx')
        create_wrong_target_wb(wrong_path)
        ok = run_test("Wrong-target", verify_fn, wrong_path, task_info,
                       expect_pass=False, expect_score_range=(0, 0))
        all_passed &= ok

        partial_path = os.path.join(tmp_dir, 'partial_production.xlsx')
        create_partial_production(partial_path)
        ok = run_test("Partial (Schedule only)", verify_fn, partial_path, task_info,
                       expect_pass=False, expect_score_range=(1, 49))
        all_passed &= ok

        # ====================================================================
        # 3. COMPENSATION EQUITY ANALYSIS
        # ====================================================================
        print("\n=== compensation_equity_analysis ===")
        verify_fn = import_verifier('compensation_equity_analysis', 'verify_compensation_equity')
        task_info = {"metadata": {}}

        wrong_path = os.path.join(tmp_dir, 'wrong_compensation.xlsx')
        create_wrong_target_wb(wrong_path)
        ok = run_test("Wrong-target", verify_fn, wrong_path, task_info,
                       expect_pass=False, expect_score_range=(0, 0))
        all_passed &= ok

        partial_path = os.path.join(tmp_dir, 'partial_compensation.xlsx')
        create_partial_compensation(partial_path)
        ok = run_test("Partial (Compa_Ratio only)", verify_fn, partial_path, task_info,
                       expect_pass=False, expect_score_range=(1, 49))
        all_passed &= ok

        # ====================================================================
        # 4. BUDGET VARIANCE DASHBOARD
        # ====================================================================
        print("\n=== budget_variance_dashboard ===")
        verify_fn = import_verifier('budget_variance_dashboard', 'verify_budget_variance')
        task_info = {"metadata": {}}

        wrong_path = os.path.join(tmp_dir, 'wrong_budget.xlsx')
        create_wrong_target_wb(wrong_path)
        ok = run_test("Wrong-target", verify_fn, wrong_path, task_info,
                       expect_pass=False, expect_score_range=(0, 0))
        all_passed &= ok

        partial_path = os.path.join(tmp_dir, 'partial_budget.xlsx')
        create_partial_budget(partial_path)
        ok = run_test("Partial (Monthly_Variance only)", verify_fn, partial_path, task_info,
                       expect_pass=False, expect_score_range=(1, 49))
        all_passed &= ok

        # ====================================================================
        # 5. LOAN PORTFOLIO AMORTIZATION
        # ====================================================================
        print("\n=== loan_portfolio_amortization ===")
        verify_fn = import_verifier('loan_portfolio_amortization', 'verify_loan_portfolio')
        task_info = {"metadata": {}}

        wrong_path = os.path.join(tmp_dir, 'wrong_loan.xlsx')
        create_wrong_target_wb(wrong_path)
        ok = run_test("Wrong-target", verify_fn, wrong_path, task_info,
                       expect_pass=False, expect_score_range=(0, 0))
        all_passed &= ok

        partial_path = os.path.join(tmp_dir, 'partial_loan.xlsx')
        create_partial_loan(partial_path)
        ok = run_test("Partial (Amort schedules only)", verify_fn, partial_path, task_info,
                       expect_pass=False, expect_score_range=(1, 49))
        all_passed &= ok

        # ====================================================================
        # SUMMARY
        # ====================================================================
        print("\n" + "=" * 60)
        if all_passed:
            print("ALL PHASE 5 TESTS PASSED")
        else:
            print("SOME TESTS FAILED — see details above")
        print("=" * 60)

    finally:
        shutil.rmtree(tmp_dir)

    return 0 if all_passed else 1


if __name__ == '__main__':
    sys.exit(main())
